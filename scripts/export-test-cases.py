"""Tổng hợp toàn bộ test case hiện có trong repo và xuất ra Excel theo định dạng Given/When/Then.

Nguồn dữ liệu: source code test tại commit HEAD của branch master. Script KHÔNG chạy test —
mọi test hiện có trong repo được coi là đã pass ở version hiện tại, nhiệm vụ của script chỉ là
tập hợp và phân loại.

Mỗi phần tử trong ROWS tương ứng với đúng một test method [Fact]/[Theory] (backend, xUnit) hoặc
một it()/test() (frontend, vitest/Playwright) có thật trong code. Một [Theory] nhiều InlineData
được gộp thành một dòng, các biến thể mô tả trong phần Given.

Chạy:  python scripts/export-test-cases.py
Ra:    docs/test-cases-<YYYY-MM-DD>.xlsx
"""

from __future__ import annotations

import subprocess
import sys
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROWS: list[dict[str, str]] = []

# ==========================================================================================
# BASKET — services/baskets/** (Backend) + frontend/apps/web/tests/basket/** (Frontend)
# ==========================================================================================

ROWS += [
    # ---- Backend / Unit / BasketLineMergeTests -------------------------------------------
    {
        "feature": "Basket",
        "sub": "Basket Domain - Gộp dòng sản phẩm",
        "layer": "Backend",
        "type": "Unit",
        "name": "AddItem_CreatesALine_WhenTheProductIsNotInTheBasketYet",
        "given": "Một giỏ hàng mới của khách phase1-stub-user, chưa có dòng sản phẩm nào.",
        "when": "Thêm sản phẩm Notebook với số lượng 1, đơn giá 12.50.",
        "then": "Giỏ có đúng 1 dòng, đúng mã sản phẩm Notebook, số lượng 1 và đơn giá 12.50.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketLineMergeTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Gộp dòng sản phẩm",
        "layer": "Backend",
        "type": "Unit",
        "name": "AddItem_IncrementsTheExistingLine_WhenTheProductIsAlreadyInTheBasket",
        "given": "Giỏ hàng đã có sẵn 1 dòng sản phẩm Notebook số lượng 1.",
        "when": "Thêm tiếp chính sản phẩm Notebook đó với số lượng 1.",
        "then": "Giỏ vẫn chỉ có 1 dòng, số lượng được cộng dồn lên 2 thay vì tạo dòng thứ hai.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketLineMergeTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Gộp dòng sản phẩm",
        "layer": "Backend",
        "type": "Unit",
        "name": "AddItem_KeepsProductsApart_WhenDifferentProductsAreAdded",
        "given": "Một giỏ hàng mới, chưa có dòng nào.",
        "when": "Thêm Notebook số lượng 1 và Apron số lượng 2.",
        "then": "Giỏ có 2 dòng tách biệt: dòng Notebook số lượng 1 và dòng Apron số lượng 2.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketLineMergeTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Gộp dòng sản phẩm",
        "layer": "Backend",
        "type": "Unit",
        "name": "AddItem_AccumulatesQuantities_AcrossManyAdditions",
        "given": "Một giỏ hàng mới, chưa có dòng nào.",
        "when": "Thêm cùng sản phẩm Notebook 5 lần liên tiếp, mỗi lần số lượng 1.",
        "then": "Giỏ chỉ có 1 dòng với số lượng đúng bằng 5, tức bằng số lần đã thêm (tiêu chí SC-003).",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketLineMergeTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Gộp dòng sản phẩm",
        "layer": "Backend",
        "type": "Unit",
        "name": "AddItem_KeepsTheOriginallyCapturedPrice_WhenTheCatalogPriceHasChanged",
        "given": "Giỏ hàng đã có Notebook được thêm với đơn giá 12.50.",
        "when": "Thêm lại chính sản phẩm đó nhưng với đơn giá mới 99.99.",
        "then": "Dòng giữ nguyên đơn giá chốt lúc đầu là 12.50 và số lượng lên 2 - giỏ hàng không bị định giá lại.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketLineMergeTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Gộp dòng sản phẩm",
        "layer": "Backend",
        "type": "Unit",
        "name": "AddItem_Rejects_AQuantityBelowOne",
        "given": "Một giỏ hàng mới; kiểm thử với các số lượng không hợp lệ là 0 và -1 (Theory, 2 bộ dữ liệu).",
        "when": "Thêm sản phẩm với số lượng nhỏ hơn 1.",
        "then": "Ném ArgumentOutOfRangeException; không cho phép tồn tại dòng có số lượng bằng 0 hoặc âm.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketLineMergeTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Gộp dòng sản phẩm",
        "layer": "Backend",
        "type": "Unit",
        "name": "AddItem_Rejects_ANegativeUnitPrice",
        "given": "Một giỏ hàng mới.",
        "when": "Thêm sản phẩm với đơn giá âm (-0.01).",
        "then": "Ném ArgumentOutOfRangeException; đơn giá âm bị từ chối.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketLineMergeTests.cs",
    },
    # ---- Backend / Unit / BasketTotalTests ------------------------------------------------
    {
        "feature": "Basket",
        "sub": "Basket Domain - Tính tổng tiền",
        "layer": "Backend",
        "type": "Unit",
        "name": "Total_IsZero_ForAnEmptyBasket",
        "given": "Một giỏ hàng mới chưa có sản phẩm nào.",
        "when": "Đọc thuộc tính tổng tiền của giỏ.",
        "then": "Tổng tiền bằng 0.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketTotalTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Tính tổng tiền",
        "layer": "Backend",
        "type": "Unit",
        "name": "Total_MultipliesQuantityByUnitPrice",
        "given": "Một giỏ hàng mới.",
        "when": "Thêm Notebook số lượng 2, đơn giá 12.50 rồi đọc tổng tiền.",
        "then": "Tổng tiền bằng 25.00, đúng bằng số lượng nhân đơn giá.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketTotalTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Tính tổng tiền",
        "layer": "Backend",
        "type": "Unit",
        "name": "Total_SumsEveryLine",
        "given": "Một giỏ hàng mới.",
        "when": "Thêm Notebook (2 x 12.50) và PourOver (1 x 48.00) rồi đọc tổng tiền.",
        "then": "Tổng tiền bằng 73.00 - cộng đủ mọi dòng trong giỏ.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketTotalTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Tính tổng tiền",
        "layer": "Backend",
        "type": "Unit",
        "name": "Total_MatchesTheWalkthroughFigure",
        "given": "Một giỏ hàng mới, dựng lại đúng kịch bản demo mô tả trong quickstart.md.",
        "when": "Thêm Notebook 2 lần (mỗi lần 1 x 12.50) và Apron 1 lần (1 x 34.25).",
        "then": "Tổng tiền bằng đúng con số 59.25 mà tài liệu demo công bố.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketTotalTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Tính tổng tiền",
        "layer": "Backend",
        "type": "Unit",
        "name": "Total_IsExact_ForAmountsThatFloatingPointWouldRound",
        "given": "Một giỏ hàng mới, dùng các mức giá mà số thực dấu chấm động sẽ làm tròn sai.",
        "when": "Thêm hai dòng đơn giá 0.10 và 0.20 rồi đọc tổng tiền.",
        "then": "Tổng tiền chính xác bằng 0.30 - tính bằng kiểu decimal chứ không phải floating point.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketTotalTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket Domain - Tính tổng tiền",
        "layer": "Backend",
        "type": "Unit",
        "name": "Clear_EmptiesTheBasket_AndZeroesTheTotal",
        "given": "Giỏ hàng đang có Notebook số lượng 3, đơn giá 12.50.",
        "when": "Gọi thao tác Clear để dọn sạch giỏ.",
        "then": "Danh sách dòng rỗng và tổng tiền trở về 0.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/BasketTotalTests.cs",
    },
    # ---- Backend / Unit / HealthCheckTests -------------------------------------------------
    {
        "feature": "Basket",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Unit",
        "name": "HealthLive_ReturnsOk",
        "given": "Baskets API được khởi động bằng WebApplicationFactory.",
        "when": "Gọi GET /health/live.",
        "then": "Trả về HTTP 200 OK, xác nhận tiến trình còn sống.",
        "file": "services/baskets/tests/Baskets.Api.UnitTests/HealthCheckTests.cs",
    },
    # ---- Backend / Integration / BasketEndpointsTests --------------------------------------
    {
        "feature": "Basket",
        "sub": "Basket API - Đọc giỏ theo Id",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetBasket_ReturnsTheBasket_WhenItExists",
        "given": "SQL Server thật (Testcontainers) đã seed một giỏ của khách read-by-id-shopper chứa 2 x Notebook giá 12.50, tenant contoso.",
        "when": "Gọi GET /baskets/{id} kèm header tenant.",
        "then": "Trả 200 kèm đúng Id, CustomerRef, 1 dòng số lượng 2 và tổng tiền 25.00.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/BasketEndpointsTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Đọc giỏ theo Id",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetBasket_ReturnsNotFound_WhenNoBasketHasThatId",
        "given": "Cơ sở dữ liệu không có giỏ hàng nào được seed, tenant hợp lệ.",
        "when": "Gọi GET /baskets/{id} với một Id ngẫu nhiên không tồn tại.",
        "then": "Trả 404 Not Found, phân biệt rõ trường hợp không có giỏ với trường hợp giỏ rỗng.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/BasketEndpointsTests.cs",
    },
    # ---- Backend / Integration / ClearBasketTests ------------------------------------------
    {
        "feature": "Basket",
        "sub": "Basket API - Dọn giỏ khi checkout",
        "layer": "Backend",
        "type": "Integration",
        "name": "Clear_RemovesEveryLine_ButKeepsTheBasket",
        "given": "Giỏ hiện tại của khách đã được thêm 2 x Notebook giá 12.50 qua API.",
        "when": "Gọi POST /baskets/current/clear rồi đọc lại giỏ hiện tại.",
        "then": "Trả 204 No Content; giỏ rỗng, tổng tiền 0, nhưng Id giỏ vẫn giữ nguyên như trước khi dọn.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/ClearBasketTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Dọn giỏ khi checkout",
        "layer": "Backend",
        "type": "Integration",
        "name": "Clear_ReturnsConflict_WhenTheBasketIsAlreadyEmpty",
        "given": "Khách chưa thêm sản phẩm nào, giỏ hiện tại đang rỗng.",
        "when": "Gọi POST /baskets/current/clear.",
        "then": "Trả 409 Conflict thay vì im lặng thành công, nhờ đó chặn được checkout trên giỏ rỗng.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/ClearBasketTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Dọn giỏ khi checkout",
        "layer": "Backend",
        "type": "Integration",
        "name": "Clear_ReturnsConflict_OnASecondClear",
        "given": "Giỏ đã có 1 x Notebook và vừa được dọn thành công (lần đầu trả 204).",
        "when": "Gọi POST /baskets/current/clear lần thứ hai.",
        "then": "Lần thứ hai trả 409 Conflict - checkout lặp lại thất bại rõ ràng thay vì âm thầm bỏ qua.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/ClearBasketTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Dọn giỏ khi checkout",
        "layer": "Backend",
        "type": "Integration",
        "name": "Clear_LeavesTheBasketUsable_ForTheNextPurchase",
        "given": "Giỏ đã được thêm 1 x Notebook rồi dọn sạch.",
        "when": "Thêm tiếp 3 x Notebook giá 12.50 vào chính giỏ đó rồi đọc lại.",
        "then": "Giỏ có 1 dòng số lượng 3, tổng tiền 37.50 - giỏ vẫn dùng được cho lần mua kế tiếp.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/ClearBasketTests.cs",
    },
    # ---- Backend / Integration / CurrentBasketTests ----------------------------------------
    {
        "feature": "Basket",
        "sub": "Basket API - Giỏ hiện tại theo người mua",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetCurrent_ReturnsAnEmptyBasket_ForACallerWhoHasNeverAddedAnything",
        "given": "Khách phase1-stub-user lần đầu truy cập, chưa từng thêm sản phẩm nào.",
        "when": "Gọi GET /baskets/current kèm header tenant và header định danh người gọi.",
        "then": "Trả 200 với giỏ rỗng, tổng tiền 0, CustomerRef đúng là người gọi - lần đầu mua hàng không bị coi là lỗi 404.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/CurrentBasketTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Giỏ hiện tại theo người mua",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetCurrent_ReturnsTheSameBasket_AcrossSeparateRequests",
        "given": "Khách đã thêm 2 x Notebook giá 12.50 vào giỏ hiện tại.",
        "when": "Gọi GET /baskets/current hai lần riêng biệt, mô phỏng hai lần tải trang.",
        "then": "Hai lần trả cùng một Id giỏ, vẫn 1 dòng số lượng 2 và tổng 25.00 - giỏ tồn tại qua refresh.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/CurrentBasketTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Giỏ hiện tại theo người mua",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetCurrent_GivesDifferentShoppersDifferentBaskets",
        "given": "Khách phase1-stub-user đã thêm 1 x Notebook vào giỏ của mình, cùng tenant contoso.",
        "when": "Một khách khác là someone-else gọi GET /baskets/current.",
        "then": "Khách thứ hai nhận giỏ rỗng mang CustomerRef của chính họ - mỗi người mua có một giỏ riêng.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/CurrentBasketTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Giỏ hiện tại theo người mua",
        "layer": "Backend",
        "type": "Integration",
        "name": "AddItem_MergesIntoTheExistingLine_WhenTheSameProductIsAddedAgain",
        "given": "Giỏ hiện tại của khách trên SQL Server thật.",
        "when": "Gọi POST /baskets/current/items hai lần cho cùng Notebook, mỗi lần số lượng 1.",
        "then": "Giỏ chỉ có 1 dòng số lượng 2, tổng tiền 25.00 - việc gộp dòng đúng cả khi đi qua hai request.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/CurrentBasketTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Giỏ hiện tại theo người mua",
        "layer": "Backend",
        "type": "Integration",
        "name": "AddItem_KeepsDistinctProductsOnSeparateLines",
        "given": "Giỏ hiện tại của khách trên SQL Server thật.",
        "when": "Thêm Notebook 2 lần (1 x 12.50) và Apron 1 lần (1 x 34.25) qua API.",
        "then": "Giỏ có 2 dòng, tổng tiền 59.25 đúng bằng con số trong kịch bản demo.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/CurrentBasketTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Giỏ hiện tại theo người mua",
        "layer": "Backend",
        "type": "Integration",
        "name": "AddItem_Rejects_AQuantityBelowOne",
        "given": "Giỏ hiện tại của khách; kiểm thử với số lượng không hợp lệ 0 và -3 (Theory, 2 bộ dữ liệu).",
        "when": "Gọi POST /baskets/current/items với số lượng nhỏ hơn 1.",
        "then": "API trả 400 Bad Request và không ghi nhận dòng không hợp lệ.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/CurrentBasketTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Basket API - Giỏ hiện tại theo người mua",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetCurrent_Fails_WhenNoCallerWasResolved",
        "given": "Request chỉ có header tenant, không có header định danh người gọi (tức không đi qua gateway).",
        "when": "Gọi GET /baskets/current.",
        "then": "Trả 500 - không xác định được người gọi thì không phát giỏ của bất kỳ ai, không có caller mặc định.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/CurrentBasketTests.cs",
    },
    # ---- Backend / Integration / ReadinessTests --------------------------------------------
    {
        "feature": "Basket",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_ReturnsOk_WhenDatabaseReachable",
        "given": "Baskets API trỏ tới SQL Server thật đang chạy qua Testcontainers.",
        "when": "Gọi GET /health/ready.",
        "then": "Trả 200 OK - readiness phản ánh đúng kết nối database.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_ReturnsServiceUnavailable_WhenDatabaseUnreachable",
        "given": "Baskets API trỏ tới connection string hợp lệ về cú pháp nhưng không thể kết nối được.",
        "when": "Gọi GET /health/ready.",
        "then": "Trả 503 Service Unavailable - readiness fail-closed chứ không báo healthy giả.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_DoesNotFallBackToAnotherServicesDatabase_WhenOwnDatabaseUnreachable",
        "given": "Database riêng của Baskets không kết nối được, nhưng cấu hình vẫn cung cấp connection string còn sống của PartiesDb, ProductsDb và OrdersDb.",
        "when": "Gọi GET /health/ready.",
        "then": "Vẫn trả 503 và nội dung phản hồi nêu rõ self-database là Unhealthy - dịch vụ không âm thầm dùng database của service khác.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/ReadinessTests.cs",
    },
    # ---- Backend / Integration / TenantEnforcementTests -------------------------------------
    {
        "feature": "Basket",
        "sub": "Tenant Enforcement",
        "layer": "Backend",
        "type": "Integration",
        "name": "ResolvingTheDbContext_Throws_WhenNoTenantHasBeenResolved",
        "given": "Một DI scope được tạo mà chưa có request HTTP nào phân giải tenant.",
        "when": "Yêu cầu resolve BasketsDbContext từ service provider.",
        "then": "Ném MissingTenantContextException - không có tenant thì không được chạm vào database.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/TenantEnforcementTests.cs",
    },
    {
        "feature": "Basket",
        "sub": "Tenant Enforcement",
        "layer": "Backend",
        "type": "Integration",
        "name": "ARequestWithoutATenant_Fails_RatherThanServingDefaultSchemaData",
        "given": "Request gọi thẳng vào Baskets API, không đi qua gateway nên không mang header tenant.",
        "when": "Gọi GET /baskets/{id}.",
        "then": "Trả 500 - thất bại rõ ràng thay vì phục vụ dữ liệu của schema mặc định.",
        "file": "services/baskets/tests/Baskets.Api.IntegrationTests/TenantEnforcementTests.cs",
    },
]

ROWS += [
    # ---- Frontend / Unit / basket/AddItemError.test.tsx --------------------------------------
    {
        "feature": "Basket",
        "sub": "Storefront - Nút thêm vào giỏ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "adds the product when the request succeeds",
        "given": "Nút AddToBasketButton cho sản phẩm Field Notes Notebook, API giả lập (MSW) trả về thành công.",
        "when": "Người mua bấm nút Add to basket.",
        "then": "Gửi đúng một request chỉ chứa productId và quantity = 1, không kèm giá - giá là do server quyết định.",
        "file": "frontend/apps/web/tests/basket/AddItemError.test.tsx",
    },
    {
        "feature": "Basket",
        "sub": "Storefront - Nút thêm vào giỏ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows a clear error when the request fails",
        "given": "API giả lập trả về lỗi 502 khi thêm sản phẩm vào giỏ.",
        "when": "Người mua bấm nút Add to basket.",
        "then": "Hiển thị thông báo lỗi dạng alert với nội dung không thể thêm - giỏ không hiển thị món chưa thực sự được thêm.",
        "file": "frontend/apps/web/tests/basket/AddItemError.test.tsx",
    },
    {
        "feature": "Basket",
        "sub": "Storefront - Nút thêm vào giỏ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "disables itself while the addition is in flight",
        "given": "API giả lập giữ request lại chưa trả lời ngay.",
        "when": "Người mua bấm nút và theo dõi trạng thái nút trong lúc chờ.",
        "then": "Nút bị vô hiệu hóa khi đang gửi và bật lại sau khi có phản hồi - bấm liên tục không xếp hàng nhiều lần thêm.",
        "file": "frontend/apps/web/tests/basket/AddItemError.test.tsx",
    },
    {
        "feature": "Basket",
        "sub": "Storefront - Nút thêm vào giỏ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "can be operated by keyboard",
        "given": "Nút AddToBasketButton hiển thị trên màn hình, API giả lập trả về thành công.",
        "when": "Nhấn Tab để đưa focus vào nút rồi nhấn Enter.",
        "then": "Nút nhận focus và request thêm sản phẩm được gửi - thao tác mua hàng không chỉ dành cho chuột (FR-017, SC-009).",
        "file": "frontend/apps/web/tests/basket/AddItemError.test.tsx",
    },
    # ---- Frontend / Unit / basket/BasketView.test.tsx -----------------------------------------
    {
        "feature": "Basket",
        "sub": "Storefront - Màn hình giỏ hàng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows each line with its name, quantity, unit price, and line total",
        "given": "API giả lập trả giỏ có 1 dòng: Field Notes Notebook, số lượng 2, đơn giá 12.50, thành tiền 25.",
        "when": "Hiển thị màn hình BasketView.",
        "then": "Dòng hiển thị tên sản phẩm, cụm Quantity: 2 × $12.50 và thành tiền $25.00.",
        "file": "frontend/apps/web/tests/basket/BasketView.test.tsx",
    },
    {
        "feature": "Basket",
        "sub": "Storefront - Màn hình giỏ hàng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows the basket total the backend reported",
        "given": "API giả lập trả giỏ 2 dòng với tổng tiền 59.25 do backend tính.",
        "when": "Hiển thị màn hình BasketView.",
        "then": "Màn hình hiện nhãn tổng và số tiền $59.25 lấy từ backend, không tự tính lại ở trình duyệt.",
        "file": "frontend/apps/web/tests/basket/BasketView.test.tsx",
    },
    {
        "feature": "Basket",
        "sub": "Storefront - Màn hình giỏ hàng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "presents the lines as a list",
        "given": "API giả lập trả giỏ có 2 dòng sản phẩm.",
        "when": "Hiển thị màn hình BasketView.",
        "then": "Nội dung được đánh dấu là danh sách có nhãn Basket với đúng 2 listitem - người dùng trình đọc màn hình biết được số dòng.",
        "file": "frontend/apps/web/tests/basket/BasketView.test.tsx",
    },
    {
        "feature": "Basket",
        "sub": "Storefront - Màn hình giỏ hàng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "tells the shopper when the basket is empty",
        "given": "API giả lập trả giỏ rỗng, tổng 0.",
        "when": "Hiển thị màn hình BasketView.",
        "then": "Hiện thông báo giỏ hàng đang trống và không có alert nào - giỏ rỗng là trạng thái hợp lệ, không phải lỗi.",
        "file": "frontend/apps/web/tests/basket/BasketView.test.tsx",
    },
    {
        "feature": "Basket",
        "sub": "Storefront - Màn hình giỏ hàng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows a readable error when the basket cannot be loaded",
        "given": "API giả lập trả lỗi 502 khi tải giỏ hàng.",
        "when": "Hiển thị màn hình BasketView.",
        "then": "Hiện alert lỗi kèm nút Try again - trang vẫn dùng được và người mua thử lại được.",
        "file": "frontend/apps/web/tests/basket/BasketView.test.tsx",
    },
]

# ==========================================================================================
# GATEWAY — services/gateway/**
# ==========================================================================================

ROWS += [
    # ---- Backend / Unit / ForwardingTimeoutBudgetTests ------------------------------------
    {
        "feature": "Gateway",
        "sub": "Ngân sách timeout chuyển tiếp",
        "layer": "Backend",
        "type": "Unit",
        "name": "TheGatewaysForwardingTimeout_IsAtLeastTheBffsTotalDownstreamBudget",
        "given": "appsettings.json của Gateway khai báo ActivityTimeout cho bff-cluster; BFF có ngân sách gọi downstream tối đa 3 giây.",
        "when": "Đọc ActivityTimeout từ file cấu hình đã commit và so với ngân sách của BFF.",
        "then": "ActivityTimeout lớn hơn hoặc bằng 3 giây - gateway không cắt request khi BFF vẫn đang chờ hợp lệ.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/ForwardingTimeoutBudgetTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Ngân sách timeout chuyển tiếp",
        "layer": "Backend",
        "type": "Unit",
        "name": "TheGatewaysForwardingTimeout_IsBounded",
        "given": "appsettings.json của Gateway khai báo ActivityTimeout cho bff-cluster.",
        "when": "Đọc giá trị ActivityTimeout đang cấu hình.",
        "then": "Giá trị khác vô hạn và nằm trong khoảng từ 3 giây đến 1 phút - không tồn tại thời gian chờ không giới hạn.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/ForwardingTimeoutBudgetTests.cs",
    },
    # ---- Backend / Unit / RouteConfigurationTests -------------------------------------------
    {
        "feature": "Gateway",
        "sub": "Cấu hình định tuyến YARP",
        "layer": "Backend",
        "type": "Unit",
        "name": "TheConfiguration_DefinesExactlyOneRoute_ToTheBffCluster",
        "given": "appsettings.json thật của Gateway được nạp qua chính cơ chế binding LoadFromConfig của YARP.",
        "when": "Đọc danh sách route trong cấu hình proxy.",
        "then": "Có đúng một route với RouteId là bff-route và ClusterId là bff-cluster.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/RouteConfigurationTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Cấu hình định tuyến YARP",
        "layer": "Backend",
        "type": "Unit",
        "name": "TheRoute_MatchesEveryPath",
        "given": "Cấu hình proxy thật của Gateway đã được nạp.",
        "when": "Đọc điều kiện Match.Path của route duy nhất.",
        "then": "Route dùng catch-all {**catch-all} - gateway không phải sửa mỗi khi BFF thêm đường dẫn mới.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/RouteConfigurationTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Cấu hình định tuyến YARP",
        "layer": "Backend",
        "type": "Unit",
        "name": "TheConfiguration_DefinesExactlyOneCluster_WithOneDestination",
        "given": "Cấu hình proxy thật của Gateway đã được nạp.",
        "when": "Đọc danh sách cluster và destination.",
        "then": "Có đúng một cluster tên bff-cluster với đúng một destination có địa chỉ không rỗng.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/RouteConfigurationTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Cấu hình định tuyến YARP",
        "layer": "Backend",
        "type": "Unit",
        "name": "TheConfiguration_NamesNoDomainServiceAsADestination",
        "given": "Cấu hình proxy thật của Gateway đã được nạp.",
        "when": "Duyệt toàn bộ địa chỉ destination của mọi cluster.",
        "then": "Không địa chỉ nào chứa tên products, baskets, orders hay parties - không thể đi tắt vào domain service mà không qua BFF.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/RouteConfigurationTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Cấu hình định tuyến YARP",
        "layer": "Backend",
        "type": "Unit",
        "name": "EveryRoute_ResolvesToADefinedCluster",
        "given": "Cấu hình proxy thật của Gateway đã được nạp.",
        "when": "Đối chiếu ClusterId của từng route với danh sách cluster đã khai báo.",
        "then": "Mọi route đều có ClusterId khác null và trỏ tới một cluster tồn tại - không có route dẫn đến hư không.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/RouteConfigurationTests.cs",
    },
    # ---- Backend / Unit / StubIdentityAuthenticationHandlerTests ----------------------------
    {
        "feature": "Gateway",
        "sub": "Stub Identity (định danh giai đoạn 1)",
        "layer": "Backend",
        "type": "Unit",
        "name": "AuthenticateAsync_Succeeds_ForAnyRequest",
        "given": "StubIdentityAuthenticationHandler được cấu hình tenant contoso và subject phase1-stub-user.",
        "when": "Thực hiện xác thực cho một request bất kỳ.",
        "then": "Xác thực thành công và không có thông tin lỗi kèm theo.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/StubIdentityAuthenticationHandlerTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Stub Identity (định danh giai đoạn 1)",
        "layer": "Backend",
        "type": "Unit",
        "name": "AuthenticateAsync_IssuesTheConfiguredTenantClaim",
        "given": "Handler được cấu hình tenant là contoso.",
        "when": "Thực hiện xác thực và đọc claim tenant trong principal trả về.",
        "then": "Principal mang claim tenant đúng bằng giá trị đã cấu hình.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/StubIdentityAuthenticationHandlerTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Stub Identity (định danh giai đoạn 1)",
        "layer": "Backend",
        "type": "Unit",
        "name": "AuthenticateAsync_IssuesTheConfiguredSubjectClaim",
        "given": "Handler được cấu hình subject là phase1-stub-user.",
        "when": "Thực hiện xác thực và đọc claim NameIdentifier.",
        "then": "Principal mang cả subject lẫn tenant, đúng hình dạng của định danh thật sẽ thay thế ở giai đoạn 3.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/StubIdentityAuthenticationHandlerTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Stub Identity (định danh giai đoạn 1)",
        "layer": "Backend",
        "type": "Unit",
        "name": "AuthenticateAsync_IgnoresTheRequest_HavingNoCredentialsToRead",
        "given": "Request mang header Authorization giả (Bearer not-a-real-token), thử với hai đường dẫn /bff/products và /anything-at-all (Theory, 2 bộ dữ liệu).",
        "when": "Thực hiện xác thực.",
        "then": "Vẫn thành công với đúng tenant đã cấu hình - nội dung request không tác động được đến kết quả xác thực.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/StubIdentityAuthenticationHandlerTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Stub Identity (định danh giai đoạn 1)",
        "layer": "Backend",
        "type": "Unit",
        "name": "AuthenticateAsync_Fails_WhenNoTenantIsConfigured",
        "given": "Handler được cấu hình tenant rỗng hoặc chỉ có khoảng trắng (Theory, 2 bộ dữ liệu).",
        "when": "Thực hiện xác thực.",
        "then": "Xác thực thất bại và không trả về principal - gateway chưa cấu hình thì không xác thực cho ai cả.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/StubIdentityAuthenticationHandlerTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Stub Identity (định danh giai đoạn 1)",
        "layer": "Backend",
        "type": "Unit",
        "name": "AuthenticateAsync_IssuesAnIdentityNamingTheStubScheme",
        "given": "Handler được cấu hình đầy đủ tenant và subject.",
        "when": "Thực hiện xác thực và đọc AuthenticationType của identity.",
        "then": "Identity mang đúng tên scheme của stub - đây là một authentication scheme thật chứ không phải mẹo gắn header.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/StubIdentityAuthenticationHandlerTests.cs",
    },
    # ---- Backend / Unit / SubjectHeaderPropagationMiddlewareTests ---------------------------
    {
        "feature": "Gateway",
        "sub": "Truyền header định danh người gọi",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_StampsTheSubjectHeader_FromTheAuthenticatedPrincipal",
        "given": "HttpContext có principal đã xác thực mang subject phase1-stub-user.",
        "when": "Chạy SubjectHeaderPropagationMiddleware.",
        "then": "Header subject trên request được gắn đúng giá trị lấy từ principal.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/SubjectHeaderPropagationMiddlewareTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Truyền header định danh người gọi",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_OverwritesACallerSuppliedSubject_NeverTrustsIt",
        "given": "Principal phân giải ra phase1-stub-user, nhưng người gọi tự gửi kèm header subject là somebody-else.",
        "when": "Chạy middleware.",
        "then": "Header bị ghi đè bằng subject đã phân giải - người gọi không thể tự khai mình là người khác để xem giỏ của họ.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/SubjectHeaderPropagationMiddlewareTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Truyền header định danh người gọi",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_RemovesTheHeader_WhenNoSubjectIsResolved",
        "given": "Không phân giải được subject (null, rỗng, hoặc khoảng trắng - Theory 3 bộ dữ liệu), trong khi người gọi vẫn gửi header subject somebody-else.",
        "when": "Chạy middleware.",
        "then": "Header bị gỡ bỏ hoàn toàn khỏi request - giá trị do người gọi tự đặt không được lọt xuống dưới.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/SubjectHeaderPropagationMiddlewareTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Truyền header định danh người gọi",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_AlwaysCallsTheRestOfThePipeline",
        "given": "Hai trường hợp: có subject phân giải được và không có subject nào (Theory, 2 bộ dữ liệu).",
        "when": "Chạy middleware với một delegate kế tiếp có đánh dấu.",
        "then": "Phần còn lại của pipeline luôn được gọi trong cả hai trường hợp.",
        "file": "services/gateway/tests/Gateway.Api.UnitTests/SubjectHeaderPropagationMiddlewareTests.cs",
    },
    # ---- Backend / Integration / CorrelationIdPropagationTests -----------------------------
    {
        "feature": "Gateway",
        "sub": "Correlation Id đầu-cuối",
        "layer": "Backend",
        "type": "Integration",
        "name": "AGeneratedCorrelationId_ReachesTheBff_AndMatchesWhatTheCallerIsGiven",
        "given": "Gateway và BFF cùng chạy trong test host; request không mang sẵn header X-Correlation-Id.",
        "when": "Gọi GET /bff/products qua gateway.",
        "then": "Correlation id gateway trả cho người gọi trùng đúng với correlationId mà BFF ghi trong ProblemDetails - id sinh ở biên được truyền tiếp, BFF không sinh id thứ hai.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/CorrelationIdPropagationTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Correlation Id đầu-cuối",
        "layer": "Backend",
        "type": "Integration",
        "name": "ACallerSuppliedCorrelationId_IsPreservedEndToEnd",
        "given": "Người gọi tự gửi kèm X-Correlation-Id là caller-supplied-correlation-id.",
        "when": "Gọi GET /bff/products qua gateway.",
        "then": "Cả header phản hồi lẫn correlationId trong body đều giữ nguyên giá trị người gọi cung cấp, không bị thay thế.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/CorrelationIdPropagationTests.cs",
    },
    # ---- Backend / Integration / DownstreamUnavailableTests --------------------------------
    {
        "feature": "Gateway",
        "sub": "Xử lý khi BFF không sẵn sàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "ARequest_ReturnsAClearError_WhenTheBffIsUnreachable",
        "given": "Gateway được cấu hình trỏ tới địa chỉ BFF không có gì lắng nghe (127.0.0.1:1), BFF không được khởi động.",
        "when": "Gọi GET /bff/products và đo thời gian phản hồi.",
        "then": "Trả 502 Bad Gateway trong dưới 5 giây - lỗi dứt khoát chứ không treo.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/DownstreamUnavailableTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Xử lý khi BFF không sẵn sàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheGatewaysOwnHealth_StaysHealthy_WhenTheBffIsUnreachable",
        "given": "Gateway trỏ tới BFF không thể kết nối.",
        "when": "Gọi GET /health/live và GET /health/ready của chính gateway.",
        "then": "Cả hai trả 200 OK - sự cố của BFF không kéo theo việc gateway bị khởi động lại.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/DownstreamUnavailableTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Xử lý khi BFF không sẵn sàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheError_LeaksNoInternalRoutingDetail",
        "given": "Gateway trỏ tới BFF không thể kết nối.",
        "when": "Gọi GET /bff/products và đọc nội dung phản hồi lỗi.",
        "then": "Body không chứa 127.0.0.1, bff-cluster, bff-route hay SocketException - không lộ chi tiết định tuyến nội bộ.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/DownstreamUnavailableTests.cs",
    },
    # ---- Backend / Integration / RoutingTests -----------------------------------------------
    {
        "feature": "Gateway",
        "sub": "Định tuyến qua Gateway",
        "layer": "Backend",
        "type": "Integration",
        "name": "ARequestToTheGateway_ReachesAResponseOnlyTheBffCanProduce",
        "given": "Gateway và BFF cùng chạy; người gọi chỉ biết đường dẫn, không biết địa chỉ hay cổng của BFF.",
        "when": "Gọi GET /openapi/v1.json qua gateway.",
        "then": "Trả 200 kèm tài liệu OpenAPI có chứa đường dẫn /bff/products - thứ chỉ BFF mới tạo ra được, chứng minh request đã tới đúng đích.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/RoutingTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Định tuyến qua Gateway",
        "layer": "Backend",
        "type": "Integration",
        "name": "AClientFacingRoute_IsForwardedToTheBffsHandler",
        "given": "Gateway và BFF cùng chạy, service products phía sau BFF không chạy.",
        "when": "Gọi GET /bff/products qua gateway.",
        "then": "Kết quả khác 404 - chứng tỏ request đã được chuyển tới handler của BFF chứ không dừng ở gateway.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/RoutingTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Định tuyến qua Gateway",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheGatewaysOwnHealthProbes_AreServedLocally_NotForwarded",
        "given": "Gateway có route catch-all; kiểm thử với /health/live và /health/ready (Theory, 2 bộ dữ liệu).",
        "when": "Gọi các endpoint health của chính gateway.",
        "then": "Trả 200 OK và được phục vụ tại chỗ - catch-all không nuốt mất health probe của gateway.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/RoutingTests.cs",
    },
    # ---- Backend / Integration / StorefrontCorsTests ---------------------------------------
    {
        "feature": "Gateway",
        "sub": "CORS cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "APreflightFromTheStorefront_IsAllowed",
        "given": "Gateway cấu hình cho phép origin http://localhost:5173.",
        "when": "Gửi preflight OPTIONS /bff/products kèm Origin và Access-Control-Request-Method: GET.",
        "then": "Phản hồi có Access-Control-Allow-Origin đúng bằng origin của storefront, do chính gateway trả lời chứ không chuyển tiếp xuống BFF.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/StorefrontCorsTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "CORS cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "APreflightFromTheStorefront_AllowsCredentials",
        "given": "Gateway cấu hình cho phép origin của storefront; client gửi kèm credentials.",
        "when": "Gửi preflight OPTIONS /bff/basket với Access-Control-Request-Method: POST.",
        "then": "Access-Control-Allow-Credentials là true và Allow-Origin không phải dấu * - đúng ràng buộc của chuẩn CORS cho request có credentials.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/StorefrontCorsTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "CORS cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "ARequestFromAnUnknownOrigin_IsNotAdmitted",
        "given": "Gateway chỉ cho phép origin của storefront.",
        "when": "Gửi preflight từ origin lạ http://evil.example.",
        "then": "Phản hồi không có header Access-Control-Allow-Origin, trình duyệt sẽ chặn - danh sách origin là allow-list chứ không mở cửa tự do.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/StorefrontCorsTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "CORS cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheAllowedOrigins_ComeFromConfiguration",
        "given": "Gateway được nạp cấu hình với origin tùy chỉnh https://storefront.example.",
        "when": "Gửi preflight từ chính origin đó.",
        "then": "Được chấp nhận với Allow-Origin tương ứng - đổi origin chỉ cần đổi cấu hình, không cần build lại.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/StorefrontCorsTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "CORS cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "ARequestWithNoOrigin_IsUntouched",
        "given": "Gateway đã bật chính sách CORS.",
        "when": "Gọi GET /health/live không kèm header Origin (request same-origin).",
        "then": "Trả 200 OK và không gắn thêm header Access-Control-Allow-Origin - chính sách CORS không ảnh hưởng request same-origin.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/StorefrontCorsTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "CORS cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "EachConfiguredOrigin_IsAdmitted",
        "given": "Gateway cấu hình đồng thời hai origin: dev server 5173 và container 4173 (Theory, 2 bộ dữ liệu).",
        "when": "Gửi preflight lần lượt từ từng origin.",
        "then": "Mỗi origin đều nhận đúng Allow-Origin của mình - chính sách hỗ trợ nhiều origin cùng lúc.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/StorefrontCorsTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "CORS cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheDevelopmentConfiguration_Admits_BothStorefrontOrigins",
        "given": "Gateway chạy với file cấu hình Development thật đã commit, không thay thế bằng cấu hình trong bộ nhớ; kiểm thử với origin 5173 và 4173 (Theory, 2 bộ dữ liệu).",
        "when": "Gửi preflight OPTIONS /bff/products từ từng origin.",
        "then": "Cả hai đều có Allow-Origin - repo thực sự được cấu hình đúng cho các storefront đang ship, không chỉ đúng về mặt cơ chế.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/StorefrontCorsTests.cs",
    },
    # ---- Backend / Integration / TenantPropagationTests ------------------------------------
    {
        "feature": "Gateway",
        "sub": "Truyền tenant xuống BFF",
        "layer": "Backend",
        "type": "Integration",
        "name": "ARequestThroughTheGateway_CarriesTheResolvedTenantToTheBff",
        "given": "Gateway chạy cùng một BFF có gắn filter ghi lại header tenant nhận được ở đầu pipeline.",
        "when": "Gọi GET /bff/products qua gateway.",
        "then": "BFF nhận đúng tenant mà gateway đã phân giải từ cấu hình StubIdentity.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/TenantPropagationTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Truyền tenant xuống BFF",
        "layer": "Backend",
        "type": "Integration",
        "name": "ACallerSuppliedTenant_IsOverwritten_NeverTrusted",
        "given": "Người gọi tự gửi kèm header X-Tenant-Id với giá trị some-other-tenant.",
        "when": "Gọi GET /bff/products qua gateway.",
        "then": "BFF nhận tenant do gateway phân giải, không phải giá trị người gọi khai - ranh giới cô lập tenant không bị phá.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/TenantPropagationTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Truyền tenant xuống BFF",
        "layer": "Backend",
        "type": "Integration",
        "name": "EveryForwardedRequest_CarriesATenant",
        "given": "Kiểm thử hai đường dẫn /bff/products và /bff/baskets/{id} (Theory, 2 bộ dữ liệu).",
        "when": "Gọi từng đường dẫn qua gateway.",
        "then": "Mọi request được chuyển tiếp đều mang tenant không rỗng, kể cả các route không dẫn tới đâu hữu ích.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/TenantPropagationTests.cs",
    },
    # ---- Backend / Integration / UnmatchedRouteTests ---------------------------------------
    {
        "feature": "Gateway",
        "sub": "Đường dẫn không khớp route",
        "layer": "Backend",
        "type": "Integration",
        "name": "AnUnknownPath_ReturnsAClearNotFound_RatherThanHanging",
        "given": "Gateway và BFF cùng chạy; kiểm thử với /no-such-path, /bff/no-such-resource và /bff/products/extra/segments (Theory, 3 bộ dữ liệu).",
        "when": "Gọi các đường dẫn không tồn tại và đo thời gian phản hồi.",
        "then": "Trả 404 Not Found trong dưới 15 giây - lỗi rõ ràng chứ không treo.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/UnmatchedRouteTests.cs",
    },
    {
        "feature": "Gateway",
        "sub": "Đường dẫn không khớp route",
        "layer": "Backend",
        "type": "Integration",
        "name": "AnUnknownPathsResponse_LeaksNoInternalRoutingDetail",
        "given": "Gateway và BFF cùng chạy.",
        "when": "Gọi GET /no-such-path và đọc nội dung phản hồi.",
        "then": "Trả 404 và body không chứa bff-cluster, bff-route, products-api hay số cổng 8080.",
        "file": "services/gateway/tests/Gateway.Api.IntegrationTests/UnmatchedRouteTests.cs",
    },
]

# ==========================================================================================
# ORDER — services/orders/** (Backend) + frontend checkout & E2E đặt hàng (Frontend)
# ==========================================================================================

ROWS += [
    # ---- Backend / Unit / HealthCheckTests --------------------------------------------------
    {
        "feature": "Order",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Unit",
        "name": "HealthLive_ReturnsOk",
        "given": "Orders API được khởi động bằng WebApplicationFactory.",
        "when": "Gọi GET /health/live.",
        "then": "Trả về HTTP 200 OK, xác nhận tiến trình còn sống.",
        "file": "services/orders/tests/Orders.Api.UnitTests/HealthCheckTests.cs",
    },
    # ---- Backend / Unit / OrderTenantTests --------------------------------------------------
    {
        "feature": "Order",
        "sub": "Order Domain - Gắn tenant cho đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_RecordsTheTenantItWasPlacedFor",
        "given": "Một dòng đơn hàng hợp lệ (Notebook, số lượng 1, đơn giá 12.50) và tenant contoso.",
        "when": "Tạo đơn hàng bằng Order.PlaceFrom.",
        "then": "Đơn hàng ghi nhận đúng TenantId là contoso - mọi đơn đều truy được về tenant đã đặt.",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTenantTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order Domain - Gắn tenant cho đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_Rejects_AnAbsentOrBlankTenant",
        "given": "Dòng đơn hàng hợp lệ nhưng tenant là null, chuỗi rỗng, khoảng trắng hoặc tab (Theory, 4 bộ dữ liệu).",
        "when": "Tạo đơn hàng bằng Order.PlaceFrom.",
        "then": "Ném ArgumentException - tenant rỗng bị coi là không có tenant, không thể tạo đơn không quy được chủ.",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTenantTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order Domain - Gắn tenant cho đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_StillRejects_AnEmptyLineSet_EvenWithATenant",
        "given": "Tenant hợp lệ nhưng danh sách dòng đơn hàng rỗng.",
        "when": "Tạo đơn hàng bằng Order.PlaceFrom.",
        "then": "Vẫn ném ArgumentException - việc bổ sung kiểm tra tenant không làm mất các luật đã có trước đó.",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTenantTests.cs",
    },
    # ---- Backend / Unit / OrderTotalTests ---------------------------------------------------
    {
        "feature": "Order",
        "sub": "Order Domain - Tính tổng đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_MultipliesQuantityByUnitPrice",
        "given": "Một dòng đơn hàng Notebook số lượng 2, đơn giá 12.50.",
        "when": "Tạo đơn hàng và đọc tổng tiền.",
        "then": "Tổng bằng 25.00 - orders service tự tính tổng chứ không nhận tổng từ bên ngoài.",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTotalTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order Domain - Tính tổng đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_SumsEveryLine",
        "given": "Hai dòng đơn hàng: Notebook 2 x 12.50 và Apron 1 x 34.25.",
        "when": "Tạo đơn hàng và đọc tổng tiền.",
        "then": "Tổng bằng 59.25, đúng con số trong kịch bản demo quickstart.md.",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTotalTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order Domain - Tính tổng đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_RecordsWhenTheOrderWasPlaced_AndGivesItAnIdentifier",
        "given": "Một dòng đơn hàng hợp lệ và thời điểm đặt hàng 2026-08-16 09:30 UTC.",
        "when": "Tạo đơn hàng bằng Order.PlaceFrom.",
        "then": "Đơn ghi đúng thời điểm đặt và được cấp một Id khác Guid rỗng.",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTotalTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order Domain - Tính tổng đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_Rejects_AnEmptyLineSet",
        "given": "Danh sách dòng đơn hàng rỗng, tenant hợp lệ.",
        "when": "Tạo đơn hàng bằng Order.PlaceFrom.",
        "then": "Ném ArgumentException - đây là lớp trong cùng trong ba lớp chặn checkout với giỏ rỗng (FR-008).",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTotalTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order Domain - Tính tổng đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_Rejects_ALineWithANonPositiveQuantity",
        "given": "Một dòng đơn hàng có số lượng 0 hoặc -1 (Theory, 2 bộ dữ liệu).",
        "when": "Tạo đơn hàng bằng Order.PlaceFrom.",
        "then": "Ném ArgumentOutOfRangeException.",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTotalTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order Domain - Tính tổng đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_Rejects_ALineWithANegativePrice",
        "given": "Một dòng đơn hàng có đơn giá âm (-0.01).",
        "when": "Tạo đơn hàng bằng Order.PlaceFrom.",
        "then": "Ném ArgumentOutOfRangeException.",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTotalTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order Domain - Tính tổng đơn hàng",
        "layer": "Backend",
        "type": "Unit",
        "name": "PlaceFrom_IsExact_ForAmountsThatFloatingPointWouldRound",
        "given": "Hai dòng đơn hàng đơn giá 0.10 và 0.20 - các mức giá mà floating point sẽ làm tròn sai.",
        "when": "Tạo đơn hàng và đọc tổng tiền.",
        "then": "Tổng chính xác bằng 0.30 - tổng đơn không bị lệch một xu trên màn hình xác nhận.",
        "file": "services/orders/tests/Orders.Api.UnitTests/OrderTotalTests.cs",
    },
    # ---- Backend / Integration / OrderEndpointsTests ----------------------------------------
    {
        "feature": "Order",
        "sub": "Order API - Đọc đơn theo Id",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetOrder_ReturnsTheOrder_WhenItExists",
        "given": "SQL Server thật đã seed một đơn hàng đặt lúc 2026-08-15 09:30 UTC, tổng 47.49, tenant contoso.",
        "when": "Gọi GET /orders/{id} kèm header tenant.",
        "then": "Trả 200 với đúng Id, thời điểm đặt và tổng tiền như đã lưu.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/OrderEndpointsTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order API - Đọc đơn theo Id",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetOrder_ReturnsTheTenantTheOrderBelongsTo",
        "given": "Một đơn hàng tổng 59.25 thuộc tenant contoso đã được seed trong database.",
        "when": "Gọi GET /orders/{id} và đọc trường TenantId trong phản hồi.",
        "then": "TenantId không rỗng và đúng bằng contoso - đọc đơn là biết được nó thuộc tenant nào mà không cần soi database.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/OrderEndpointsTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order API - Đọc đơn theo Id",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetOrder_ReturnsNotFound_WhenNoOrderHasThatId",
        "given": "Database không có đơn hàng nào được seed, tenant hợp lệ.",
        "when": "Gọi GET /orders/{id} với Id ngẫu nhiên không tồn tại.",
        "then": "Trả 404 Not Found.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/OrderEndpointsTests.cs",
    },
    # ---- Backend / Integration / PlaceOrderTests ---------------------------------------------
    {
        "feature": "Order",
        "sub": "Order API - Đặt hàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "PlaceOrder_CreatesTheOrder_AndComputesItsTotal",
        "given": "Orders API chạy trên SQL Server thật, request mang đủ header tenant và người mua.",
        "when": "POST /orders với hai dòng: Notebook 2 x 12.50 và Apron 1 x 34.25.",
        "then": "Trả 201 Created, đơn có Id khác rỗng và tổng tiền do server tự tính bằng 59.25.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/PlaceOrderTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order API - Đặt hàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "PlaceOrder_ReturnsAnIdentifier_ThatReadsBackAsTheSameOrder",
        "given": "Một đơn hàng vừa được đặt qua POST /orders.",
        "when": "Dùng chính Id trả về để gọi GET /orders/{id}.",
        "then": "Đơn đọc lại trùng Id, tổng tiền và thời điểm đặt - mã đơn trên màn hình xác nhận đúng là đơn đã tạo (SC-005).",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/PlaceOrderTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order API - Đặt hàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "PlaceOrder_ReturnsALocationHeader_ForTheCreatedOrder",
        "given": "Request đặt hàng hợp lệ với một dòng Notebook 1 x 12.50.",
        "when": "POST /orders và đọc header Location của phản hồi.",
        "then": "Location trỏ đúng tới /orders/{id} của đơn vừa tạo.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/PlaceOrderTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order API - Đặt hàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "PlaceOrder_Rejects_ARequestWithNoLines",
        "given": "Request đặt hàng có danh sách items rỗng, gọi thẳng vào API bỏ qua kiểm tra phía giao diện.",
        "when": "POST /orders.",
        "then": "Trả 400 Bad Request - không thể tạo đơn cho giỏ rỗng ngay cả khi gọi trực tiếp API.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/PlaceOrderTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order API - Đặt hàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "PlaceOrder_Rejects_ALineWithANonPositiveQuantity",
        "given": "Request đặt hàng có một dòng với số lượng bằng 0.",
        "when": "POST /orders.",
        "then": "Trả 400 Bad Request.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/PlaceOrderTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order API - Đặt hàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "PlaceOrder_Fails_WhenNoCallerWasResolved",
        "given": "Request chỉ có header tenant, không có header định danh người mua (không đi qua gateway).",
        "when": "POST /orders với một dòng hợp lệ.",
        "then": "Trả 500 - đơn hàng phải thuộc về một ai đó, không xác định được người mua thì không được tạo đơn.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/PlaceOrderTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order API - Đặt hàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "PlaceOrder_PersistsTheResolvedTenant_OnTheOrderRow",
        "given": "Request đặt hàng hợp lệ với tenant contoso được phân giải từ header.",
        "when": "POST /orders rồi đọc trực tiếp bản ghi trong database.",
        "then": "Cột TenantId của bản ghi đúng bằng contoso - tenant thực sự được lưu chứ không chỉ vang lại trong phản hồi.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/PlaceOrderTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Order API - Đặt hàng",
        "layer": "Backend",
        "type": "Integration",
        "name": "PlaceOrder_IgnoresATenantNamedInTheRequestBody",
        "given": "Request đặt hàng cố tình nhét thêm trường tenantId = someone-elses-tenant vào body, trong khi header phân giải ra contoso.",
        "when": "POST /orders rồi đọc bản ghi đã lưu trong database.",
        "then": "TenantId lưu vẫn là contoso - người gọi không thể tự chọn tenant cho đơn hàng của mình.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/PlaceOrderTests.cs",
    },
    # ---- Backend / Integration / ReadinessTests ---------------------------------------------
    {
        "feature": "Order",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_ReturnsOk_WhenDatabaseReachable",
        "given": "Orders API trỏ tới SQL Server thật đang chạy qua Testcontainers.",
        "when": "Gọi GET /health/ready.",
        "then": "Trả 200 OK - readiness phản ánh đúng kết nối database.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_ReturnsServiceUnavailable_WhenDatabaseUnreachable",
        "given": "Orders API trỏ tới connection string hợp lệ về cú pháp nhưng không thể kết nối được.",
        "when": "Gọi GET /health/ready.",
        "then": "Trả 503 Service Unavailable - readiness fail-closed chứ không báo healthy giả.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_DoesNotFallBackToAnotherServicesDatabase_WhenOwnDatabaseUnreachable",
        "given": "Database riêng của Orders không kết nối được, nhưng cấu hình vẫn cung cấp connection string còn sống của PartiesDb, ProductsDb và BasketsDb.",
        "when": "Gọi GET /health/ready.",
        "then": "Vẫn trả 503 và phản hồi nêu rõ self-database là Unhealthy - không âm thầm dùng database của service khác.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/ReadinessTests.cs",
    },
    # ---- Backend / Integration / TenantEnforcementTests --------------------------------------
    {
        "feature": "Order",
        "sub": "Tenant Enforcement",
        "layer": "Backend",
        "type": "Integration",
        "name": "ResolvingTheDbContext_Throws_WhenNoTenantHasBeenResolved",
        "given": "Một DI scope được tạo mà chưa có request HTTP nào phân giải tenant.",
        "when": "Yêu cầu resolve OrdersDbContext từ service provider.",
        "then": "Ném MissingTenantContextException - không có tenant thì không được chạm vào database.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/TenantEnforcementTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Tenant Enforcement",
        "layer": "Backend",
        "type": "Integration",
        "name": "ARequestWithoutATenant_Fails_RatherThanServingDefaultSchemaData",
        "given": "Request gọi thẳng vào Orders API, không mang header tenant.",
        "when": "Gọi GET /orders/{id}.",
        "then": "Trả 500 - thất bại rõ ràng thay vì phục vụ dữ liệu của schema mặc định.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/TenantEnforcementTests.cs",
    },
    {
        "feature": "Order",
        "sub": "Tenant Enforcement",
        "layer": "Backend",
        "type": "Integration",
        "name": "AWriteWithoutATenant_CreatesNoOrder",
        "given": "Đếm số bản ghi đơn hàng trước khi thực hiện, request ghi không mang header tenant.",
        "when": "POST /orders với một dòng hợp lệ.",
        "then": "Trả 500 và số bản ghi đơn hàng không thay đổi - không chỉ bị từ chối mà thực sự không ghi gì vào database.",
        "file": "services/orders/tests/Orders.Api.IntegrationTests/TenantEnforcementTests.cs",
    },
]

ROWS += [
    # ---- Frontend / Unit / checkout/Confirmation.test.tsx -------------------------------------
    {
        "feature": "Order",
        "sub": "Storefront - Màn hình xác nhận đơn",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows the order identifier verbatim",
        "given": "Màn hình Confirmation nhận đơn hàng có mã aaaaaaaa-0000-4000-8000-000000000001, tổng 59.25.",
        "when": "Hiển thị màn hình xác nhận.",
        "then": "Mã đơn hiện nguyên vẹn không rút gọn - người mua đọc và trích dẫn được, đối chiếu đúng với đơn trong hệ thống.",
        "file": "frontend/apps/web/tests/checkout/Confirmation.test.tsx",
    },
    {
        "feature": "Order",
        "sub": "Storefront - Màn hình xác nhận đơn",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows the order total in the single Phase 1 currency",
        "given": "Đơn hàng có tổng tiền 59.25.",
        "when": "Hiển thị màn hình xác nhận.",
        "then": "Hiện $59.25 - định dạng tiền tệ duy nhất của giai đoạn 1.",
        "file": "frontend/apps/web/tests/checkout/Confirmation.test.tsx",
    },
    {
        "feature": "Order",
        "sub": "Storefront - Màn hình xác nhận đơn",
        "layer": "Frontend",
        "type": "Unit",
        "name": "tells the shopper their order was placed",
        "given": "Đơn hàng hợp lệ được truyền vào màn hình xác nhận.",
        "when": "Hiển thị màn hình xác nhận.",
        "then": "Có tiêu đề báo đơn đã được đặt thành công.",
        "file": "frontend/apps/web/tests/checkout/Confirmation.test.tsx",
    },
    {
        "feature": "Order",
        "sub": "Storefront - Màn hình xác nhận đơn",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows a nothing-to-show state when there is no order",
        "given": "Người dùng vào thẳng địa chỉ màn hình xác nhận mà chưa đặt hàng, không có dữ liệu đơn.",
        "when": "Hiển thị màn hình xác nhận.",
        "then": "Hiện trạng thái không có đơn gần đây và tuyệt đối không bịa ra mã đơn nào.",
        "file": "frontend/apps/web/tests/checkout/Confirmation.test.tsx",
    },
    # ---- Frontend / Unit / checkout/DoubleSubmit.test.tsx --------------------------------------
    {
        "feature": "Order",
        "sub": "Storefront - Chống đặt hàng trùng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "issues exactly one checkout request when clicked twice in rapid succession",
        "given": "Giỏ có 1 sản phẩm; API giả lập giữ request checkout lại chưa trả lời.",
        "when": "Bấm nút Check out lần đầu, chờ nút bị vô hiệu hóa, rồi bấm tiếp lần hai khi request đầu còn đang chạy.",
        "then": "Chỉ đúng một request checkout được gửi đi (FR-016, SC-008).",
        "file": "frontend/apps/web/tests/checkout/DoubleSubmit.test.tsx",
    },
    {
        "feature": "Order",
        "sub": "Storefront - Chống đặt hàng trùng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "issues one request even when both clicks land before React re-renders",
        "given": "Giỏ có 1 sản phẩm; hai cú click được bắn bằng fireEvent trong cùng một tick, trước khi React kịp render lại trạng thái đang chờ.",
        "when": "Bấm nút Check out hai lần liên tiếp tức thì.",
        "then": "Vẫn chỉ một request được gửi - đúng tình huống đã từng tạo ra 2 đơn cách nhau 6 mili giây trên môi trường container.",
        "file": "frontend/apps/web/tests/checkout/DoubleSubmit.test.tsx",
    },
    {
        "feature": "Order",
        "sub": "Storefront - Chống đặt hàng trùng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "reports the created order exactly once",
        "given": "API giả lập trả về 201 với đơn hàng đã tạo.",
        "when": "Bấm nút Check out một lần.",
        "then": "Callback onCheckedOut được gọi đúng một lần với đúng mã đơn.",
        "file": "frontend/apps/web/tests/checkout/DoubleSubmit.test.tsx",
    },
    {
        "feature": "Order",
        "sub": "Storefront - Chống đặt hàng trùng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows an error and reports no order when checkout fails",
        "given": "API giả lập trả lỗi 502 cho request checkout.",
        "when": "Bấm nút Check out.",
        "then": "Hiện alert lỗi và không báo về đơn hàng nào - không có xác nhận giả, giỏ giữ nguyên để thử lại.",
        "file": "frontend/apps/web/tests/checkout/DoubleSubmit.test.tsx",
    },
    # ---- Frontend / Unit / checkout/EmptyBasketBlocks.test.tsx ---------------------------------
    {
        "feature": "Order",
        "sub": "Storefront - Chặn checkout khi giỏ rỗng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "is not operable",
        "given": "CheckoutButton nhận itemCount = 0 (giỏ rỗng).",
        "when": "Hiển thị nút Check out.",
        "then": "Nút ở trạng thái bị vô hiệu hóa.",
        "file": "frontend/apps/web/tests/checkout/EmptyBasketBlocks.test.tsx",
    },
    {
        "feature": "Order",
        "sub": "Storefront - Chặn checkout khi giỏ rỗng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "sends no checkout request when the shopper tries anyway",
        "given": "Giỏ rỗng, nút bị vô hiệu hóa; test cố tình bỏ qua kiểm tra pointer để click thật vào nút.",
        "when": "Người mua vẫn cố bấm nút Check out.",
        "then": "Không request checkout nào được gửi - tiêu chí là không gửi request, chứ không phải server từ chối (FR-008, SC-004).",
        "file": "frontend/apps/web/tests/checkout/EmptyBasketBlocks.test.tsx",
    },
    {
        "feature": "Order",
        "sub": "Storefront - Chặn checkout khi giỏ rỗng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "becomes operable once the basket holds something",
        "given": "CheckoutButton nhận itemCount = 1.",
        "when": "Bấm nút Check out.",
        "then": "Nút ở trạng thái bật và gửi đúng một request checkout.",
        "file": "frontend/apps/web/tests/checkout/EmptyBasketBlocks.test.tsx",
    },
    # ---- Frontend / E2E / e2e/walkthrough.spec.ts ----------------------------------------------
    {
        "feature": "Order",
        "sub": "E2E - Walkthrough mua hàng (dev server)",
        "layer": "Frontend",
        "type": "E2E",
        "name": "browse, add to basket, check out, and see the confirmation",
        "given": "Toàn bộ stack backend và gateway đang chạy, storefront chạy trên Vite dev server, giỏ đã được dọn sạch trước mỗi test.",
        "when": "Duyệt catalog, thêm 2 Notebook và 1 Apron, xem giỏ, tải lại trang, checkout rồi đọc lại đơn qua gateway.",
        "then": "Giỏ gộp 2 Notebook thành 1 dòng, tổng $59.25 vẫn còn sau khi refresh mà không dùng localStorage/sessionStorage; xác nhận hiện mã đơn đọc lại được đúng tổng 59.25; giỏ rỗng sau đó; mọi request chỉ đi tới gateway; không có lỗi console nào.",
        "file": "frontend/apps/web/e2e/walkthrough.spec.ts",
    },
    {
        "feature": "Order",
        "sub": "E2E - Walkthrough mua hàng (dev server)",
        "layer": "Frontend",
        "type": "E2E",
        "name": "checkout is blocked, and unsent, when the basket is empty",
        "given": "Giỏ hàng đang rỗng, mở trang /basket trên trình duyệt thật.",
        "when": "Cố bấm nút Check out (kể cả ép click).",
        "then": "Nút bị vô hiệu hóa và không request /bff/checkout nào được ghi nhận.",
        "file": "frontend/apps/web/e2e/walkthrough.spec.ts",
    },
    {
        "feature": "Order",
        "sub": "E2E - Walkthrough mua hàng (dev server)",
        "layer": "Frontend",
        "type": "E2E",
        "name": "checking out twice in rapid succession creates exactly one order",
        "given": "Giỏ có 1 Notebook, nút Check out đang bật.",
        "when": "Bắn hai cú click gần như đồng thời vào nút Check out.",
        "then": "Màn hình xác nhận hiện ra và chỉ đúng một request POST /bff/checkout được gửi.",
        "file": "frontend/apps/web/e2e/walkthrough.spec.ts",
    },
    {
        "feature": "Order",
        "sub": "E2E - Walkthrough mua hàng (dev server)",
        "layer": "Frontend",
        "type": "E2E",
        "name": "the whole flow can be completed using only the keyboard",
        "given": "Storefront chạy trên trình duyệt thật, không dùng chuột trong suốt kịch bản.",
        "when": "Dùng Tab và Enter để thêm sản phẩm, chuyển sang trang giỏ và bấm Check out.",
        "then": "Từng điều khiển nhận được focus, viền focus hiển thị rõ (outline khác none) và màn hình xác nhận đơn hiện ra - hoàn tất toàn bộ luồng chỉ bằng bàn phím (FR-017, SC-009).",
        "file": "frontend/apps/web/e2e/walkthrough.spec.ts",
    },
    # ---- Frontend / E2E / demo/order-demo.spec.ts ----------------------------------------------
    {
        "feature": "Order",
        "sub": "E2E - Demo đặt hàng trên stack container",
        "layer": "Frontend",
        "type": "E2E",
        "name": "one order, placed end to end, on the running stack",
        "given": "Stack container đang chạy (storefront cổng 4173, gateway 5300) do script demo dựng lên, giỏ đã được dọn sạch.",
        "when": "Duyệt catalog, thêm 2 Notebook và 1 Apron (chờ từng request ghi xong), checkout, rồi đọc lại đơn qua gateway.",
        "then": "Giỏ gộp 2 Notebook thành 1 dòng với tổng $59.25; xác nhận hiện mã đơn dạng GUID; đơn đọc lại đúng mã và tổng 59.25; giỏ rỗng sau đó; đồng thời lưu 4 ảnh chụp màn hình vào docs/demo và ghi mã đơn cùng tổng tiền ra artifacts/demo.",
        "file": "frontend/apps/web/demo/order-demo.spec.ts",
    },
]

# ==========================================================================================
# PARTY — services/parties/**
# ==========================================================================================

ROWS += [
    {
        "feature": "Party",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Unit",
        "name": "HealthLive_ReturnsOk",
        "given": "Parties API được khởi động bằng WebApplicationFactory.",
        "when": "Gọi GET /health/live.",
        "then": "Trả về HTTP 200 OK, xác nhận tiến trình còn sống.",
        "file": "services/parties/tests/Parties.Api.UnitTests/HealthCheckTests.cs",
    },
    {
        "feature": "Party",
        "sub": "Party API - Đọc đối tác theo Id",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetParty_ReturnsTheParty_WhenItExists",
        "given": "SQL Server thật (Testcontainers) đã seed một party tên hiển thị Ada Lovelace, tenant contoso.",
        "when": "Gọi GET /parties/{id} kèm header tenant.",
        "then": "Trả 200 với đúng Id và DisplayName như đã lưu.",
        "file": "services/parties/tests/Parties.Api.IntegrationTests/PartyEndpointsTests.cs",
    },
    {
        "feature": "Party",
        "sub": "Party API - Đọc đối tác theo Id",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetParty_ReturnsNotFound_WhenNoPartyHasThatId",
        "given": "Database không có party nào được seed, tenant hợp lệ.",
        "when": "Gọi GET /parties/{id} với Id ngẫu nhiên không tồn tại.",
        "then": "Trả 404 Not Found.",
        "file": "services/parties/tests/Parties.Api.IntegrationTests/PartyEndpointsTests.cs",
    },
    {
        "feature": "Party",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_ReturnsOk_WhenDatabaseReachable",
        "given": "Parties API trỏ tới SQL Server thật đang chạy qua Testcontainers.",
        "when": "Gọi GET /health/ready.",
        "then": "Trả 200 OK - readiness phản ánh đúng kết nối database.",
        "file": "services/parties/tests/Parties.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Party",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_ReturnsServiceUnavailable_WhenDatabaseUnreachable",
        "given": "Parties API trỏ tới connection string hợp lệ về cú pháp nhưng không thể kết nối được.",
        "when": "Gọi GET /health/ready.",
        "then": "Trả 503 Service Unavailable - readiness fail-closed chứ không báo healthy giả.",
        "file": "services/parties/tests/Parties.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Party",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_DoesNotFallBackToAnotherServicesDatabase_WhenOwnDatabaseUnreachable",
        "given": "Database riêng của Parties không kết nối được, nhưng cấu hình vẫn cung cấp connection string còn sống của ProductsDb, BasketsDb và OrdersDb.",
        "when": "Gọi GET /health/ready.",
        "then": "Vẫn trả 503 và phản hồi nêu rõ self-database là Unhealthy - không âm thầm dùng database của service khác.",
        "file": "services/parties/tests/Parties.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Party",
        "sub": "Tenant Enforcement",
        "layer": "Backend",
        "type": "Integration",
        "name": "ResolvingTheDbContext_Throws_WhenNoTenantHasBeenResolved",
        "given": "Một DI scope được tạo mà chưa có request HTTP nào phân giải tenant.",
        "when": "Yêu cầu resolve PartiesDbContext từ service provider.",
        "then": "Ném MissingTenantContextException - không có tenant thì không được chạm vào database.",
        "file": "services/parties/tests/Parties.Api.IntegrationTests/TenantEnforcementTests.cs",
    },
    {
        "feature": "Party",
        "sub": "Tenant Enforcement",
        "layer": "Backend",
        "type": "Integration",
        "name": "ARequestWithoutATenant_Fails_RatherThanServingDefaultSchemaData",
        "given": "Request gọi thẳng vào Parties API, không mang header tenant.",
        "when": "Gọi GET /parties/{id}.",
        "then": "Trả 500 - thất bại rõ ràng thay vì phục vụ dữ liệu của schema mặc định.",
        "file": "services/parties/tests/Parties.Api.IntegrationTests/TenantEnforcementTests.cs",
    },
]

# ==========================================================================================
# PRODUCT — services/products/** (Backend) + frontend/apps/web/tests/catalog/** (Frontend)
# ==========================================================================================

ROWS += [
    {
        "feature": "Product",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Unit",
        "name": "HealthLive_ReturnsOk",
        "given": "Products API được khởi động bằng WebApplicationFactory.",
        "when": "Gọi GET /health/live.",
        "then": "Trả về HTTP 200 OK, xác nhận tiến trình còn sống.",
        "file": "services/products/tests/Products.Api.UnitTests/HealthCheckTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Catalog API - Danh sách sản phẩm",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetProducts_ReturnsEveryProduct_WithIdNameAndPrice",
        "given": "SQL Server thật đã seed 2 sản phẩm: Ceramic mug giá 12.50 và Cafetiere giá 34.99, tenant contoso.",
        "when": "Gọi GET /products kèm header tenant.",
        "then": "Trả 200 với đủ 2 sản phẩm, từng sản phẩm đúng cả Id, Name và Price - không trường nào bị rơi hay đổi tên.",
        "file": "services/products/tests/Products.Api.IntegrationTests/CatalogEndpointsTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Catalog API - Danh sách sản phẩm",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetProducts_ReturnsEmptyArray_WhenCatalogIsEmpty",
        "given": "Catalog đã được dọn sạch, không còn sản phẩm nào.",
        "when": "Gọi GET /products.",
        "then": "Trả 200 với mảng rỗng - catalog rỗng là trạng thái hợp lệ, không phải lỗi 404.",
        "file": "services/products/tests/Products.Api.IntegrationTests/CatalogEndpointsTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Catalog Seed - Dữ liệu mẫu theo migration",
        "layer": "Backend",
        "type": "Integration",
        "name": "ApplyingMigrations_SeedsTheCatalog_WithTheThreeKnownProducts",
        "given": "Một database mới hoàn toàn, chỉ chạy migration chứ không seed thủ công.",
        "when": "Gọi GET /products sau khi migration hoàn tất.",
        "then": "Trả 200 và có đủ 3 sản phẩm mẫu với đúng Id, tên và giá như CatalogSeed khai báo - đúng tên và giá mà kịch bản demo và Playwright dựa vào.",
        "file": "services/products/tests/Products.Api.IntegrationTests/CatalogSeedTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Catalog Seed - Dữ liệu mẫu theo migration",
        "layer": "Backend",
        "type": "Integration",
        "name": "ApplyingMigrations_LeavesAPurchasableProduct_WithoutAnyManualSetup",
        "given": "Một database mới chỉ vừa chạy migration, không có thao tác chuẩn bị dữ liệu nào khác.",
        "when": "Gọi GET /products.",
        "then": "Danh sách không rỗng và mọi sản phẩm đều có tên không rỗng, giá lớn hơn 0 - luồng demo chạy được ngay mà không cần nhập liệu tay (FR-018).",
        "file": "services/products/tests/Products.Api.IntegrationTests/CatalogSeedTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Catalog Seed - Dữ liệu mẫu theo migration",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheSeededIdentifiers_AreStableAcrossFreshDatabases",
        "given": "Hai database mới độc lập, cùng chạy migration.",
        "when": "Đọc danh sách sản phẩm từ cả hai và so sánh tập Id.",
        "then": "Tập Id giống hệt nhau - mã sản phẩm mẫu cố định, không sinh mới theo từng môi trường.",
        "file": "services/products/tests/Products.Api.IntegrationTests/CatalogSeedTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_ReturnsOk_WhenDatabaseReachable",
        "given": "Products API trỏ tới SQL Server thật đang chạy qua Testcontainers.",
        "when": "Gọi GET /health/ready.",
        "then": "Trả 200 OK - readiness phản ánh đúng kết nối database.",
        "file": "services/products/tests/Products.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_ReturnsServiceUnavailable_WhenDatabaseUnreachable",
        "given": "Products API trỏ tới connection string hợp lệ về cú pháp nhưng không thể kết nối được.",
        "when": "Gọi GET /health/ready.",
        "then": "Trả 503 Service Unavailable - readiness fail-closed chứ không báo healthy giả.",
        "file": "services/products/tests/Products.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Health & Readiness",
        "layer": "Backend",
        "type": "Integration",
        "name": "HealthReady_DoesNotFallBackToAnotherServicesDatabase_WhenOwnDatabaseUnreachable",
        "given": "Database riêng của Products không kết nối được, nhưng cấu hình vẫn cung cấp connection string còn sống của PartiesDb, BasketsDb và OrdersDb.",
        "when": "Gọi GET /health/ready.",
        "then": "Vẫn trả 503 và phản hồi nêu rõ self-database là Unhealthy - không âm thầm dùng database của service khác.",
        "file": "services/products/tests/Products.Api.IntegrationTests/ReadinessTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Tenant Enforcement",
        "layer": "Backend",
        "type": "Integration",
        "name": "ResolvingTheDbContext_Throws_WhenNoTenantHasBeenResolved",
        "given": "Một DI scope được tạo mà chưa có request HTTP nào phân giải tenant.",
        "when": "Yêu cầu resolve ProductsDbContext từ service provider.",
        "then": "Ném MissingTenantContextException - không tồn tại khoảnh khắc nào DbContext có mặt mà chưa qua kiểm tra tenant.",
        "file": "services/products/tests/Products.Api.IntegrationTests/TenantEnforcementTests.cs",
    },
    {
        "feature": "Product",
        "sub": "Tenant Enforcement",
        "layer": "Backend",
        "type": "Integration",
        "name": "ARequestWithoutATenant_Fails_RatherThanServingDefaultSchemaData",
        "given": "Request gọi thẳng vào Products API với database hoạt động bình thường, nhưng không mang header tenant.",
        "when": "Gọi GET /products.",
        "then": "Trả 500 thay vì 200 kèm catalog của tenant mặc định - đây là bằng chứng cổng chặn tenant thực sự tồn tại.",
        "file": "services/products/tests/Products.Api.IntegrationTests/TenantEnforcementTests.cs",
    },
]

ROWS += [
    # ---- Frontend / Unit / catalog/ProductList.test.tsx ---------------------------------------
    {
        "feature": "Product",
        "sub": "Storefront - Danh sách sản phẩm",
        "layer": "Frontend",
        "type": "Unit",
        "name": "lists every product with its name and price",
        "given": "API giả lập trả 2 sản phẩm: Field Notes Notebook giá 12.5 và Ceramic Pour-Over Set giá 48.",
        "when": "Hiển thị component ProductList.",
        "then": "Cả hai tên sản phẩm và giá đã định dạng $12.50, $48.00 đều hiển thị - dữ liệu lấy từ backend chứ không nhúng trong client.",
        "file": "frontend/apps/web/tests/catalog/ProductList.test.tsx",
    },
    {
        "feature": "Product",
        "sub": "Storefront - Danh sách sản phẩm",
        "layer": "Frontend",
        "type": "Unit",
        "name": "presents the catalog as a list",
        "given": "API giả lập trả 1 sản phẩm Linen Apron.",
        "when": "Hiển thị component ProductList.",
        "then": "Nội dung được đánh dấu là danh sách có nhãn products với đúng 1 listitem - trình đọc màn hình biết được số lượng sản phẩm trước khi đọc.",
        "file": "frontend/apps/web/tests/catalog/ProductList.test.tsx",
    },
    {
        "feature": "Product",
        "sub": "Storefront - Danh sách sản phẩm",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows the empty state when the catalog holds nothing",
        "given": "API giả lập trả danh sách sản phẩm rỗng.",
        "when": "Hiển thị component ProductList.",
        "then": "Hiện thông báo không có sản phẩm và không render danh sách - trang không trông như bị hỏng.",
        "file": "frontend/apps/web/tests/catalog/ProductList.test.tsx",
    },
    {
        "feature": "Product",
        "sub": "Storefront - Danh sách sản phẩm",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows a readable error when the backend fails",
        "given": "API giả lập trả lỗi 502 khi lấy danh sách sản phẩm.",
        "when": "Hiển thị component ProductList.",
        "then": "Hiện alert lỗi kèm nút Try again - không màn hình trắng, không spinner quay mãi (FR-012).",
        "file": "frontend/apps/web/tests/catalog/ProductList.test.tsx",
    },
    {
        "feature": "Product",
        "sub": "Storefront - Danh sách sản phẩm",
        "layer": "Frontend",
        "type": "Unit",
        "name": "requests the catalog from the configured gateway origin",
        "given": "API client được cấu hình baseUrl là origin của gateway.",
        "when": "Hiển thị ProductList và ghi lại các URL thực sự được gọi.",
        "then": "Đúng một request và đúng tới địa chỉ gateway đã cấu hình - không có URL cứng lọt vào component (SC-010).",
        "file": "frontend/apps/web/tests/catalog/ProductList.test.tsx",
    },
    # ---- Frontend / Unit / catalog/EmptyCatalog.test.tsx ---------------------------------------
    {
        "feature": "Product",
        "sub": "Storefront - Trạng thái catalog rỗng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "tells the shopper there is nothing to buy yet",
        "given": "Component EmptyCatalog được hiển thị.",
        "when": "Render màn hình.",
        "then": "Hiện thông báo hiện chưa có sản phẩm nào - trạng thái rỗng được nói rõ chứ không để trang trắng.",
        "file": "frontend/apps/web/tests/catalog/EmptyCatalog.test.tsx",
    },
    {
        "feature": "Product",
        "sub": "Storefront - Trạng thái catalog rỗng",
        "layer": "Frontend",
        "type": "Unit",
        "name": "does not present itself as an error",
        "given": "Component EmptyCatalog được hiển thị.",
        "when": "Render màn hình.",
        "then": "Không có phần tử role alert - catalog rỗng không phải lỗi, không cắt ngang người dùng trình đọc màn hình để báo một chuyện bình thường.",
        "file": "frontend/apps/web/tests/catalog/EmptyCatalog.test.tsx",
    },
    # ---- Frontend / Unit / catalog/CatalogError.test.tsx ---------------------------------------
    {
        "feature": "Product",
        "sub": "Storefront - Trạng thái lỗi catalog",
        "layer": "Frontend",
        "type": "Unit",
        "name": "shows a readable message and announces it",
        "given": "Component ErrorState nhận thông điệp không tải được sản phẩm.",
        "when": "Render màn hình.",
        "then": "Phần tử role alert chứa đúng thông điệp đó - lỗi được công bố cho cả trình đọc màn hình.",
        "file": "frontend/apps/web/tests/catalog/CatalogError.test.tsx",
    },
    {
        "feature": "Product",
        "sub": "Storefront - Trạng thái lỗi catalog",
        "layer": "Frontend",
        "type": "Unit",
        "name": "offers a retry the shopper can operate",
        "given": "Component ErrorState có truyền hàm onRetry.",
        "when": "Bấm nút Try again.",
        "then": "Hàm onRetry được gọi đúng một lần - trang vẫn dùng được sau lỗi.",
        "file": "frontend/apps/web/tests/catalog/CatalogError.test.tsx",
    },
    {
        "feature": "Product",
        "sub": "Storefront - Trạng thái lỗi catalog",
        "layer": "Frontend",
        "type": "Unit",
        "name": "reaches and fires retry by keyboard alone",
        "given": "Component ErrorState có nút Try again.",
        "when": "Nhấn Tab để đưa focus vào nút rồi nhấn Enter.",
        "then": "Nút nhận focus và onRetry được gọi - đường phục hồi sau lỗi không chỉ dành cho chuột (FR-017, SC-009).",
        "file": "frontend/apps/web/tests/catalog/CatalogError.test.tsx",
    },
    {
        "feature": "Product",
        "sub": "Storefront - Trạng thái lỗi catalog",
        "layer": "Frontend",
        "type": "Unit",
        "name": "omits the retry control when no retry is possible",
        "given": "Component ErrorState không được truyền hàm onRetry.",
        "when": "Render màn hình.",
        "then": "Không hiển thị nút nào - không đưa ra điều khiển vô tác dụng.",
        "file": "frontend/apps/web/tests/catalog/CatalogError.test.tsx",
    },
]

# ==========================================================================================
# COMMON — services/bff/**, shared/Tenancy.UnitTests, tests/** (convention),
#          frontend shared / accessibility / app
# ==========================================================================================

ROWS += [
    # ---- BFF / Unit / DownstreamServiceClientOptionsTests -----------------------------------
    {
        "feature": "Common",
        "sub": "BFF - Cấu hình downstream client",
        "layer": "Backend",
        "type": "Unit",
        "name": "Validation_Fails_WhenBaseUrlIsMissing",
        "given": "Cấu hình không khai báo Services:ProductsApi:BaseUrl.",
        "when": "Đọc options của downstream client ProductsApi.",
        "then": "Ném OptionsValidationException và thông báo nêu đúng tên khóa cấu hình bị thiếu - lỗi cấu hình chặn ngay khi khởi động chứ không lỗi từng request.",
        "file": "services/bff/tests/Bff.Api.UnitTests/DownstreamServiceClientOptionsTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Cấu hình downstream client",
        "layer": "Backend",
        "type": "Unit",
        "name": "Validation_Fails_WhenBaseUrlIsNotAnAbsoluteUri",
        "given": "BaseUrl được khai báo là đường dẫn tương đối /products.",
        "when": "Đọc options của downstream client.",
        "then": "Ném OptionsValidationException - BaseUrl phải là URI tuyệt đối.",
        "file": "services/bff/tests/Bff.Api.UnitTests/DownstreamServiceClientOptionsTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Cấu hình downstream client",
        "layer": "Backend",
        "type": "Unit",
        "name": "Validation_Succeeds_WhenBaseUrlIsAnAbsoluteUri",
        "given": "BaseUrl được khai báo hợp lệ là http://products-api:8080.",
        "when": "Đọc options của downstream client.",
        "then": "Options nạp thành công với đúng BaseUrl và ServiceName là ProductsApi.",
        "file": "services/bff/tests/Bff.Api.UnitTests/DownstreamServiceClientOptionsTests.cs",
    },
    # ---- BFF / Unit / ResponseMappingTests ---------------------------------------------------
    {
        "feature": "Common",
        "sub": "BFF - Ánh xạ dữ liệu phản hồi",
        "layer": "Backend",
        "type": "Unit",
        "name": "ProductSummary_CarriesEveryFieldFromTheDownstreamProduct",
        "given": "Một ProductResource từ downstream: Ceramic mug, giá 12.50.",
        "when": "Ánh xạ sang ProductSummary bằng ProductsEndpoints.ToSummary.",
        "then": "Cả Id, Name và Price đều được giữ nguyên.",
        "file": "services/bff/tests/Bff.Api.UnitTests/ResponseMappingTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Ánh xạ dữ liệu phản hồi",
        "layer": "Backend",
        "type": "Unit",
        "name": "BasketItem_JoinsTheProductName_AndPassesEveryOtherFieldThrough",
        "given": "Một dòng giỏ hàng (2 x 12.50, thành tiền 25.00) và từ điển tên sản phẩm tra được Field Notes Notebook.",
        "when": "Ánh xạ sang BasketItem bằng BasketsEndpoints.ToItem.",
        "then": "Tên sản phẩm được ghép vào, các trường còn lại giữ nguyên; thành tiền được truyền thẳng chứ không tính lại - phép tính tiền thuộc về baskets service.",
        "file": "services/bff/tests/Bff.Api.UnitTests/ResponseMappingTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Ánh xạ dữ liệu phản hồi",
        "layer": "Backend",
        "type": "Unit",
        "name": "BasketItem_SurvivesAProductMissingFromTheCatalog",
        "given": "Một dòng giỏ hàng có sản phẩm không còn trong catalog (từ điển tên rỗng).",
        "when": "Ánh xạ sang BasketItem.",
        "then": "Dòng vẫn tồn tại với thành tiền 9.99 và tên không rỗng - không tự ý bỏ dòng mà khách đã chọn và đang bị tính tiền.",
        "file": "services/bff/tests/Bff.Api.UnitTests/ResponseMappingTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Ánh xạ dữ liệu phản hồi",
        "layer": "Backend",
        "type": "Unit",
        "name": "OrderResponse_CarriesEveryFieldFromTheDownstreamOrder",
        "given": "Một OrderResource từ downstream: đặt lúc 2026-08-15 09:30 UTC, tổng 47.49.",
        "when": "Ánh xạ sang OrderResponse bằng OrdersEndpoints.ToResponse.",
        "then": "Id, thời điểm đặt và tổng tiền đều giữ nguyên, không bị hoán vị giữa các trường cùng kiểu.",
        "file": "services/bff/tests/Bff.Api.UnitTests/ResponseMappingTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Ánh xạ dữ liệu phản hồi",
        "layer": "Backend",
        "type": "Unit",
        "name": "OrderResponse_PreservesTheUtcKindOfThePlacedTimestamp",
        "given": "Thời điểm đặt hàng mang DateTimeKind.Utc.",
        "when": "Ánh xạ sang OrderResponse.",
        "then": "Kind của mốc thời gian vẫn là Utc - giao diện không hiển thị sai múi giờ mà không có dấu hiệu nào.",
        "file": "services/bff/tests/Bff.Api.UnitTests/ResponseMappingTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Ánh xạ dữ liệu phản hồi",
        "layer": "Backend",
        "type": "Unit",
        "name": "PartyResponse_CarriesEveryFieldFromTheDownstreamParty",
        "given": "Một PartyResource từ downstream: Ada Lovelace.",
        "when": "Ánh xạ sang PartyResponse bằng PartiesEndpoints.ToResponse.",
        "then": "Id và DisplayName đều được giữ nguyên.",
        "file": "services/bff/tests/Bff.Api.UnitTests/ResponseMappingTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Ánh xạ dữ liệu phản hồi",
        "layer": "Backend",
        "type": "Unit",
        "name": "ProductSummary_PreservesPricePrecisionExactly",
        "given": "Giá sản phẩm ở các mức 0.01, 12.50 và 999999999999.99 (Theory, 3 bộ dữ liệu).",
        "when": "Ánh xạ sang ProductSummary rồi in ra chuỗi.",
        "then": "Giá trị và cả số 0 ở phần thập phân giữ nguyên - không bị làm tròn do chuyển decimal sang double.",
        "file": "services/bff/tests/Bff.Api.UnitTests/ResponseMappingTests.cs",
    },
    # ---- BFF / Integration / BasketFlowTests -------------------------------------------------
    {
        "feature": "Common",
        "sub": "BFF - Luồng giỏ hàng cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetBasket_ReturnsAnEmptyBasket_ForAShopperWhoHasAddedNothing",
        "given": "BFF chạy cùng Products API và Baskets API thật, khách chưa thêm gì.",
        "when": "Gọi GET /bff/basket.",
        "then": "Trả 200 với giỏ rỗng và tổng tiền 0.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/BasketFlowTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Luồng giỏ hàng cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "AddItem_ReturnsTheBasket_WithTheProductsNameAndResolvedPrice",
        "given": "BFF chạy cùng Products API (đã seed) và Baskets API thật.",
        "when": "POST /bff/basket/items với productId của Notebook và số lượng 1, không kèm giá.",
        "then": "Trả 200; dòng giỏ có tên Field Notes Notebook, đơn giá 12.50 do BFF tra từ catalog, thành tiền và tổng đều 12.50.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/BasketFlowTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Luồng giỏ hàng cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "AddItem_IgnoresAPriceSuppliedByTheClient",
        "given": "Client cố tình gửi kèm unitPrice = 0.01 trong khi giá thật của sản phẩm là 12.50.",
        "when": "POST /bff/basket/items.",
        "then": "Dòng giỏ mang đơn giá 12.50 lấy từ catalog - client không thể tự đặt giá cho mình.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/BasketFlowTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Luồng giỏ hàng cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "AddItem_MergesIntoTheExistingLine_WhenTheSameProductIsAddedAgain",
        "given": "BFF chạy cùng các service thật.",
        "when": "POST /bff/basket/items hai lần cho cùng Notebook, mỗi lần số lượng 1, rồi đọc lại giỏ.",
        "then": "Giỏ chỉ có 1 dòng số lượng 2, tổng tiền 25.00.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/BasketFlowTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Luồng giỏ hàng cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "AddItem_ReturnsNotFound_WhenNoSuchProductExists",
        "given": "productId là một Guid ngẫu nhiên không có trong catalog.",
        "when": "POST /bff/basket/items.",
        "then": "Trả 404 Not Found chứ không phải 502 - downstream trả lời đúng, chỉ là request sai.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/BasketFlowTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Luồng giỏ hàng cho storefront",
        "layer": "Backend",
        "type": "Integration",
        "name": "AddItem_Rejects_AQuantityBelowOne",
        "given": "Số lượng không hợp lệ là 0 và -2 (Theory, 2 bộ dữ liệu).",
        "when": "POST /bff/basket/items.",
        "then": "Trả 400 Bad Request.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/BasketFlowTests.cs",
    },
    # ---- BFF / Integration / BasketsRouteTests ------------------------------------------------
    {
        "feature": "Common",
        "sub": "BFF - Route proxy tới Baskets",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetBasket_ReturnsShapedBasketFromTheBasketsService",
        "given": "Baskets API thật đọc từ database thật, đã có giỏ rỗng của khách bff-route-shopper.",
        "when": "Gọi GET /bff/baskets/{id} qua BFF.",
        "then": "Trả 200 với đúng Id, CustomerRef và danh sách dòng rỗng - giỏ rỗng không cần gọi sang Products.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/BasketsRouteTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Route proxy tới Baskets",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetBasket_ReturnsNotFound_WhenTheBasketsServiceHasNoSuchBasket",
        "given": "Baskets API không có giỏ nào với Id được hỏi.",
        "when": "Gọi GET /bff/baskets/{id} qua BFF.",
        "then": "Trả 404 chứ không phải 502 hay 200 rỗng - phân biệt giữa downstream trả lời đúng và downstream hỏng.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/BasketsRouteTests.cs",
    },
    # ---- BFF / Integration / CheckoutTests ----------------------------------------------------
    {
        "feature": "Common",
        "sub": "BFF - Checkout đầu-cuối qua 3 service",
        "layer": "Backend",
        "type": "Integration",
        "name": "Checkout_CreatesAnOrder_ForWhatIsInTheBasket",
        "given": "Products, Baskets và Orders API thật cùng chạy sau BFF; giỏ có 2 x Notebook và 1 x Apron.",
        "when": "POST /bff/checkout.",
        "then": "Trả 201 Created với mã đơn khác rỗng và tổng tiền 59.25 - con số đi qua đủ 3 service.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/CheckoutTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Checkout đầu-cuối qua 3 service",
        "layer": "Backend",
        "type": "Integration",
        "name": "Checkout_ReturnsAReference_ThatReadsBackAsTheSameOrder",
        "given": "Giỏ có 1 x Notebook và vừa checkout thành công.",
        "when": "Dùng mã đơn trong xác nhận gọi GET /bff/orders/{id}.",
        "then": "Đơn đọc lại trùng Id và tổng tiền - mã đơn trên màn hình xác nhận đúng là đơn có thật trong hệ thống.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/CheckoutTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Checkout đầu-cuối qua 3 service",
        "layer": "Backend",
        "type": "Integration",
        "name": "Checkout_EmptiesTheBasket",
        "given": "Giỏ có 1 x Notebook.",
        "when": "POST /bff/checkout rồi gọi GET /bff/basket.",
        "then": "Giỏ rỗng và tổng tiền 0 sau khi đặt hàng thành công.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/CheckoutTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Checkout đầu-cuối qua 3 service",
        "layer": "Backend",
        "type": "Integration",
        "name": "Checkout_ReturnsConflict_WhenTheBasketIsEmpty",
        "given": "Khách chưa thêm sản phẩm nào, giỏ đang rỗng.",
        "when": "POST /bff/checkout.",
        "then": "Trả 409 Conflict - server tự từ chối chứ không dựa vào việc giao diện đã chặn.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/CheckoutTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Checkout đầu-cuối qua 3 service",
        "layer": "Backend",
        "type": "Integration",
        "name": "Checkout_CreatesExactlyOneOrder_WhenAttemptedTwice",
        "given": "Giỏ có 1 x Notebook.",
        "when": "Gọi POST /bff/checkout hai lần liên tiếp.",
        "then": "Lần đầu trả 201, lần hai trả 409 - bấm đặt hàng hai lần không tạo ra hai đơn.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/CheckoutTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Checkout đầu-cuối qua 3 service",
        "layer": "Backend",
        "type": "Integration",
        "name": "Checkout_OrdersOnlyTheCallersOwnBasket",
        "given": "Hai khách khác nhau, mỗi người có giỏ riêng: một người có Notebook, người kia có Apron.",
        "when": "Khách thứ nhất checkout.",
        "then": "Đơn của khách thứ nhất tổng 12.50 và giỏ của khách thứ hai vẫn nguyên vẹn 1 dòng.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/CheckoutTests.cs",
    },
    # ---- BFF / Integration / DownstreamUnavailableTests ---------------------------------------
    {
        "feature": "Common",
        "sub": "BFF - Xử lý downstream lỗi",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetProducts_ReturnsBadGateway_WhenTheProductsServiceIsUnreachable",
        "given": "BFF chạy với transport tới Products API luôn lỗi kết nối, không có host downstream nào được khởi động.",
        "when": "Gọi GET /bff/products và đo thời gian.",
        "then": "Trả 502 Bad Gateway trong dưới 5 giây (SC-003).",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/DownstreamUnavailableTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Xử lý downstream lỗi",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetProducts_ReturnsGatewayTimeout_WhenTheProductsServiceNeverAnswers",
        "given": "Products API được giả lập ở trạng thái không bao giờ trả lời (chậm chứ không chết hẳn).",
        "when": "Gọi GET /bff/products và đo thời gian.",
        "then": "Trả 504 Gateway Timeout trong dưới 5 giây - phân biệt được dịch vụ đã chết với dịch vụ đang quá tải.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/DownstreamUnavailableTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Xử lý downstream lỗi",
        "layer": "Backend",
        "type": "Integration",
        "name": "ADownstreamFailure_ReturnsProblemDetailsCarryingTheCorrelationId",
        "given": "Transport tới Products API luôn lỗi.",
        "when": "Gọi GET /bff/products và đọc nội dung lỗi.",
        "then": "Content-Type là application/problem+json với type, title, status 502 và correlationId trùng đúng header X-Correlation-Id của phản hồi.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/DownstreamUnavailableTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Xử lý downstream lỗi",
        "layer": "Backend",
        "type": "Integration",
        "name": "ADownstreamFailure_NamesTheLogicalServiceOnly_NeverItsAddress",
        "given": "Transport tới Products API luôn lỗi.",
        "when": "Gọi GET /bff/products và đọc body lỗi.",
        "then": "Body có tên logic ProductsApi nhưng không chứa host, scheme http://, tên thư viện Polly hay stack trace - chẩn đoán được mà không lộ topology.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/DownstreamUnavailableTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Xử lý downstream lỗi",
        "layer": "Backend",
        "type": "Integration",
        "name": "EveryRoute_FailsAsAProblemDetails_WhenItsDownstreamIsUnreachable",
        "given": "Lần lượt làm hỏng BasketsApi, OrdersApi và PartiesApi (Theory, 3 bộ dữ liệu).",
        "when": "Gọi route tương ứng của từng service.",
        "then": "Mọi route đều trả 502 với Content-Type application/problem+json - không route nào rơi vào lỗi 500 trần.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/DownstreamUnavailableTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Xử lý downstream lỗi",
        "layer": "Backend",
        "type": "Integration",
        "name": "ADownstreamFailure_IsBoundedAndStructured_AgainstARealUnreachableHost",
        "given": "BFF trỏ tới host thật không thể phân giải (products-service.invalid), dùng transport thật không giả lập.",
        "when": "Gọi GET /bff/products và đo thời gian.",
        "then": "Trả 502 hoặc 504 (tùy tốc độ phân giải DNS) dưới 5 giây, dạng problem+json, detail nêu ProductsApi và có correlationId.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/DownstreamUnavailableTests.cs",
    },
    # ---- BFF / Integration / GeneratedContractTests -------------------------------------------
    {
        "feature": "Common",
        "sub": "BFF - Hợp đồng OpenAPI sinh tự động",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheDocument_DescribesEveryClientFacingRoute",
        "given": "BFF chạy ở môi trường Development, nơi tài liệu OpenAPI được publish.",
        "when": "Tải /openapi/v1.json và đọc danh sách paths.",
        "then": "Tài liệu mô tả đủ 4 route /bff/products, /bff/baskets/{basketId}, /bff/orders/{orderId}, /bff/parties/{partyId}.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/GeneratedContractTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Hợp đồng OpenAPI sinh tự động",
        "layer": "Backend",
        "type": "Integration",
        "name": "EveryRoute_DeclaresItsDownstreamFailureResponses",
        "given": "Kiểm thử lần lượt 4 route client-facing (Theory, 4 bộ dữ liệu).",
        "when": "Đọc danh sách mã phản hồi khai báo cho từng route.",
        "then": "Mỗi route đều khai báo 200, 502 và 504 - client sinh từ tài liệu này biết được các tình huống lỗi.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/GeneratedContractTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Hợp đồng OpenAPI sinh tự động",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheByIdRoutes_DeclareNotFound",
        "given": "Kiểm thử 3 route theo Id: baskets, orders, parties (Theory, 3 bộ dữ liệu).",
        "when": "Đọc danh sách mã phản hồi khai báo cho từng route.",
        "then": "Mỗi route đều khai báo 404 - client phân nhánh được trường hợp không tìm thấy thay vì coi đó là lỗi.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/GeneratedContractTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Hợp đồng OpenAPI sinh tự động",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheProductListingSchema_MatchesTheHandAuthoredContract",
        "given": "Tài liệu OpenAPI sinh tự động và bản hợp đồng viết tay bff-openapi.yaml.",
        "when": "Đối chiếu schema ProductListResponse và ProductSummary.",
        "then": "ProductListResponse có thuộc tính items; ProductSummary có đủ id, name, price.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/GeneratedContractTests.cs",
    },
    # ---- BFF / Integration / OrdersRouteTests / PartiesRouteTests / ProductsRouteTests --------
    {
        "feature": "Common",
        "sub": "BFF - Route proxy tới Orders",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetOrder_ReturnsShapedOrderFromTheOrdersService",
        "given": "Orders API thật đọc database thật, đã có đơn đặt lúc 2026-08-15 09:30 UTC tổng 47.49.",
        "when": "Gọi GET /bff/orders/{id} qua BFF.",
        "then": "Trả 200 với đúng Id, thời điểm đặt và tổng tiền - SPA không phải gọi thẳng orders service.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/OrdersRouteTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Route proxy tới Orders",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetOrder_ReturnsNotFound_WhenTheOrdersServiceHasNoSuchOrder",
        "given": "Orders API không có đơn nào với Id được hỏi.",
        "when": "Gọi GET /bff/orders/{id} qua BFF.",
        "then": "Trả 404 Not Found.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/OrdersRouteTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Route proxy tới Parties",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetParty_ReturnsShapedPartyFromThePartiesService",
        "given": "Parties API thật đọc database thật, đã có party Ada Lovelace.",
        "when": "Gọi GET /bff/parties/{id} qua BFF.",
        "then": "Trả 200 với đúng Id và DisplayName.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/PartiesRouteTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Route proxy tới Parties",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetParty_ReturnsNotFound_WhenThePartiesServiceHasNoSuchParty",
        "given": "Parties API không có party nào với Id được hỏi.",
        "when": "Gọi GET /bff/parties/{id} qua BFF.",
        "then": "Trả 404 Not Found.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/PartiesRouteTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Route proxy tới Products",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetProducts_ReturnsShapedListingFromTheProductsService",
        "given": "Products API thật đã seed Ceramic mug 12.50 và Cafetiere 34.99.",
        "when": "Gọi GET /bff/products qua BFF.",
        "then": "Trả 200 với 2 sản phẩm, từng sản phẩm đúng Id, Name, Price theo schema ProductSummary.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/ProductsRouteTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Route proxy tới Products",
        "layer": "Backend",
        "type": "Integration",
        "name": "GetProducts_ReturnsEmptyItemsEnvelope_WhenTheCatalogIsEmpty",
        "given": "Catalog của Products API đã được dọn sạch.",
        "when": "Gọi GET /bff/products qua BFF.",
        "then": "Trả 200 với đối tượng bọc ngoài có items rỗng - SPA luôn đọc được items mà không phải phân nhánh theo hình dạng phản hồi.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/ProductsRouteTests.cs",
    },
    # ---- BFF / Integration / SubjectPropagationTests & TenantPropagationTests -----------------
    {
        "feature": "Common",
        "sub": "BFF - Truyền định danh xuống domain service",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheBffsOutboundCall_CarriesTheSubjectTheBffReceived",
        "given": "BFF nhận request có header tenant contoso và subject phase1-stub-user; một handler ghi lại request đi ra tới Products API.",
        "when": "Gọi GET /bff/products.",
        "then": "Request đi ra mang đúng subject mà BFF đã nhận - typed HttpClient không tự truyền header nên đây là bằng chứng handler truyền tiếp hoạt động.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/SubjectPropagationTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Truyền định danh xuống domain service",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheBffsOutboundCall_CarriesNoSubject_WhenTheBffItselfHasNone",
        "given": "Request vào BFF chỉ có tenant, không có subject (gateway bị bỏ qua).",
        "when": "Gọi GET /bff/products.",
        "then": "Mọi request đi ra đều không có header subject - BFF chỉ chuyển tiếp chứ không tự bịa ra người gọi mặc định.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/SubjectPropagationTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Truyền định danh xuống domain service",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheBffsOutboundCall_CarriesTheTenantTheBffReceived",
        "given": "BFF nhận request có header tenant contoso; handler ghi lại request đi ra tới Products API.",
        "when": "Gọi GET /bff/products.",
        "then": "Request đi ra mang đúng tenant contoso - chuỗi truyền tenant không đứt ở BFF.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/TenantPropagationTests.cs",
    },
    {
        "feature": "Common",
        "sub": "BFF - Truyền định danh xuống domain service",
        "layer": "Backend",
        "type": "Integration",
        "name": "TheBffsOutboundCall_CarriesNoTenant_WhenTheBffItselfHasNone",
        "given": "Request vào BFF không mang header tenant nào.",
        "when": "Gọi GET /bff/products, downstream từ chối nên pipeline resilience thực hiện retry.",
        "then": "Tất cả các lần gọi ra, kể cả các lần retry, đều không có header tenant - không có fallback bịa tenant ở lần thử lại.",
        "file": "services/bff/tests/Bff.Api.IntegrationTests/TenantPropagationTests.cs",
    },
]

ROWS += [
    # ---- Shared Tenancy / Unit / CallerContextTests -------------------------------------------
    {
        "feature": "Common",
        "sub": "Tenancy - CallerContext",
        "layer": "Backend",
        "type": "Unit",
        "name": "RequireSubjectId_ReturnsTheResolvedSubject_WhenOneHasBeenSet",
        "given": "CallerContext đã được gán SubjectId là phase1-stub-user.",
        "when": "Gọi RequireSubjectId().",
        "then": "Trả về đúng subject đã phân giải.",
        "file": "shared/Tenancy.UnitTests/CallerContextTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - CallerContext",
        "layer": "Backend",
        "type": "Unit",
        "name": "CallerContext_IsUnresolved_BeforeAnythingSetsIt",
        "given": "Một CallerContext vừa khởi tạo, chưa ai gán giá trị.",
        "when": "Đọc thuộc tính SubjectId.",
        "then": "Giá trị là null - trạng thái mặc định là Chưa phân giải.",
        "file": "shared/Tenancy.UnitTests/CallerContextTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - CallerContext",
        "layer": "Backend",
        "type": "Unit",
        "name": "RequireSubjectId_Throws_WhenNoSubjectHasBeenResolved",
        "given": "Một CallerContext chưa được phân giải.",
        "when": "Gọi RequireSubjectId().",
        "then": "Ném MissingCallerContextException.",
        "file": "shared/Tenancy.UnitTests/CallerContextTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - CallerContext",
        "layer": "Backend",
        "type": "Unit",
        "name": "RequireSubjectId_Throws_WhenTheResolvedSubjectIsBlank",
        "given": "SubjectId được gán chuỗi rỗng, khoảng trắng hoặc tab (Theory, 3 bộ dữ liệu).",
        "when": "Gọi RequireSubjectId().",
        "then": "Ném MissingCallerContextException - subject rỗng là chưa phân giải, không phải một người mua tên rỗng dùng chung giỏ hàng.",
        "file": "shared/Tenancy.UnitTests/CallerContextTests.cs",
    },
    # ---- Shared Tenancy / Unit / CallerContextMiddlewareTests ---------------------------------
    {
        "feature": "Common",
        "sub": "Tenancy - CallerContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_ResolvesTheCallerContext_FromTheInboundHeader",
        "given": "Request mang header subject với giá trị phase1-stub-user.",
        "when": "Chạy CallerContextMiddleware.",
        "then": "CallerContext phân giải đúng subject từ header, không tự suy ra hay đặt mặc định.",
        "file": "shared/Tenancy.UnitTests/CallerContextMiddlewareTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - CallerContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_LeavesTheCallerContextUnresolved_WhenTheHeaderIsAbsentOrEmpty",
        "given": "Header subject vắng mặt, rỗng hoặc chỉ có khoảng trắng (Theory, 3 bộ dữ liệu).",
        "when": "Chạy middleware rồi gọi RequireSubjectId().",
        "then": "Ném MissingCallerContextException - vắng và rỗng đều là chưa phân giải.",
        "file": "shared/Tenancy.UnitTests/CallerContextMiddlewareTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - CallerContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_PushesTheResolvedSubjectIntoTheLoggingScope",
        "given": "Request mang header subject phase1-stub-user, dùng logger ghi lại các scope.",
        "when": "Chạy middleware.",
        "then": "Có đúng một logging scope chứa khóa SubjectId với giá trị tương ứng - mọi dòng log của request đều truy được người gọi.",
        "file": "shared/Tenancy.UnitTests/CallerContextMiddlewareTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - CallerContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_PushesNoSubjectScope_WhenTheRequestIsUnresolved",
        "given": "Request không mang header subject.",
        "when": "Chạy middleware.",
        "then": "Không mở logging scope nào - không ghi SubjectId rỗng như thể có người gọi.",
        "file": "shared/Tenancy.UnitTests/CallerContextMiddlewareTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - CallerContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_AlwaysCallsTheRestOfThePipeline",
        "given": "Hai trường hợp: có header subject và không có header (Theory, 2 bộ dữ liệu).",
        "when": "Chạy middleware.",
        "then": "Pipeline phía sau luôn được gọi - health probe không đi qua gateway vẫn hoạt động, việc bắt buộc có người gọi thuộc về từng route.",
        "file": "shared/Tenancy.UnitTests/CallerContextMiddlewareTests.cs",
    },
    # ---- Shared Tenancy / Unit / TenantContextTests -------------------------------------------
    {
        "feature": "Common",
        "sub": "Tenancy - TenantContext",
        "layer": "Backend",
        "type": "Unit",
        "name": "RequireTenantId_ReturnsTheResolvedTenant_WhenOneHasBeenSet",
        "given": "TenantContext đã được gán TenantId là acme.",
        "when": "Gọi RequireTenantId().",
        "then": "Trả về đúng tenant đã phân giải.",
        "file": "shared/Tenancy.UnitTests/TenantContextTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - TenantContext",
        "layer": "Backend",
        "type": "Unit",
        "name": "TenantContext_IsUnresolved_BeforeAnythingSetsIt",
        "given": "Một TenantContext vừa khởi tạo, chưa ai gán giá trị.",
        "when": "Đọc thuộc tính TenantId.",
        "then": "Giá trị là null - chỉ có hai trạng thái Đã phân giải và Chưa phân giải.",
        "file": "shared/Tenancy.UnitTests/TenantContextTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - TenantContext",
        "layer": "Backend",
        "type": "Unit",
        "name": "RequireTenantId_Throws_WhenNoTenantHasBeenResolved",
        "given": "Một TenantContext chưa được phân giải.",
        "when": "Gọi RequireTenantId().",
        "then": "Ném MissingTenantContextException - không tồn tại trạng thái thứ ba là phân giải về tenant mặc định.",
        "file": "shared/Tenancy.UnitTests/TenantContextTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - TenantContext",
        "layer": "Backend",
        "type": "Unit",
        "name": "RequireTenantId_Throws_WhenTheResolvedTenantIsBlank",
        "given": "TenantId được gán chuỗi rỗng, khoảng trắng hoặc tab (Theory, 3 bộ dữ liệu).",
        "when": "Gọi RequireTenantId().",
        "then": "Ném MissingTenantContextException - header X-Tenant-Id rỗng không lọt qua được cổng chặn để chạm tới database.",
        "file": "shared/Tenancy.UnitTests/TenantContextTests.cs",
    },
    # ---- Shared Tenancy / Unit / TenantContextMiddlewareTests ---------------------------------
    {
        "feature": "Common",
        "sub": "Tenancy - TenantContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_ResolvesTheTenantContext_FromTheInboundHeader",
        "given": "Request mang header tenant với giá trị acme.",
        "when": "Chạy TenantContextMiddleware.",
        "then": "TenantContext phân giải đúng tenant từ header, không tự suy lại hay đặt mặc định.",
        "file": "shared/Tenancy.UnitTests/TenantContextMiddlewareTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - TenantContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_LeavesTheTenantContextUnresolved_WhenTheHeaderIsAbsentOrEmpty",
        "given": "Header tenant vắng mặt, rỗng hoặc chỉ có khoảng trắng (Theory, 3 bộ dữ liệu).",
        "when": "Chạy middleware rồi gọi RequireTenantId().",
        "then": "Ném MissingTenantContextException - không bao giờ có tenant dự phòng.",
        "file": "shared/Tenancy.UnitTests/TenantContextMiddlewareTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - TenantContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_PushesTheResolvedTenantIntoTheLoggingScope",
        "given": "Request mang header tenant acme, dùng logger ghi lại các scope.",
        "when": "Chạy middleware.",
        "then": "Có đúng một logging scope chứa khóa TenantId với giá trị acme - tenant xuất hiện trên mọi dòng log mà không cần cấu hình riêng cho từng service.",
        "file": "shared/Tenancy.UnitTests/TenantContextMiddlewareTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - TenantContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_PushesNoTenantScope_WhenTheRequestIsUnresolved",
        "given": "Request không mang header tenant.",
        "when": "Chạy middleware.",
        "then": "Không mở logging scope nào - sự vắng mặt chính là tín hiệu, không ghi TenantId rỗng.",
        "file": "shared/Tenancy.UnitTests/TenantContextMiddlewareTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Tenancy - TenantContextMiddleware",
        "layer": "Backend",
        "type": "Unit",
        "name": "InvokeAsync_AlwaysCallsTheRestOfThePipeline",
        "given": "Hai trường hợp: có header tenant acme và không có header (Theory, 2 bộ dữ liệu).",
        "when": "Chạy middleware.",
        "then": "Pipeline phía sau luôn được gọi trong cả hai trường hợp.",
        "file": "shared/Tenancy.UnitTests/TenantContextMiddlewareTests.cs",
    },
    # ---- Convention / ContainerConventionTests ------------------------------------------------
    {
        "feature": "Common",
        "sub": "Convention - Dockerfile & shared project",
        "layer": "Backend",
        "type": "Unit",
        "name": "EveryServiceImage_ReceivesEverySharedProject_ItCompilesAgainst",
        "given": "Toàn bộ Dockerfile và file .csproj đã commit trong repo.",
        "when": "Quét đối chiếu các shared project mà mỗi service tham chiếu với các thư mục Dockerfile thực sự COPY vào.",
        "then": "Không có vi phạm nào - mọi image build được từ một bản checkout sạch (FR-014).",
        "file": "tests/ContainerConventionTests/DockerfileSharedProjectTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Dockerfile & shared project",
        "layer": "Backend",
        "type": "Unit",
        "name": "TheScan_Examined_EveryService",
        "given": "Bộ quét Dockerfile chạy trên thư mục gốc repo.",
        "when": "Đọc danh sách service và Dockerfile mà bộ quét đã thực sự xem xét.",
        "then": "Đúng 6 service baskets, bff, gateway, orders, parties, products và đủ 6 Dockerfile - bộ quét không trỏ nhầm thư mục rồi báo sạch.",
        "file": "tests/ContainerConventionTests/DockerfileSharedProjectTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Dockerfile & shared project",
        "layer": "Backend",
        "type": "Unit",
        "name": "TheScan_Observed_SharedProjectReferences",
        "given": "Bộ quét Dockerfile chạy trên thư mục gốc repo.",
        "when": "Đọc danh sách tham chiếu shared project mà bộ quét ghi nhận được.",
        "then": "Danh sách không rỗng và mỗi service đều có ít nhất một tham chiếu - bộ quét không bị mù do regex hỏng.",
        "file": "tests/ContainerConventionTests/DockerfileSharedProjectTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Dockerfile & shared project",
        "layer": "Backend",
        "type": "Unit",
        "name": "ServicesThatUseTheTenancyLibrary_CopyIt",
        "given": "Năm service dùng thư viện Tenancy: baskets, bff, orders, parties, products (Theory, 5 bộ dữ liệu).",
        "when": "Quét Dockerfile của từng service.",
        "then": "Không service nào bị đánh dấu thiếu COPY thư viện Tenancy - đúng lỗi hồi quy đã từng làm 5/6 image không build được.",
        "file": "tests/ContainerConventionTests/DockerfileSharedProjectTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Dockerfile & shared project",
        "layer": "Backend",
        "type": "Unit",
        "name": "TheGateway_DoesNotReferenceTheTenancyLibrary",
        "given": "File Gateway.Api.csproj đã commit trong repo.",
        "when": "Đọc danh sách shared project mà gateway tham chiếu.",
        "then": "Không có Tenancy - gateway sinh ra header tenant và subject từ claim chứ không đọc chúng, nên ngoại lệ này là hợp lệ.",
        "file": "tests/ContainerConventionTests/DockerfileSharedProjectTests.cs",
    },
    # ---- Convention / ConnectionStringIsolationTests ------------------------------------------
    {
        "feature": "Common",
        "sub": "Convention - Cô lập database giữa các service",
        "layer": "Backend",
        "type": "Unit",
        "name": "NoServiceConfiguration_NamesAnotherServicesDatabase",
        "given": "Toàn bộ file appsettings đã commit của 6 service.",
        "when": "Quét mọi connection string và đối chiếu với chủ sở hữu database.",
        "then": "Không vi phạm nào - không service nào đọc hay ghi database của service khác (SC-003).",
        "file": "tests/CrossServiceIsolation.Tests/ConnectionStringIsolationTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Cô lập database giữa các service",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_ActuallyExaminesEveryServicesConfiguration",
        "given": "Bộ quét connection string chạy trên thư mục services thật.",
        "when": "Đọc danh sách service, file cấu hình và số connection string đã quét.",
        "then": "Đủ 6 service, mỗi service có file cấu hình được quét, và số connection string ít nhất bằng số service sở hữu database - kết quả sạch không phải do quét nhầm chỗ.",
        "file": "tests/CrossServiceIsolation.Tests/ConnectionStringIsolationTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Cô lập database giữa các service",
        "layer": "Backend",
        "type": "Unit",
        "name": "NoStatelessService_DeclaresAConnectionString",
        "given": "Các file appsettings của hai service không sở hữu dữ liệu là bff và gateway.",
        "when": "Đọc từng file và tìm mục ConnectionStrings.",
        "then": "Tìm thấy file cấu hình và không file nào khai báo ConnectionStrings - service không sở hữu dữ liệu thì không được cấp database.",
        "file": "tests/CrossServiceIsolation.Tests/ConnectionStringIsolationTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Cô lập database giữa các service",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_FlagsAConfigurationThatReachesAnotherServicesDatabase",
        "given": "Một cây thư mục services giả được dựng tạm với cấu hình cố ý sai, 3 biến thể vi phạm khác nhau (Theory, 3 bộ dữ liệu).",
        "when": "Chạy bộ quét trên cây thư mục giả đó.",
        "then": "Phát hiện đúng một vi phạm, nêu đúng service vi phạm là parties và service bị xâm phạm là orders - chứng minh bộ quét thật sự phát hiện được lỗi.",
        "file": "tests/CrossServiceIsolation.Tests/ConnectionStringIsolationTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Cô lập database giữa các service",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_AllowsAServiceToNameItsOwnDatabase",
        "given": "Cây thư mục services giả với parties trỏ PartiesDb và orders trỏ OrdersDb - đều đúng chủ.",
        "when": "Chạy bộ quét.",
        "then": "Không có vi phạm nào - bộ quét không báo động giả.",
        "file": "tests/CrossServiceIsolation.Tests/ConnectionStringIsolationTests.cs",
    },
    # ---- Convention / TenantGatedConnectionTests -----------------------------------------------
    {
        "feature": "Common",
        "sub": "Convention - DbContext phải chặn theo tenant",
        "layer": "Backend",
        "type": "Unit",
        "name": "EveryDatabaseOwningService_HasExactlyOneDbContextRegistration",
        "given": "Mã nguồn Program.cs đã commit của 4 service sở hữu database.",
        "when": "Quét các điểm gọi AddDbContext.",
        "then": "Mỗi service có đúng một điểm đăng ký DbContext - chỉ một cửa duy nhất để canh.",
        "file": "tests/CrossServiceIsolation.Tests/TenantGatedConnectionTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - DbContext phải chặn theo tenant",
        "layer": "Backend",
        "type": "Unit",
        "name": "EveryDbContextRegistration_IsGatedOnAResolvedTenant",
        "given": "Mã nguồn Program.cs đã commit của 4 service sở hữu database.",
        "when": "Quét xem mỗi điểm đăng ký DbContext có gọi RequireTenantId trước khi mở kết nối không.",
        "then": "Số điểm được chặn bằng đúng tổng số điểm đăng ký - không tồn tại đường nào chạm database mà chưa phân giải tenant.",
        "file": "tests/CrossServiceIsolation.Tests/TenantGatedConnectionTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - DbContext phải chặn theo tenant",
        "layer": "Backend",
        "type": "Unit",
        "name": "NoStatelessService_RegistersADbContext",
        "given": "Kết quả quét của các service không sở hữu database (bff, gateway).",
        "when": "Đếm số điểm đăng ký DbContext của các service này.",
        "then": "Bằng 0 - service không sở hữu dữ liệu thì không mở kết nối database, dù có chặn tenant hay không.",
        "file": "tests/CrossServiceIsolation.Tests/TenantGatedConnectionTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - DbContext phải chặn theo tenant",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_ActuallyExaminesEveryServicesRegistration",
        "given": "Bộ quét chạy trên thư mục services thật.",
        "when": "Đọc danh sách service và số kết quả bộ quét ghi nhận.",
        "then": "Đủ 6 service và 6 kết quả - kết quả sạch không phải do quét nhầm thư mục.",
        "file": "tests/CrossServiceIsolation.Tests/TenantGatedConnectionTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - DbContext phải chặn theo tenant",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_FlagsAnUngatedRegistration",
        "given": "Một Program.cs giả đăng ký AddDbContext mà không gọi RequireTenantId.",
        "when": "Chạy bộ quét.",
        "then": "Ghi nhận 1 điểm đăng ký và 0 điểm được chặn - bộ quét phát hiện đúng vi phạm.",
        "file": "tests/CrossServiceIsolation.Tests/TenantGatedConnectionTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - DbContext phải chặn theo tenant",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_AcceptsAGatedRegistration",
        "given": "Một Program.cs giả đăng ký AddDbContext có gọi RequireTenantId trước khi tạo kết nối.",
        "when": "Chạy bộ quét.",
        "then": "Ghi nhận 1 điểm đăng ký và 1 điểm được chặn - bộ quét không báo động giả với mã đúng.",
        "file": "tests/CrossServiceIsolation.Tests/TenantGatedConnectionTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - DbContext phải chặn theo tenant",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_DoesNotAcceptAGuardThatOnlyAppearsInAComment",
        "given": "Một Program.cs giả chỉ có dòng comment nhắc tới RequireTenantId, còn mã thật thì không gọi.",
        "when": "Chạy bộ quét.",
        "then": "Số điểm được chặn bằng 0 - xóa cổng chặn mà để lại comment mô tả sẽ không qua được kiểm tra.",
        "file": "tests/CrossServiceIsolation.Tests/TenantGatedConnectionTests.cs",
    },
    # ---- Convention / VerticalSliceStructureTests ----------------------------------------------
    {
        "feature": "Common",
        "sub": "Convention - Cấu trúc vertical slice",
        "layer": "Backend",
        "type": "Unit",
        "name": "NoService_HasATopLevelTechnicalLayerFolder",
        "given": "Cấu trúc thư mục dự án API thật của 6 service.",
        "when": "Quét tìm các thư mục cấp cao mang tên tầng kỹ thuật như Controllers, Services, Repositories.",
        "then": "Không có vi phạm nào - mã được tổ chức theo nghiệp vụ, không theo tầng kỹ thuật (SC-004).",
        "file": "tests/StructureConventionTests/VerticalSliceStructureTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Cấu trúc vertical slice",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_ActuallyExaminesEveryServicesApiProject",
        "given": "Bộ quét cấu trúc chạy trên thư mục services thật.",
        "when": "Đọc danh sách service và dự án đã quét.",
        "then": "Đủ 6 service và 6 dự án API - kết quả sạch không phải do quét nhầm thư mục.",
        "file": "tests/StructureConventionTests/VerticalSliceStructureTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Cấu trúc vertical slice",
        "layer": "Backend",
        "type": "Unit",
        "name": "EveryService_OrganisesAtLeastOneCapabilityUnderFeatures",
        "given": "Cấu trúc thư mục dự án API thật của 6 service.",
        "when": "Quét các thư mục nghiệp vụ nằm dưới Features.",
        "then": "Mỗi service đều có ít nhất một thư mục nghiệp vụ - không đủ nếu chỉ vắng mặt thư mục tầng kỹ thuật.",
        "file": "tests/StructureConventionTests/VerticalSliceStructureTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Cấu trúc vertical slice",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_FlagsATopLevelTechnicalLayerFolder",
        "given": "Cây thư mục giả có thư mục cấp cao mang tên Controllers, Services, Repositories hoặc repositories (Theory, 4 bộ dữ liệu).",
        "when": "Chạy bộ quét.",
        "then": "Phát hiện đúng một vi phạm, nêu đúng tên service và tên thư mục vi phạm.",
        "file": "tests/StructureConventionTests/VerticalSliceStructureTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Cấu trúc vertical slice",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_AllowsCapabilityFoldersAndNonLayerFolders",
        "given": "Cây thư mục giả có Features/HealthCheck, Data và Properties.",
        "when": "Chạy bộ quét.",
        "then": "Không có vi phạm - các thư mục nghiệp vụ và thư mục không phải tầng kỹ thuật đều hợp lệ.",
        "file": "tests/StructureConventionTests/VerticalSliceStructureTests.cs",
    },
    {
        "feature": "Common",
        "sub": "Convention - Cấu trúc vertical slice",
        "layer": "Backend",
        "type": "Unit",
        "name": "Scan_AllowsATechnicalNameNestedInsideACapability",
        "given": "Cây thư mục giả có Features/HealthCheck/Services - tên tầng kỹ thuật nằm bên trong một nghiệp vụ.",
        "when": "Chạy bộ quét.",
        "then": "Không có vi phạm - mã vẫn nằm cùng nghiệp vụ mà nó phục vụ, đó là chuyện nội bộ của nghiệp vụ đó.",
        "file": "tests/StructureConventionTests/VerticalSliceStructureTests.cs",
    },
]

ROWS += [
    # ---- Frontend / Unit / shared/money.test.ts ------------------------------------------------
    {
        "feature": "Common",
        "sub": "Storefront - Định dạng tiền tệ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "formats %d as %s",
        "given": "Các số tiền 12.5, 48, 34.25 và 0 (it.each, 4 bộ dữ liệu).",
        "when": "Gọi formatMoney với từng giá trị.",
        "then": "Trả về $12.50, $48.00, $34.25 và $0.00 - luôn có ký hiệu đô la và đúng hai chữ số thập phân (FR-024).",
        "file": "frontend/apps/web/tests/shared/money.test.ts",
    },
    {
        "feature": "Common",
        "sub": "Storefront - Định dạng tiền tệ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "groups thousands so a large total stays readable",
        "given": "Số tiền lớn 1234.5.",
        "when": "Gọi formatMoney.",
        "then": "Trả về $1,234.50 - có dấu phân cách hàng nghìn cho dễ đọc.",
        "file": "frontend/apps/web/tests/shared/money.test.ts",
    },
    {
        "feature": "Common",
        "sub": "Storefront - Định dạng tiền tệ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "rounds to two decimal places",
        "given": "Các số tiền có nhiều hơn hai chữ số thập phân: 12.499 và 12.494.",
        "when": "Gọi formatMoney.",
        "then": "Trả về $12.50 và $12.49 - luôn đúng hai chữ số thập phân để giá hiển thị khớp với tổng tính ra.",
        "file": "frontend/apps/web/tests/shared/money.test.ts",
    },
    {
        "feature": "Common",
        "sub": "Storefront - Định dạng tiền tệ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "keeps the sign on a negative amount",
        "given": "Số tiền âm -5.",
        "when": "Gọi formatMoney.",
        "then": "Trả về -$5.00 - dấu âm hiển thị rõ chứ không bị âm thầm bỏ đi.",
        "file": "frontend/apps/web/tests/shared/money.test.ts",
    },
    {
        "feature": "Common",
        "sub": "Storefront - Định dạng tiền tệ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "accepts the string form the contract also permits",
        "given": "Số tiền ở dạng chuỗi '12.50' và '48' - dạng mà hợp đồng OpenAPI cho phép để giữ độ chính xác decimal.",
        "when": "Gọi formatMoney.",
        "then": "Trả về $12.50 và $48.00 - màn hình không hiện NaN khi backend trả dạng chuỗi.",
        "file": "frontend/apps/web/tests/shared/money.test.ts",
    },
    {
        "feature": "Common",
        "sub": "Storefront - Định dạng tiền tệ",
        "layer": "Frontend",
        "type": "Unit",
        "name": "refuses a value that is not an amount at all",
        "given": "Giá trị đầu vào là chuỗi 'not-a-price'.",
        "when": "Gọi formatMoney.",
        "then": "Ném TypeError thay vì hiển thị giá trị vô nghĩa.",
        "file": "frontend/apps/web/tests/shared/money.test.ts",
    },
    # ---- Frontend / Unit / accessibility.test.tsx ----------------------------------------------
    {
        "feature": "Common",
        "sub": "Storefront - Khả năng tiếp cận của shell",
        "layer": "Frontend",
        "type": "Unit",
        "name": "offers a skip link as the first focusable element",
        "given": "Ứng dụng App được render tại trang gốc, API sản phẩm và giỏ hàng được giả lập.",
        "when": "Nhấn Tab lần đầu tiên.",
        "then": "Liên kết Skip to content nhận focus - đúng WCAG 2.4.1, bỏ qua được khối điều hướng.",
        "file": "frontend/apps/web/tests/accessibility.test.tsx",
    },
    {
        "feature": "Common",
        "sub": "Storefront - Khả năng tiếp cận của shell",
        "layer": "Frontend",
        "type": "Unit",
        "name": "gives each screen its own document title",
        "given": "Ứng dụng App được render tại trang Products.",
        "when": "Chuyển sang màn hình Basket bằng liên kết điều hướng.",
        "then": "Tiêu đề tài liệu đổi từ Products sang Basket - đúng WCAG 2.4.2, mỗi màn hình tự xưng danh khác nhau.",
        "file": "frontend/apps/web/tests/accessibility.test.tsx",
    },
    {
        "feature": "Common",
        "sub": "Storefront - Khả năng tiếp cận của shell",
        "layer": "Frontend",
        "type": "Unit",
        "name": "moves focus to the new screen after navigating",
        "given": "Ứng dụng App được render tại trang gốc.",
        "when": "Bấm liên kết Basket và chờ nội dung màn hình mới xuất hiện.",
        "then": "Focus chuyển sang vùng main của màn hình mới - người dùng bàn phím không bị kẹt lại ở header trong khi nội dung bên dưới đã đổi.",
        "file": "frontend/apps/web/tests/accessibility.test.tsx",
    },
    {
        "feature": "Common",
        "sub": "Storefront - Khả năng tiếp cận của shell",
        "layer": "Frontend",
        "type": "Unit",
        "name": "does not steal focus on first render",
        "given": "Ứng dụng App vừa được render lần đầu.",
        "when": "Kiểm tra vị trí focus.",
        "then": "Vùng main không giữ focus - lần render đầu không cướp focus của người dùng.",
        "file": "frontend/apps/web/tests/accessibility.test.tsx",
    },
    # ---- Frontend / Unit / app.test.tsx --------------------------------------------------------
    {
        "feature": "Common",
        "sub": "Storefront - Application shell",
        "layer": "Frontend",
        "type": "Unit",
        "name": "renders the landing route inside the shell",
        "given": "Ứng dụng App với API sản phẩm được giả lập tại origin gateway mặc định (localhost:5300).",
        "when": "Render App.",
        "then": "Tiêu đề Products xuất hiện - provider được gắn, router phân giải và route mặc định render thành công.",
        "file": "frontend/apps/web/tests/app.test.tsx",
    },
    {
        "feature": "Common",
        "sub": "Storefront - Application shell",
        "layer": "Frontend",
        "type": "Unit",
        "name": "exposes navigation to the shopper",
        "given": "Ứng dụng App được render.",
        "when": "Tìm vùng điều hướng chính.",
        "then": "Có vùng navigation tên Main và liên kết Basket - người mua chuyển màn hình được.",
        "file": "frontend/apps/web/tests/app.test.tsx",
    },
]


# ==========================================================================================
# CÁCH LÀM CHO TEST CASE BỊ FAIL — viết theo góc nhìn QA
#
# Mỗi mô tả là một thao tác QA thực hiện được: bấm trên giao diện, gọi API, bật/tắt container,
# hoặc sửa mã nguồn dưới máy local để tái hiện. Không đề xuất cách sửa lỗi — chỉ nêu cách làm
# cho kỳ vọng của test không còn đúng.
#
# Khoá tra cứu là cặp (file nguồn, tên test gốc), đã kiểm chứng là duy nhất trên cả 249 case.
# ==========================================================================================

BREAK_HINTS: dict[tuple[str, str], str] = {}

_COMPOSE = "docker compose -f docker-compose.local.yml"

# Bốn service sở hữu database, giống nhau ở ba nhóm test bên dưới nên sinh bằng vòng lặp.
_SERVICES = {
    "baskets": {"cap": "Baskets", "api": "baskets-api", "db": "baskets-db", "port": 5188,
                "read": "/baskets/{id bất kỳ}", "dbport": 14332},
    "orders": {"cap": "Orders", "api": "orders-api", "db": "orders-db", "port": 5041,
               "read": "/orders/{id bất kỳ}", "dbport": 14333},
    "parties": {"cap": "Parties", "api": "parties-api", "db": "parties-db", "port": 5204,
                "read": "/parties/{id bất kỳ}", "dbport": 14330},
    "products": {"cap": "Products", "api": "products-api", "db": "products-db", "port": 5088,
                 "read": "/products", "dbport": 14331},
}

for _name, _s in _SERVICES.items():
    _cap, _api, _db, _port, _read = _s["cap"], _s["api"], _s["db"], _s["port"], _s["read"]

    BREAK_HINTS[(f"services/{_name}/tests/{_cap}.Api.UnitTests/HealthCheckTests.cs",
                 "HealthLive_ReturnsOk")] = (
        f"Dừng container {_api} (`{_COMPOSE} stop {_api}`) rồi gọi GET http://localhost:{_port}/health/live "
        f"— không nhận được phản hồi là case đỏ. Lưu ý: tắt database {_db} KHÔNG làm case này đỏ, vì "
        f"liveness cố ý không chạm tới database."
    )

    BREAK_HINTS[(f"services/{_name}/tests/{_cap}.Api.IntegrationTests/ReadinessTests.cs",
                 "HealthReady_ReturnsOk_WhenDatabaseReachable")] = (
        f"Tắt database riêng của service (`{_COMPOSE} stop {_db}`) rồi gọi GET "
        f"http://localhost:{_port}/health/ready — nhận 503 thay vì 200 là case đỏ. Bật lại {_db} để về xanh."
    )

    BREAK_HINTS[(f"services/{_name}/tests/{_cap}.Api.IntegrationTests/ReadinessTests.cs",
                 "HealthReady_ReturnsServiceUnavailable_WhenDatabaseUnreachable")] = (
        f"Cần làm cho readiness không phát hiện được database chết. Dưới local, sửa readiness của "
        f"{_cap} để bỏ phần mở kết nối database (chỉ báo tiến trình còn sống), rồi tắt {_db} và gọi "
        f"GET http://localhost:{_port}/health/ready — vẫn trả 200 là case đỏ."
    )

    BREAK_HINTS[(f"services/{_name}/tests/{_cap}.Api.IntegrationTests/ReadinessTests.cs",
                 "HealthReady_DoesNotFallBackToAnotherServicesDatabase_WhenOwnDatabaseUnreachable")] = (
        f"Dưới local, cấu hình {_cap} để khi database riêng hỏng thì quay sang dùng database của "
        f"service khác, rồi tắt {_db} và gọi GET http://localhost:{_port}/health/ready — trả 200, hoặc "
        f"nội dung phản hồi không còn nêu check 'self-database' là Unhealthy, đều là case đỏ."
    )

    BREAK_HINTS[(f"services/{_name}/tests/{_cap}.Api.IntegrationTests/TenantEnforcementTests.cs",
                 "ResolvingTheDbContext_Throws_WhenNoTenantHasBeenResolved")] = (
        f"Dưới local, cho {_cap} một tenant mặc định khi không phân giải được (gán sẵn giá trị lúc "
        f"khởi tạo thay vì bắt buộc có header). Khi đó truy cập database không còn ném lỗi và case đỏ."
    )

    BREAK_HINTS[(f"services/{_name}/tests/{_cap}.Api.IntegrationTests/TenantEnforcementTests.cs",
                 "ARequestWithoutATenant_Fails_RatherThanServingDefaultSchemaData")] = (
        f"Gọi GET http://localhost:{_port}{_read} KHÔNG kèm header X-Tenant-Id: nếu nhận 200 kèm dữ liệu "
        f"thay vì lỗi 500 thì case đỏ. Tái hiện bằng cách sửa {_cap} dưới local để dùng tenant mặc định "
        f"khi thiếu header."
    )

BREAK_HINTS[("services/orders/tests/Orders.Api.IntegrationTests/TenantEnforcementTests.cs",
             "AWriteWithoutATenant_CreatesNoOrder")] = (
    "Đếm số đơn trong database orders (kết nối localhost,14333), gọi POST http://localhost:5041/orders "
    "KHÔNG kèm X-Tenant-Id, rồi đếm lại. Số đơn tăng lên là case đỏ — kể cả khi API vẫn trả lỗi 500, "
    "vì điều được canh ở đây là không có bản ghi nào được tạo."
)

# ---------------------------------------------------------------- BASKET (backend)

_F = "services/baskets/tests/Baskets.Api.UnitTests/BasketLineMergeTests.cs"
BREAK_HINTS.update({
    (_F, "AddItem_CreatesALine_WhenTheProductIsNotInTheBasketYet"):
        "Từ giỏ rỗng, thêm 1 Notebook rồi gọi GET http://localhost:5188/baskets/current — nếu giỏ "
        "không có đúng 1 dòng đúng mã sản phẩm, số lượng 1, đơn giá 12.50 thì case đỏ. Dưới local, "
        "sửa Baskets để lần thêm đầu tiên không tạo dòng nào là tái hiện được.",
    (_F, "AddItem_IncrementsTheExistingLine_WhenTheProductIsAlreadyInTheBasket"):
        "Thêm cùng một sản phẩm hai lần rồi xem giỏ: nếu ra 2 dòng riêng thay vì 1 dòng số lượng 2 thì "
        "case đỏ. Dưới local, sửa Baskets để mỗi lần thêm luôn tạo dòng mới thay vì cộng dồn.",
    (_F, "AddItem_KeepsProductsApart_WhenDifferentProductsAreAdded"):
        "Thêm Notebook rồi thêm Apron: nếu hai sản phẩm khác nhau bị dồn vào chung một dòng thì case đỏ. "
        "Dưới local, sửa luật gộp để nó gộp theo bất kỳ sản phẩm nào thay vì chỉ gộp khi trùng mã.",
    (_F, "AddItem_AccumulatesQuantities_AcrossManyAdditions"):
        "Bấm 'Thêm vào giỏ' 5 lần cho cùng một sản phẩm rồi xem giỏ: số lượng phải đúng bằng 5. Ra 1, "
        "ra 2, hay ra 5 dòng riêng đều là case đỏ.",
    (_F, "AddItem_KeepsTheOriginallyCapturedPrice_WhenTheCatalogPriceHasChanged"):
        "Thêm sản phẩm vào giỏ, sau đó đổi giá sản phẩm đó trong database products (localhost,14331), "
        "rồi thêm chính nó lần nữa — nếu dòng trong giỏ đổi sang giá mới thì case đỏ. Giá phải giữ "
        "nguyên mức đã chốt lúc thêm lần đầu.",
    (_F, "AddItem_Rejects_AQuantityBelowOne"):
        "Gọi POST http://localhost:5188/baskets/current/items với quantity 0 hoặc số âm — nếu được chấp "
        "nhận và giỏ xuất hiện dòng số lượng 0/âm thay vì bị từ chối thì case đỏ.",
    (_F, "AddItem_Rejects_ANegativeUnitPrice"):
        "Gọi POST http://localhost:5188/baskets/current/items với unitPrice âm (ví dụ -0.01) — nếu được "
        "chấp nhận thay vì bị từ chối thì case đỏ.",
})

_F = "services/baskets/tests/Baskets.Api.UnitTests/BasketTotalTests.cs"
BREAK_HINTS.update({
    (_F, "Total_IsZero_ForAnEmptyBasket"):
        "Xem giỏ khi chưa thêm gì: nếu tổng tiền khác 0 thì case đỏ. Dưới local, sửa phép tính tổng để "
        "cộng thêm một khoản cố định là tái hiện được.",
    (_F, "Total_MultipliesQuantityByUnitPrice"):
        "Thêm 2 Notebook giá 12.50 rồi xem giỏ: tổng phải là 25.00. Nếu ra 12.50 (quên nhân số lượng) "
        "hoặc bất kỳ con số nào khác thì case đỏ.",
    (_F, "Total_SumsEveryLine"):
        "Thêm 2 Notebook (12.50) và 1 Pour-Over (48.00) rồi xem giỏ: tổng phải là 73.00. Nếu chỉ tính "
        "một dòng, hoặc bỏ sót dòng cuối, thì case đỏ.",
    (_F, "Total_MatchesTheWalkthroughFigure"):
        "Dựng đúng giỏ của kịch bản demo — 2 Notebook + 1 Apron — rồi xem tổng: khác 59.25 là case đỏ. "
        "Đây là con số mọi tài liệu demo và kịch bản E2E đang trích dẫn.",
    (_F, "Total_IsExact_ForAmountsThatFloatingPointWouldRound"):
        "Tạo giỏ có hai dòng giá 0.10 và 0.20, xem tổng: phải đúng 0.30. Ra 0.30000000000000004 là case "
        "đỏ — tái hiện bằng cách sửa kiểu số tiền trong Baskets từ decimal sang double dưới local.",
    (_F, "Clear_EmptiesTheBasket_AndZeroesTheTotal"):
        "Cho giỏ có hàng rồi gọi POST http://localhost:5188/baskets/current/clear: nếu sau đó giỏ vẫn "
        "còn dòng, hoặc danh sách rỗng nhưng tổng tiền vẫn khác 0, thì case đỏ.",
})

_F = "services/baskets/tests/Baskets.Api.IntegrationTests/BasketEndpointsTests.cs"
BREAK_HINTS.update({
    (_F, "GetBasket_ReturnsTheBasket_WhenItExists"):
        "Tạo giỏ có 2 Notebook, lấy id của nó rồi gọi GET http://localhost:5188/baskets/{id} kèm "
        "X-Tenant-Id — nếu trả về sai id, sai CustomerRef, thiếu danh sách dòng, hoặc tổng khác 25.00 "
        "thì case đỏ.",
    (_F, "GetBasket_ReturnsNotFound_WhenNoBasketHasThatId"):
        "Gọi GET http://localhost:5188/baskets/{một GUID không tồn tại} — nếu trả 200 kèm giỏ rỗng thay "
        "vì 404 thì case đỏ. Hai tình huống 'không có giỏ' và 'giỏ không có gì' phải phân biệt được.",
})

_F = "services/baskets/tests/Baskets.Api.IntegrationTests/ClearBasketTests.cs"
BREAK_HINTS.update({
    (_F, "Clear_RemovesEveryLine_ButKeepsTheBasket"):
        "Ghi lại id giỏ, thêm hàng, gọi POST /baskets/current/clear rồi xem lại giỏ: nếu id giỏ đổi "
        "thành id mới thì case đỏ (chứng tỏ giỏ bị xoá rồi tạo lại chứ không phải chỉ dọn hàng). Trả về "
        "khác 204 cũng đỏ.",
    (_F, "Clear_ReturnsConflict_WhenTheBasketIsAlreadyEmpty"):
        "Gọi POST http://localhost:5188/baskets/current/clear khi giỏ đang rỗng — nếu trả 204 (im lặng "
        "thành công) thay vì 409 thì case đỏ.",
    (_F, "Clear_ReturnsConflict_OnASecondClear"):
        "Thêm hàng, dọn giỏ (204), rồi dọn lần nữa ngay: lần thứ hai phải là 409. Nếu vẫn 204 thì case "
        "đỏ — và đó cũng là lỗ hổng khiến bấm đặt hàng hai lần có thể tạo hai đơn.",
    (_F, "Clear_LeavesTheBasketUsable_ForTheNextPurchase"):
        "Sau khi dọn giỏ, thêm tiếp 3 Notebook rồi xem giỏ: phải có 1 dòng số lượng 3, tổng 37.50. Nếu "
        "thêm không được, hoặc giỏ vẫn rỗng, thì case đỏ.",
})

_F = "services/baskets/tests/Baskets.Api.IntegrationTests/CurrentBasketTests.cs"
BREAK_HINTS.update({
    (_F, "GetCurrent_ReturnsAnEmptyBasket_ForACallerWhoHasNeverAddedAnything"):
        "Dùng một X-Subject-Id chưa từng dùng bao giờ rồi gọi GET http://localhost:5188/baskets/current "
        "— phải nhận 200 với giỏ rỗng. Nếu trả 404 (coi 'chưa từng mua' là lỗi) thì case đỏ.",
    (_F, "GetCurrent_ReturnsTheSameBasket_AcrossSeparateRequests"):
        "Thêm 2 Notebook rồi gọi GET /baskets/current hai lần liên tiếp: nếu hai lần trả về id giỏ khác "
        "nhau, hoặc lần sau mất hàng, thì case đỏ. Tương đương việc F5 trang giỏ hàng mà giỏ trống trơn.",
    (_F, "GetCurrent_GivesDifferentShoppersDifferentBaskets"):
        "Gọi /baskets/current với X-Subject-Id: alice, thêm hàng; rồi gọi lại với X-Subject-Id: bob — "
        "nếu bob nhìn thấy hàng của alice, hoặc hai người ra cùng một id giỏ, thì case đỏ.",
    (_F, "AddItem_MergesIntoTheExistingLine_WhenTheSameProductIsAddedAgain"):
        "Gọi POST /baskets/current/items hai lần cho cùng sản phẩm rồi xem giỏ: phải là 1 dòng số lượng "
        "2, tổng 25.00. Ra 2 dòng là case đỏ — khác với case unit ở chỗ đây kiểm tra việc gộp còn đúng "
        "khi đi qua hai request riêng biệt và ghi xuống database thật.",
    (_F, "AddItem_KeepsDistinctProductsOnSeparateLines"):
        "Thêm Notebook 2 lần và Apron 1 lần: giỏ phải có 2 dòng, tổng 59.25. Ra 1 dòng hoặc 3 dòng, "
        "hoặc tổng khác, đều là case đỏ.",
    (_F, "AddItem_Rejects_AQuantityBelowOne"):
        "Gọi POST http://localhost:5188/baskets/current/items với quantity 0 hoặc -3 — nhận 200 hay 500 "
        "thay vì 400 đều là case đỏ. API phải từ chối rõ ràng là do dữ liệu gửi lên sai.",
    (_F, "GetCurrent_Fails_WhenNoCallerWasResolved"):
        "Gọi GET http://localhost:5188/baskets/current có X-Tenant-Id nhưng KHÔNG có X-Subject-Id — nếu "
        "vẫn trả về một giỏ nào đó thay vì lỗi 500 thì case đỏ, vì khi đó mọi request thiếu subject sẽ "
        "dùng chung một giỏ.",
})

# ---------------------------------------------------------------- BASKET (frontend)

_F = "frontend/apps/web/tests/basket/AddItemError.test.tsx"
BREAK_HINTS.update({
    (_F, "adds the product when the request succeeds"):
        "Mở tab Network, bấm 'Add to basket' trên http://localhost:4173 và xem nội dung request gửi đi — "
        "nếu ngoài productId và quantity còn kèm cả giá thì case đỏ, vì khi đó giá là do client tự khai. "
        "Không gửi được request nào cũng đỏ.",
    (_F, "shows a clear error when the request fails"):
        "Tắt baskets-api rồi bấm 'Add to basket' — nếu không hiện thông báo lỗi nào, hoặc tệ hơn là giỏ "
        "hiển thị món vừa bấm dù server đã từ chối, thì case đỏ. Trường hợp thứ hai nguy hiểm hơn: khách "
        "sẽ đặt hàng trong khi tin rằng đã mua được món đó.",
    (_F, "disables itself while the addition is in flight"):
        "Bấm 'Add to basket' liên tục thật nhanh khi mạng chậm (dùng Network throttling trong DevTools) "
        "— nếu đếm được nhiều request hơn số lần thao tác hợp lệ thì case đỏ. Tái hiện bằng cách bỏ đoạn "
        "vô hiệu hoá nút trong lúc đang gửi.",
    (_F, "can be operated by keyboard"):
        "Không dùng chuột: Tab tới nút 'Add to basket' rồi Enter — nếu không Tab tới được, hoặc Enter "
        "không gửi request, thì case đỏ.",
})

_F = "frontend/apps/web/tests/basket/BasketView.test.tsx"
BREAK_HINTS.update({
    (_F, "shows each line with its name, quantity, unit price, and line total"):
        "Thêm 2 Notebook rồi mở http://localhost:4173/basket — thiếu tên sản phẩm, thiếu cụm số lượng x "
        "đơn giá, hoặc thiếu thành tiền của dòng, đều là case đỏ.",
    (_F, "shows the basket total the backend reported"):
        "Thêm 2 Notebook và 1 Apron rồi xem trang giỏ — tổng phải hiện $59.25. Case đỏ nếu con số khác, "
        "hoặc nếu giao diện tự cộng lại thay vì hiển thị tổng server trả về (tái hiện bằng cách sửa "
        "giao diện tự tính tổng từ các dòng — khi đó tổng có thể lệch với số tiền thực bị tính).",
    (_F, "presents the lines as a list"):
        "Xem cấu trúc trang giỏ trong DevTools hoặc bằng trình đọc màn hình — nếu các dòng hàng không "
        "còn được đánh dấu là danh sách có nhãn thì case đỏ.",
    (_F, "tells the shopper when the basket is empty"):
        "Dọn sạch giỏ rồi mở trang giỏ hàng — nếu thấy trang trắng, hoặc hiện thông báo lỗi, thay vì "
        "dòng chữ báo giỏ đang trống thì case đỏ. Giỏ rỗng là trạng thái bình thường, nhất là ngay sau "
        "khi đặt hàng xong.",
    (_F, "shows a readable error when the basket cannot be loaded"):
        "Tắt baskets-api rồi mở trang giỏ hàng — nếu không có thông báo lỗi đọc được, hoặc có thông báo "
        "nhưng không kèm nút 'Try again', thì case đỏ.",
})

# ---------------------------------------------------------------- GATEWAY

_F = "services/gateway/tests/Gateway.Api.UnitTests/ForwardingTimeoutBudgetTests.cs"
BREAK_HINTS.update({
    (_F, "TheGatewaysForwardingTimeout_IsAtLeastTheBffsTotalDownstreamBudget"):
        "Mở services/gateway/src/Gateway.Api/appsettings.json, hạ ActivityTimeout của bff-cluster xuống "
        "dưới 3 giây (ví dụ 00:00:02) — case đỏ ngay. Hậu quả thật: gateway cắt request trong khi BFF "
        "vẫn đang chờ hợp lệ, người dùng nhận lỗi mà log của BFF không giải thích được.",
    (_F, "TheGatewaysForwardingTimeout_IsBounded"):
        "Xoá hẳn dòng ActivityTimeout khỏi appsettings.json của gateway, hoặc đặt nó thành một giá trị "
        "rất lớn (trên 1 phút) — case đỏ. Đây là case chặn tình huống 'chờ vô hạn'.",
})

_F = "services/gateway/tests/Gateway.Api.UnitTests/RouteConfigurationTests.cs"
BREAK_HINTS.update({
    (_F, "TheConfiguration_DefinesExactlyOneRoute_ToTheBffCluster"):
        "Thêm một route thứ hai vào mục ReverseProxy trong appsettings.json của gateway, hoặc đổi tên "
        "route/cluster hiện có khác 'bff-route'/'bff-cluster' — case đỏ.",
    (_F, "TheRoute_MatchesEveryPath"):
        "Đổi Match.Path của route từ '{**catch-all}' sang một đường dẫn cụ thể như '/bff/{**rest}' — "
        "case đỏ. Hậu quả thật: mỗi lần BFF thêm đường dẫn mới lại phải sửa gateway.",
    (_F, "TheConfiguration_DefinesExactlyOneCluster_WithOneDestination"):
        "Thêm cluster thứ hai, thêm destination thứ hai vào bff-cluster, hoặc để trống địa chỉ "
        "destination — cả ba đều làm case đỏ.",
    (_F, "TheConfiguration_NamesNoDomainServiceAsADestination"):
        "Thêm vào cấu hình gateway một destination trỏ thẳng tới products-api, baskets-api, orders-api "
        "hoặc parties-api — case đỏ. Đây là case chặn việc mở đường đi tắt vòng qua BFF.",
    (_F, "EveryRoute_ResolvesToADefinedCluster"):
        "Sửa ClusterId của route thành một tên không tồn tại (ví dụ 'bff-cluster-2'), hoặc xoá hẳn "
        "ClusterId — case đỏ, vì khi đó có route nhưng không dẫn đi đâu cả.",
})

_F = "services/gateway/tests/Gateway.Api.UnitTests/StubIdentityAuthenticationHandlerTests.cs"
BREAK_HINTS.update({
    (_F, "AuthenticateAsync_Succeeds_ForAnyRequest"):
        "Dưới local, sửa stub identity của gateway để nó từ chối xác thực trong một điều kiện nào đó "
        "(ví dụ khi thiếu header Authorization) — case đỏ. Giai đoạn 1 chưa có thông tin đăng nhập để "
        "đọc, nên mọi request đều phải xác thực thành công.",
    (_F, "AuthenticateAsync_IssuesTheConfiguredTenantClaim"):
        "Đổi StubIdentity:TenantId trong cấu hình gateway thành giá trị khác rồi kiểm tra tenant gắn vào "
        "request; hoặc sửa stub để không phát ra claim tenant nữa — case đỏ.",
    (_F, "AuthenticateAsync_IssuesTheConfiguredSubjectClaim"):
        "Sửa stub identity để chỉ phát ra tenant mà không phát ra subject — case đỏ. Hậu quả thật: mọi "
        "endpoint theo người mua (giỏ hàng, đặt hàng) sẽ lỗi 500.",
    (_F, "AuthenticateAsync_IgnoresTheRequest_HavingNoCredentialsToRead"):
        "Sửa stub để nó đọc header Authorization của người gọi và đổi kết quả theo đó — case đỏ. Ở giai "
        "đoạn 1, nội dung request không được phép ảnh hưởng tới danh tính phân giải ra.",
    (_F, "AuthenticateAsync_Fails_WhenNoTenantIsConfigured"):
        "Sửa stub để khi StubIdentity:TenantId bị bỏ trống thì vẫn xác thực thành công (ví dụ tự điền "
        "một tenant mặc định) — case đỏ. Một principal không có tenant là request chưa phân giải đội lốt "
        "đã phân giải.",
    (_F, "AuthenticateAsync_IssuesAnIdentityNamingTheStubScheme"):
        "Sửa stub để tạo identity không đặt tên scheme (hoặc đặt tên khác) — case đỏ. Case này giữ cho "
        "stub vẫn là một cơ chế xác thực thật, không tụt xuống thành mẹo gắn header.",
})

_F = "services/gateway/tests/Gateway.Api.UnitTests/SubjectHeaderPropagationMiddlewareTests.cs"
BREAK_HINTS.update({
    (_F, "InvokeAsync_StampsTheSubjectHeader_FromTheAuthenticatedPrincipal"):
        "Dưới local, gỡ bước gắn header X-Subject-Id ở gateway rồi mua hàng qua giao diện "
        "http://localhost:4173 — giỏ hàng lỗi 500 vì phía dưới không biết người mua là ai. Case đỏ.",
    (_F, "InvokeAsync_OverwritesACallerSuppliedSubject_NeverTrustsIt"):
        "Sửa gateway để giữ lại header X-Subject-Id do client gửi lên thay vì ghi đè, rồi gọi "
        "http://localhost:5300/bff/basket kèm 'X-Subject-Id: alice' — nếu giỏ trả về customerRef là "
        "alice thì case đỏ, và đó là lỗ hổng cho phép xem giỏ của người khác.",
    (_F, "InvokeAsync_RemovesTheHeader_WhenNoSubjectIsResolved"):
        "Sửa gateway để khi không phân giải được subject thì cứ để nguyên header client gửi lên (thay vì "
        "gỡ bỏ) — case đỏ. Đây chính là đường tuồn danh tính mà case trên đang bịt.",
    (_F, "InvokeAsync_AlwaysCallsTheRestOfThePipeline"):
        "Sửa gateway để chặn request ngay tại middleware khi không phân giải được subject (ví dụ trả 401 "
        "và dừng) — case đỏ. Hậu quả thật: các endpoint không cần danh tính, như health probe, cũng bị "
        "chặn theo.",
})

_F = "services/gateway/tests/Gateway.Api.IntegrationTests/CorrelationIdPropagationTests.cs"
BREAK_HINTS.update({
    (_F, "AGeneratedCorrelationId_ReachesTheBff_AndMatchesWhatTheCallerIsGiven"):
        "Gọi http://localhost:5300/bff/products khi products-api đang tắt, so header X-Correlation-Id "
        "trong phản hồi với trường correlationId trong nội dung lỗi — hai giá trị khác nhau là case đỏ, "
        "vì khi đó mã tra cứu đưa cho khách không tìm được gì trong log của BFF.",
    (_F, "ACallerSuppliedCorrelationId_IsPreservedEndToEnd"):
        "Gọi qua gateway kèm sẵn 'X-Correlation-Id: ma-cua-toi' — nếu phản hồi trả về một mã khác do hệ "
        "thống tự sinh thì case đỏ, vì khách hàng mất đường nối giữa log của họ và log của mình.",
})

_F = "services/gateway/tests/Gateway.Api.IntegrationTests/DownstreamUnavailableTests.cs"
BREAK_HINTS.update({
    (_F, "ARequest_ReturnsAClearError_WhenTheBffIsUnreachable"):
        "Tắt bff-api (`docker compose -f docker-compose.local.yml stop bff-api`) rồi gọi "
        "http://localhost:5300/bff/products và bấm giờ — nếu treo quá 5 giây, hoặc trả về lỗi socket "
        "thô thay vì mã lỗi rõ ràng, thì case đỏ.",
    (_F, "TheGatewaysOwnHealth_StaysHealthy_WhenTheBffIsUnreachable"):
        "Tắt bff-api rồi gọi http://localhost:5300/health/live và /health/ready — nếu gateway báo không "
        "khoẻ theo thì case đỏ. Hậu quả thật: một sự cố ở BFF làm toàn bộ gateway bị khởi động lại.",
    (_F, "TheError_LeaksNoInternalRoutingDetail"):
        "Tắt bff-api rồi đọc kỹ nội dung lỗi trả về từ http://localhost:5300/bff/products — nếu thấy "
        "127.0.0.1, tên bff-cluster, bff-route hay SocketException thì case đỏ.",
})

_F = "services/gateway/tests/Gateway.Api.IntegrationTests/RoutingTests.cs"
BREAK_HINTS.update({
    (_F, "ARequestToTheGateway_ReachesAResponseOnlyTheBffCanProduce"):
        "Gọi http://localhost:5300/openapi/v1.json — nếu không nhận được tài liệu OpenAPI có đường dẫn "
        "/bff/products thì case đỏ, vì đó là bằng chứng request thật sự đi tới BFF chứ không dừng ở "
        "gateway. Tái hiện bằng cách trỏ destination của gateway sang một địa chỉ khác.",
    (_F, "AClientFacingRoute_IsForwardedToTheBffsHandler"):
        "Gọi http://localhost:5300/bff/products — nhận 404 là case đỏ (nghĩa là gateway không chuyển "
        "tiếp). Tái hiện bằng cách thu hẹp Match.Path của route để nó không còn khớp /bff/products.",
    (_F, "TheGatewaysOwnHealthProbes_AreServedLocally_NotForwarded"):
        "Sửa cấu hình route của gateway để /health/live và /health/ready cũng bị chuyển tiếp xuống BFF, "
        "rồi tắt bff-api và gọi hai đường dẫn đó — không còn 200 là case đỏ.",
})

_F = "services/gateway/tests/Gateway.Api.IntegrationTests/StorefrontCorsTests.cs"
BREAK_HINTS.update({
    (_F, "APreflightFromTheStorefront_IsAllowed"):
        "Xoá http://localhost:5173 và http://localhost:4173 khỏi danh sách origin được phép của gateway, "
        "rồi gửi OPTIONS http://localhost:5300/bff/products kèm 'Origin: http://localhost:4173' — không "
        "còn header Access-Control-Allow-Origin là case đỏ. Trên giao diện, biểu hiện là trang trắng vì "
        "trình duyệt chặn mọi lời gọi.",
    (_F, "APreflightFromTheStorefront_AllowsCredentials"):
        "Sửa chính sách CORS của gateway để dùng origin dấu * hoặc bỏ phần cho phép credentials, rồi gửi "
        "lại preflight — thiếu 'Access-Control-Allow-Credentials: true', hoặc Allow-Origin thành '*', "
        "đều là case đỏ.",
    (_F, "ARequestFromAnUnknownOrigin_IsNotAdmitted"):
        "Sửa gateway để chấp nhận mọi origin, rồi gửi preflight kèm 'Origin: http://evil.example' — nếu "
        "nhận được Access-Control-Allow-Origin thì case đỏ.",
    (_F, "TheAllowedOrigins_ComeFromConfiguration"):
        "Sửa gateway để danh sách origin được viết cứng trong code thay vì đọc từ cấu hình, rồi khai một "
        "origin mới trong cấu hình và gửi preflight từ đó — không được chấp nhận là case đỏ.",
    (_F, "ARequestWithNoOrigin_IsUntouched"):
        "Gọi http://localhost:5300/health/live không kèm header Origin — nếu phản hồi vẫn có header "
        "Access-Control-Allow-Origin, hoặc không còn trả 200, thì case đỏ.",
    (_F, "EachConfiguredOrigin_IsAdmitted"):
        "Bỏ bớt một trong hai origin đã cấu hình (5173 của dev server, 4173 của storefront container), "
        "rồi gửi preflight từ chính origin vừa bỏ — case đỏ.",
    (_F, "TheDevelopmentConfiguration_Admits_BothStorefrontOrigins"):
        "Xoá hoặc sửa mục Cors trong services/gateway/src/Gateway.Api/appsettings.Development.json — "
        "case đỏ. Case này khác các case trên ở chỗ nó đọc file cấu hình thật đã commit, chính là loại "
        "lỗi từng lọt lưới: cơ chế thì đúng nhưng repo lại quên cấu hình.",
})

_F = "services/gateway/tests/Gateway.Api.IntegrationTests/TenantPropagationTests.cs"
BREAK_HINTS.update({
    (_F, "ARequestThroughTheGateway_CarriesTheResolvedTenantToTheBff"):
        "Gỡ bước gắn header X-Tenant-Id ở gateway dưới local, rồi mua hàng qua http://localhost:4173 — "
        "mọi màn hình lỗi vì phía dưới không phân giải được tenant. Case đỏ.",
    (_F, "ACallerSuppliedTenant_IsOverwritten_NeverTrusted"):
        "Sửa gateway để giữ lại X-Tenant-Id do client gửi thay vì ghi đè, rồi gọi "
        "http://localhost:5300/bff/products kèm 'X-Tenant-Id: tenant-khac' — nếu request đi xuống mang "
        "tenant đó thì case đỏ, và đó là lỗ hổng vượt ranh giới cô lập dữ liệu.",
    (_F, "EveryForwardedRequest_CarriesATenant"):
        "Sửa gateway để chỉ gắn tenant cho một số đường dẫn nhất định, rồi gọi một đường dẫn nằm ngoài "
        "danh sách đó — request đi xuống không mang tenant là case đỏ.",
})

_F = "services/gateway/tests/Gateway.Api.IntegrationTests/UnmatchedRouteTests.cs"
BREAK_HINTS.update({
    (_F, "AnUnknownPath_ReturnsAClearNotFound_RatherThanHanging"):
        "Gọi http://localhost:5300/no-such-path và bấm giờ — nếu treo lâu (trên 15 giây) hoặc trả mã "
        "khác 404 thì case đỏ. Tái hiện bằng cách trỏ destination của gateway sang một địa chỉ không "
        "phản hồi.",
    (_F, "AnUnknownPathsResponse_LeaksNoInternalRoutingDetail"):
        "Đọc nội dung phản hồi 404 của http://localhost:5300/no-such-path — nếu thấy bff-cluster, "
        "bff-route, products-api hay số cổng 8080 thì case đỏ. Tái hiện bằng cách bật chế độ hiển thị "
        "lỗi chi tiết cho gateway dưới local.",
})

# ---------------------------------------------------------------- ORDER (backend)

_F = "services/orders/tests/Orders.Api.UnitTests/OrderTenantTests.cs"
BREAK_HINTS.update({
    (_F, "PlaceFrom_RecordsTheTenantItWasPlacedFor"):
        "Đặt một đơn qua http://localhost:5300/bff/checkout rồi đọc đơn đó tại "
        "http://localhost:5041/orders/{id} kèm X-Tenant-Id — nếu trường tenantId trống hoặc sai thì "
        "case đỏ. Tái hiện bằng cách sửa Orders dưới local để không gán tenant khi tạo đơn.",
    (_F, "PlaceFrom_Rejects_AnAbsentOrBlankTenant"):
        "Sửa Orders dưới local để chấp nhận tenant rỗng (thay vì bắt buộc phải có), rồi đặt đơn không "
        "kèm tenant — đơn được tạo với tenantId rỗng là case đỏ. Hậu quả thật: có bản ghi đơn hàng "
        "không quy được về khách nào.",
    (_F, "PlaceFrom_StillRejects_AnEmptyLineSet_EvenWithATenant"):
        "Gọi POST http://localhost:5041/orders với items rỗng nhưng có đủ header tenant — nếu đơn vẫn "
        "được tạo thì case đỏ. Case này canh việc thêm luật tenant không làm mất luật cũ.",
})

_F = "services/orders/tests/Orders.Api.UnitTests/OrderTotalTests.cs"
BREAK_HINTS.update({
    (_F, "PlaceFrom_MultipliesQuantityByUnitPrice"):
        "Đặt đơn gồm 2 Notebook giá 12.50 rồi đọc lại đơn: tổng phải là 25.00. Ra 12.50 (quên nhân số "
        "lượng) là case đỏ.",
    (_F, "PlaceFrom_SumsEveryLine"):
        "Đặt đơn gồm 2 Notebook và 1 Apron rồi đọc lại: tổng phải là 59.25. Nếu chỉ tính một dòng, hoặc "
        "bỏ sót dòng cuối, thì case đỏ.",
    (_F, "PlaceFrom_RecordsWhenTheOrderWasPlaced_AndGivesItAnIdentifier"):
        "Đặt đơn rồi xem phản hồi: thiếu thời điểm đặt, hoặc mã đơn là GUID toàn số 0, đều là case đỏ. "
        "Trên giao diện, biểu hiện là màn hình xác nhận không có mã đơn để đọc cho khách.",
    (_F, "PlaceFrom_Rejects_AnEmptyLineSet"):
        "Gọi POST http://localhost:5041/orders với danh sách items rỗng — nếu tạo ra một đơn tổng 0 thay "
        "vì bị từ chối thì case đỏ.",
    (_F, "PlaceFrom_Rejects_ALineWithANonPositiveQuantity"):
        "Gọi POST http://localhost:5041/orders với một dòng số lượng 0 hoặc âm — nếu đơn vẫn được tạo "
        "thì case đỏ.",
    (_F, "PlaceFrom_Rejects_ALineWithANegativePrice"):
        "Gọi POST http://localhost:5041/orders với đơn giá âm (-0.01) — nếu đơn vẫn được tạo thì case "
        "đỏ. Hậu quả thật: đơn hàng có tổng tiền âm.",
    (_F, "PlaceFrom_IsExact_ForAmountsThatFloatingPointWouldRound"):
        "Đặt đơn có hai dòng giá 0.10 và 0.20, đọc lại tổng: phải đúng 0.30. Ra 0.30000000000000004 là "
        "case đỏ — tái hiện bằng cách đổi kiểu số tiền trong Orders từ decimal sang double dưới local.",
})

_F = "services/orders/tests/Orders.Api.IntegrationTests/OrderEndpointsTests.cs"
BREAK_HINTS.update({
    (_F, "GetOrder_ReturnsTheOrder_WhenItExists"):
        "Đặt một đơn, ghi lại mã, rồi gọi GET http://localhost:5041/orders/{id} kèm X-Tenant-Id — sai "
        "mã, sai thời điểm đặt, hoặc sai tổng tiền so với lúc tạo đều là case đỏ.",
    (_F, "GetOrder_ReturnsTheTenantTheOrderBelongsTo"):
        "Đọc một đơn tại http://localhost:5041/orders/{id} — nếu phản hồi không có trường tenantId, "
        "hoặc trường đó rỗng, thì case đỏ. Đây là trường mà kịch bản demo dựa vào để chứng minh đơn "
        "hàng thuộc đúng tenant.",
    (_F, "GetOrder_ReturnsNotFound_WhenNoOrderHasThatId"):
        "Gọi GET http://localhost:5041/orders/{GUID không tồn tại} — trả 200 kèm đơn rỗng, hoặc lỗi 500, "
        "thay vì 404 đều là case đỏ.",
})

_F = "services/orders/tests/Orders.Api.IntegrationTests/PlaceOrderTests.cs"
BREAK_HINTS.update({
    (_F, "PlaceOrder_CreatesTheOrder_AndComputesItsTotal"):
        "Gọi POST http://localhost:5041/orders với 2 Notebook và 1 Apron — nếu không nhận 201, hoặc tổng "
        "trả về khác 59.25, thì case đỏ. Thử gửi kèm trường total trong body: nếu server nghe theo giá "
        "trị đó thay vì tự tính thì cũng đỏ.",
    (_F, "PlaceOrder_ReturnsAnIdentifier_ThatReadsBackAsTheSameOrder"):
        "Đặt đơn rồi lấy đúng mã trong phản hồi gọi GET /orders/{id} — nếu không đọc lại được, hoặc "
        "tổng/thời điểm khác lúc tạo, thì case đỏ. Trên giao diện, đây là tình huống mã đơn in cho "
        "khách không tra được.",
    (_F, "PlaceOrder_ReturnsALocationHeader_ForTheCreatedOrder"):
        "Đặt đơn và xem header phản hồi — thiếu header Location, hoặc Location trỏ sai đường dẫn "
        "/orders/{id}, là case đỏ.",
    (_F, "PlaceOrder_Rejects_ARequestWithNoLines"):
        "Gọi POST http://localhost:5041/orders với items rỗng, bỏ qua mọi kiểm tra phía giao diện — nếu "
        "nhận 201 thay vì 400 thì case đỏ. Đây là lớp chặn cuối cùng cho tình huống đặt hàng giỏ rỗng.",
    (_F, "PlaceOrder_Rejects_ALineWithANonPositiveQuantity"):
        "Gọi POST http://localhost:5041/orders với quantity 0 — nhận 201 thay vì 400 là case đỏ.",
    (_F, "PlaceOrder_Fails_WhenNoCallerWasResolved"):
        "Gọi POST http://localhost:5041/orders có X-Tenant-Id nhưng KHÔNG có X-Subject-Id — nếu đơn vẫn "
        "được tạo thì case đỏ, vì khi đó tồn tại đơn hàng không thuộc về ai.",
    (_F, "PlaceOrder_PersistsTheResolvedTenant_OnTheOrderRow"):
        "Đặt đơn rồi mở database orders (localhost,14333) xem cột TenantId của bản ghi vừa tạo — trống "
        "hoặc sai là case đỏ. Khác case đọc qua API ở chỗ đây kiểm tra dữ liệu thật sự nằm trong "
        "database, không phải giá trị được vang lại trong phản hồi.",
    (_F, "PlaceOrder_IgnoresATenantNamedInTheRequestBody"):
        "Gọi POST http://localhost:5041/orders với body cố tình thêm \"tenantId\": \"tenant-khac\" trong "
        "khi header vẫn là contoso, rồi kiểm tra tenant của đơn vừa tạo — nếu ghi theo body thì case đỏ, "
        "và đó là lỗ hổng cho phép người gọi tự chọn tenant cho đơn của mình.",
})

# ---------------------------------------------------------------- ORDER (frontend)

_F = "frontend/apps/web/tests/checkout/Confirmation.test.tsx"
BREAK_HINTS.update({
    (_F, "shows the order identifier verbatim"):
        "Mua hàng tới màn hình xác nhận trên http://localhost:4173 — nếu mã đơn bị rút gọn (kiểu "
        "'aaaaaaaa…') hoặc không hiển thị thì case đỏ, vì khách không còn đọc và trích dẫn được mã đầy "
        "đủ. Tái hiện bằng cách sửa màn hình xác nhận để cắt bớt mã.",
    (_F, "shows the order total in the single Phase 1 currency"):
        "Xem màn hình xác nhận sau khi đặt đơn 59.25 — nếu hiện '59.25' trần, thiếu ký hiệu $, hoặc "
        "thiếu hai chữ số thập phân, thì case đỏ.",
    (_F, "tells the shopper their order was placed"):
        "Xem màn hình xác nhận — nếu không còn tiêu đề báo đơn đã đặt thành công thì case đỏ. Tái hiện "
        "bằng cách đổi tiêu đề màn hình xác nhận thành một câu chung chung không nói đơn đã được đặt.",
    (_F, "shows a nothing-to-show state when there is no order"):
        "Gõ thẳng địa chỉ màn hình xác nhận http://localhost:4173/confirmation khi chưa đặt đơn nào — "
        "nếu màn hình vỡ, hiện lỗi, hoặc tệ hơn là hiện một mã đơn bịa ra, thì case đỏ.",
})

_F = "frontend/apps/web/tests/checkout/DoubleSubmit.test.tsx"
BREAK_HINTS.update({
    (_F, "issues exactly one checkout request when clicked twice in rapid succession"):
        "Mở tab Network của trình duyệt, bấm 'Check out' rồi bấm tiếp ngay lần nữa khi request đầu chưa "
        "xong — nếu thấy hai request POST /bff/checkout thì case đỏ. Tái hiện bằng cách bỏ đoạn vô hiệu "
        "hoá nút trong lúc đang gửi.",
    (_F, "issues one request even when both clicks land before React re-renders"):
        "Bấm hai lần thật nhanh trong cùng một khoảnh khắc (nhanh hơn tốc độ vẽ lại giao diện) — vẫn "
        "phải chỉ một request. Đây đúng là tình huống từng tạo ra hai đơn cách nhau 6 mili giây trên "
        "môi trường container, nên nếu chỉ dựa vào việc nút bị mờ đi thì case đỏ.",
    (_F, "reports the created order exactly once"):
        "Đặt hàng một lần và đếm số lần màn hình xác nhận được kích hoạt — nếu chuyển màn hình hai lần, "
        "hoặc không chuyển lần nào dù đơn đã tạo, thì case đỏ.",
    (_F, "shows an error and reports no order when checkout fails"):
        "Tắt orders-api rồi bấm 'Check out' — nếu không hiện thông báo lỗi, hoặc tệ hơn là vẫn nhảy sang "
        "màn hình xác nhận với một mã đơn không có thật, thì case đỏ.",
})

_F = "frontend/apps/web/tests/checkout/EmptyBasketBlocks.test.tsx"
BREAK_HINTS.update({
    (_F, "is not operable"):
        "Mở http://localhost:4173/basket khi giỏ đang rỗng — nếu nút 'Check out' vẫn bấm được thì case "
        "đỏ. Tái hiện bằng cách bỏ điều kiện vô hiệu hoá nút theo số lượng hàng trong giỏ.",
    (_F, "sends no checkout request when the shopper tries anyway"):
        "Mở tab Network, giỏ rỗng, cố bấm 'Check out' — nếu thấy bất kỳ request POST /bff/checkout nào "
        "thì case đỏ, kể cả khi server đã từ chối bằng 409. Tiêu chí ở đây là không gửi request nào cả.",
    (_F, "becomes operable once the basket holds something"):
        "Thêm 1 sản phẩm vào giỏ — nếu nút 'Check out' vẫn mờ và không bấm được thì case đỏ (chặn nhầm "
        "cả trường hợp hợp lệ).",
})

# ---------------------------------------------------------------- ORDER (E2E)

_F = "frontend/apps/web/e2e/walkthrough.spec.ts"
BREAK_HINTS.update({
    (_F, "browse, add to basket, check out, and see the confirmation"):
        "Chạy trọn luồng mua hàng trên http://localhost:4173 và mở sẵn Console + Network. Case đỏ nếu "
        "bất kỳ điều nào xảy ra: hai Notebook không gộp thành một dòng, tổng khác $59.25, F5 giữa chừng "
        "làm mất giỏ, mã đơn không đọc lại được, giỏ không rỗng sau khi đặt, có request đi tới địa chỉ "
        "khác ngoài gateway, hoặc Console xuất hiện lỗi đỏ.",
    (_F, "checkout is blocked, and unsent, when the basket is empty"):
        "Mở /basket với giỏ rỗng, mở tab Network, cố bấm 'Check out' — thấy request POST /bff/checkout "
        "là case đỏ.",
    (_F, "checking out twice in rapid succession creates exactly one order"):
        "Bấm 'Check out' hai lần gần như đồng thời trên trình duyệt thật, đếm số request POST "
        "/bff/checkout trong tab Network — nhiều hơn một là case đỏ.",
    (_F, "the whole flow can be completed using only the keyboard"):
        "Rút chuột ra, chỉ dùng Tab và Enter để đi từ trang sản phẩm tới màn hình xác nhận. Case đỏ nếu "
        "có điều khiển nào không Tab tới được, hoặc phần tử đang được chọn không hiện viền focus nhìn "
        "thấy được. Tái hiện bằng cách bỏ style :focus-visible trong CSS.",
})

_F = "frontend/apps/web/demo/order-demo.spec.ts"
BREAK_HINTS.update({
    (_F, "one order, placed end to end, on the running stack"):
        "Chạy ./scripts/demo.ps1 nhưng bỏ bước dọn giỏ, hoặc chạy khi stack container chưa lên hẳn — "
        "case đỏ ngay ở bước đầu vì giỏ không rỗng. Ngoài ra case đỏ nếu tổng khác $59.25, đơn không "
        "đọc lại được qua gateway, giỏ không rỗng sau khi đặt, hoặc 4 ảnh chụp trong docs/demo/ không "
        "được sinh ra.",
})

# ---------------------------------------------------------------- PARTY

_F = "services/parties/tests/Parties.Api.IntegrationTests/PartyEndpointsTests.cs"
BREAK_HINTS.update({
    (_F, "GetParty_ReturnsTheParty_WhenItExists"):
        "Cần có sẵn dữ liệu: chèn tay một bản ghi vào bảng Parties trong database parties "
        "(localhost,14330), rồi gọi GET http://localhost:5204/parties/{id} kèm X-Tenant-Id — sai id hoặc "
        "sai DisplayName so với bản ghi đã chèn là case đỏ. Lưu ý bảng này không có dữ liệu seed và "
        "cũng không có API tạo party, nên phải chèn bằng SQL.",
    (_F, "GetParty_ReturnsNotFound_WhenNoPartyHasThatId"):
        "Gọi GET http://localhost:5204/parties/{GUID không tồn tại} kèm X-Tenant-Id — trả 200 kèm dữ "
        "liệu rỗng, hoặc lỗi 500, thay vì 404 đều là case đỏ.",
})

# ---------------------------------------------------------------- PRODUCT (backend)

_F = "services/products/tests/Products.Api.IntegrationTests/CatalogEndpointsTests.cs"
BREAK_HINTS.update({
    (_F, "GetProducts_ReturnsEveryProduct_WithIdNameAndPrice"):
        "Gọi GET http://localhost:5088/products kèm X-Tenant-Id và đối chiếu từng sản phẩm với bảng "
        "Products trong database (localhost,14331) — thiếu sản phẩm, thiếu trường name hoặc price, hay "
        "sai giá đều là case đỏ. Trên giao diện, biểu hiện là danh sách sản phẩm có ô trống.",
    (_F, "GetProducts_ReturnsEmptyArray_WhenCatalogIsEmpty"):
        "Xoá hết bản ghi trong bảng Products rồi gọi GET http://localhost:5088/products — nếu trả 404 "
        "hoặc lỗi thay vì mảng rỗng thì case đỏ, vì catalog rỗng là trạng thái hợp lệ chứ không phải "
        "sự cố.",
})

_F = "services/products/tests/Products.Api.IntegrationTests/CatalogSeedTests.cs"
BREAK_HINTS.update({
    (_F, "ApplyingMigrations_SeedsTheCatalog_WithTheThreeKnownProducts"):
        "Dựng lại stack từ database trắng (`./scripts/local-down.ps1 -DiscardData` rồi "
        "`./scripts/local-up.ps1`) và gọi GET http://localhost:5088/products — nếu thiếu bất kỳ sản "
        "phẩm nào trong ba sản phẩm mẫu, hoặc tên/giá khác (12.50 / 48.00 / 34.25), thì case đỏ. Đó "
        "cũng là lúc kịch bản demo và Playwright hỏng theo vì chúng chọn sản phẩm theo tên.",
    (_F, "ApplyingMigrations_LeavesAPurchasableProduct_WithoutAnyManualSetup"):
        "Dựng lại từ database trắng rồi mở http://localhost:4173 — nếu trang sản phẩm trống, hoặc có "
        "sản phẩm nhưng thiếu tên / giá bằng 0, thì case đỏ. Tiêu chí là mở lên mua được ngay, không "
        "phải nhập liệu tay.",
    (_F, "TheSeededIdentifiers_AreStableAcrossFreshDatabases"):
        "Ghi lại mã ba sản phẩm, dựng lại từ database trắng, rồi so mã lần nữa — nếu mã đổi sau mỗi lần "
        "dựng thì case đỏ, vì mọi tài liệu và kịch bản test đang trỏ tới các mã cố định đó.",
})

# ---------------------------------------------------------------- PRODUCT (frontend)

_F = "frontend/apps/web/tests/catalog/ProductList.test.tsx"
BREAK_HINTS.update({
    (_F, "lists every product with its name and price"):
        "Mở http://localhost:4173 — nếu thiếu sản phẩm nào, hoặc giá hiện dạng thô như '48' thay vì "
        "'$48.00', thì case đỏ. Tái hiện bằng cách bỏ bước định dạng tiền khi hiển thị.",
    (_F, "presents the catalog as a list"):
        "Dùng trình đọc màn hình (hoặc xem cấu trúc trong DevTools) trên trang sản phẩm — nếu danh sách "
        "không còn được đánh dấu là list có nhãn thì case đỏ, người dùng khiếm thị mất khả năng biết "
        "trước có bao nhiêu sản phẩm.",
    (_F, "shows the empty state when the catalog holds nothing"):
        "Xoá hết sản phẩm trong database rồi mở trang chủ — nếu thấy trang trắng, hoặc vòng xoay chờ "
        "mãi không dứt, thay vì thông báo chưa có sản phẩm thì case đỏ.",
    (_F, "shows a readable error when the backend fails"):
        "Tắt products-api rồi mở trang chủ — nếu trang trắng, hoặc quay vòng vô tận, hoặc có thông báo "
        "lỗi nhưng không có nút thử lại, thì case đỏ.",
    (_F, "requests the catalog from the configured gateway origin"):
        "Mở tab Network khi vào trang chủ — nếu thấy request đi thẳng tới cổng 5088 (products) hay bất "
        "kỳ địa chỉ nào ngoài gateway 5300 thì case đỏ. Tái hiện bằng cách viết cứng một URL khác trong "
        "component thay vì dùng địa chỉ đã cấu hình.",
})

_F = "frontend/apps/web/tests/catalog/EmptyCatalog.test.tsx"
BREAK_HINTS.update({
    (_F, "tells the shopper there is nothing to buy yet"):
        "Xoá hết sản phẩm rồi mở trang chủ — không thấy dòng thông báo chưa có sản phẩm là case đỏ.",
    (_F, "does not present itself as an error"):
        "Ở trạng thái catalog rỗng, kiểm tra xem thông báo có bị đánh dấu là cảnh báo lỗi không — nếu "
        "trình đọc màn hình đọc nó như một thông báo lỗi cắt ngang thì case đỏ, vì catalog rỗng là "
        "chuyện bình thường chứ không phải sự cố.",
})

_F = "frontend/apps/web/tests/catalog/CatalogError.test.tsx"
BREAK_HINTS.update({
    (_F, "shows a readable message and announces it"):
        "Tắt products-api rồi mở trang chủ — nếu thông báo lỗi không được đánh dấu để trình đọc màn hình "
        "công bố, hoặc nội dung không đúng thông điệp, thì case đỏ.",
    (_F, "offers a retry the shopper can operate"):
        "Ở màn hình lỗi, bấm nút 'Try again' — nếu bấm không có phản ứng gì thì case đỏ.",
    (_F, "reaches and fires retry by keyboard alone"):
        "Ở màn hình lỗi, chỉ dùng Tab rồi Enter để kích hoạt nút thử lại — nếu không Tab tới được, hoặc "
        "Enter không kích hoạt, thì case đỏ. Đường phục hồi sau lỗi không được chỉ dành cho chuột.",
    (_F, "omits the retry control when no retry is possible"):
        "Ở trạng thái lỗi không thể thử lại, nếu vẫn hiện một nút bấm vào không làm gì thì case đỏ. Tái "
        "hiện bằng cách luôn vẽ nút thử lại bất kể có hành động phía sau hay không.",
})

# ---------------------------------------------------------------- COMMON: BFF

_F = "services/bff/tests/Bff.Api.UnitTests/DownstreamServiceClientOptionsTests.cs"
BREAK_HINTS.update({
    (_F, "Validation_Fails_WhenBaseUrlIsMissing"):
        "Xoá biến môi trường Services__ProductsApi__BaseUrl của bff-api trong docker-compose.local.yml "
        "rồi khởi động lại stack — nếu BFF vẫn lên và vẫn báo healthy, chỉ hỏng khi có người gọi, thì "
        "case đỏ. Nó phải từ chối khởi động, và thông báo phải nêu đúng tên khoá cấu hình thiếu.",
    (_F, "Validation_Fails_WhenBaseUrlIsNotAnAbsoluteUri"):
        "Đặt Services__ProductsApi__BaseUrl thành một đường dẫn tương đối như '/products' rồi khởi động "
        "lại — nếu BFF khởi động bình thường thì case đỏ.",
    (_F, "Validation_Succeeds_WhenBaseUrlIsAnAbsoluteUri"):
        "Đặt BaseUrl thành một địa chỉ tuyệt đối hợp lệ (http://products-api:8080) — nếu BFF vẫn từ chối "
        "khởi động thì case đỏ (chặn nhầm cấu hình đúng). Tái hiện bằng cách siết luật kiểm tra quá tay, "
        "ví dụ bắt buộc phải là https.",
})

_F = "services/bff/tests/Bff.Api.UnitTests/ResponseMappingTests.cs"
BREAK_HINTS.update({
    (_F, "ProductSummary_CarriesEveryFieldFromTheDownstreamProduct"):
        "So dữ liệu tại http://localhost:5088/products (gọi thẳng) với http://localhost:5301/bff/products "
        "— nếu qua BFF bị mất trường, hoặc tên và giá bị tráo chỗ cho nhau, thì case đỏ. Trên giao diện, "
        "biểu hiện là thẻ sản phẩm hiện tên ở ô giá.",
    (_F, "BasketItem_JoinsTheProductName_AndPassesEveryOtherFieldThrough"):
        "So một dòng giỏ tại http://localhost:5188/baskets/current với cùng dòng đó qua "
        "http://localhost:5301/bff/basket — bản qua BFF phải có thêm trường name, các trường còn lại "
        "giữ nguyên. Thiếu name (giao diện hiện GUID thay vì tên sản phẩm), hoặc thành tiền bị tính lại "
        "khác đi, đều là case đỏ.",
    (_F, "BasketItem_SurvivesAProductMissingFromTheCatalog"):
        "Thêm một sản phẩm vào giỏ, rồi xoá chính sản phẩm đó khỏi bảng Products (localhost,14331), rồi "
        "mở trang giỏ hàng — nếu dòng đó biến mất khỏi giỏ thì case đỏ, vì khách vẫn đang bị tính tiền "
        "cho nó mà không thấy nó đâu. Nó phải ở lại kèm tên thay thế.",
    (_F, "OrderResponse_CarriesEveryFieldFromTheDownstreamOrder"):
        "So một đơn tại http://localhost:5041/orders/{id} với cùng đơn đó qua "
        "http://localhost:5301/bff/orders/{id} — mất trường, hoặc tổng tiền và thời điểm đặt bị tráo, "
        "là case đỏ.",
    (_F, "OrderResponse_PreservesTheUtcKindOfThePlacedTimestamp"):
        "Đặt đơn rồi xem thời điểm đặt hiển thị trên màn hình xác nhận — nếu lệch múi giờ so với giờ "
        "thật thì case đỏ. Tái hiện bằng cách bỏ nhãn UTC của mốc thời gian ở bước ánh xạ trong BFF.",
    (_F, "PartyResponse_CarriesEveryFieldFromTheDownstreamParty"):
        "So dữ liệu party gọi thẳng http://localhost:5204/parties/{id} với qua "
        "http://localhost:5301/bff/parties/{id} — mất id hoặc mất DisplayName là case đỏ. Cần chèn tay "
        "một party vào database trước vì bảng này không có dữ liệu seed.",
    (_F, "ProductSummary_PreservesPricePrecisionExactly"):
        "Sửa giá một sản phẩm trong database thành số nhiều chữ số như 999999999999.99 hoặc 0.01, rồi "
        "xem giá đó qua http://localhost:5301/bff/products — sai số, hoặc mất số 0 ở cuối (12.50 thành "
        "12.5), là case đỏ. Tái hiện bằng cách đổi kiểu số tiền trong BFF từ decimal sang double.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/BasketFlowTests.cs"
BREAK_HINTS.update({
    (_F, "GetBasket_ReturnsAnEmptyBasket_ForAShopperWhoHasAddedNothing"):
        "Dọn giỏ rồi gọi http://localhost:5301/bff/basket kèm header tenant và subject — nếu trả 404 "
        "hoặc lỗi thay vì giỏ rỗng tổng 0 thì case đỏ.",
    (_F, "AddItem_ReturnsTheBasket_WithTheProductsNameAndResolvedPrice"):
        "Gọi POST http://localhost:5301/bff/basket/items chỉ với productId và quantity — nếu dòng trả về "
        "thiếu tên sản phẩm, hoặc đơn giá khác 12.50 lấy từ catalog, thì case đỏ.",
    (_F, "AddItem_IgnoresAPriceSuppliedByTheClient"):
        "Gọi POST /bff/basket/items kèm thêm \"unitPrice\": 0.01 trong body — nếu giỏ ghi nhận giá 0.01 "
        "thay vì 12.50 thì case đỏ, và đó là lỗ hổng cho phép khách tự đặt giá cho mình.",
    (_F, "AddItem_MergesIntoTheExistingLine_WhenTheSameProductIsAddedAgain"):
        "Gọi POST /bff/basket/items hai lần cho cùng sản phẩm rồi đọc giỏ qua BFF — ra 2 dòng thay vì 1 "
        "dòng số lượng 2, tổng 25.00, là case đỏ.",
    (_F, "AddItem_ReturnsNotFound_WhenNoSuchProductExists"):
        "Gọi POST /bff/basket/items với một productId ngẫu nhiên không có trong catalog — nếu trả 502 "
        "(coi như downstream hỏng) hoặc thêm thành công thì case đỏ; phải là 404.",
    (_F, "AddItem_Rejects_AQuantityBelowOne"):
        "Gọi POST /bff/basket/items với quantity 0 hoặc -2 — nhận 200 hoặc 500 thay vì 400 là case đỏ.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/BasketsRouteTests.cs"
BREAK_HINTS.update({
    (_F, "GetBasket_ReturnsShapedBasketFromTheBasketsService"):
        "Tạo một giỏ rỗng, lấy id rồi gọi http://localhost:5301/bff/baskets/{id} — sai id, sai "
        "CustomerRef, hoặc danh sách dòng không rỗng đều là case đỏ.",
    (_F, "GetBasket_ReturnsNotFound_WhenTheBasketsServiceHasNoSuchBasket"):
        "Gọi http://localhost:5301/bff/baskets/{GUID không tồn tại} — trả 502 hay 200 rỗng thay vì 404 "
        "là case đỏ. Case này canh việc không lẫn lộn 'dịch vụ hỏng' với 'không có dữ liệu'.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/CheckoutTests.cs"
BREAK_HINTS.update({
    (_F, "Checkout_CreatesAnOrder_ForWhatIsInTheBasket"):
        "Thêm 2 Notebook và 1 Apron rồi gọi POST http://localhost:5300/bff/checkout — không nhận 201, "
        "hoặc tổng đơn khác 59.25, là case đỏ. Đây là con số đi qua đủ ba service.",
    (_F, "Checkout_ReturnsAReference_ThatReadsBackAsTheSameOrder"):
        "Đặt hàng rồi lấy mã đơn trong màn hình xác nhận gọi GET /bff/orders/{id} — không đọc lại được, "
        "hoặc tổng khác, là case đỏ.",
    (_F, "Checkout_EmptiesTheBasket"):
        "Đặt hàng xong rồi mở lại trang giỏ — nếu hàng vẫn còn nguyên trong giỏ thì case đỏ. Ngoài việc "
        "gây nhầm lẫn cho khách, đây còn là thứ khiến bấm đặt hàng lần hai tạo được đơn thứ hai.",
    (_F, "Checkout_ReturnsConflict_WhenTheBasketIsEmpty"):
        "Gọi POST http://localhost:5300/bff/checkout khi giỏ đang rỗng, bỏ qua giao diện — nếu tạo ra "
        "một đơn rỗng thay vì trả 409 thì case đỏ.",
    (_F, "Checkout_CreatesExactlyOneOrder_WhenAttemptedTwice"):
        "Gọi POST /bff/checkout hai lần liên tiếp cho cùng một giỏ — nếu lần thứ hai cũng trả 201 thì "
        "case đỏ: khách bị tính tiền hai lần.",
    (_F, "Checkout_OrdersOnlyTheCallersOwnBasket"):
        "Cho hai subject khác nhau (alice, bob) mỗi người một giỏ có hàng khác nhau, rồi cho alice đặt "
        "hàng — nếu đơn của alice gồm cả hàng của bob, hoặc giỏ của bob bị dọn theo, thì case đỏ.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/DownstreamUnavailableTests.cs"
BREAK_HINTS.update({
    (_F, "GetProducts_ReturnsBadGateway_WhenTheProductsServiceIsUnreachable"):
        "Tắt products-api rồi gọi http://localhost:5301/bff/products và bấm giờ — nếu treo quá 5 giây, "
        "hoặc trả 200 với dữ liệu cũ, hoặc trả 500 trần, thì case đỏ. Lưu ý dừng container thường cho "
        "504 (hết thời gian chờ) thay vì 502 (không kết nối được); cả hai đều là câu trả lời hợp lệ.",
    (_F, "GetProducts_ReturnsGatewayTimeout_WhenTheProductsServiceNeverAnswers"):
        "Làm cho products-api nhận request nhưng không bao giờ trả lời (ví dụ tạm dừng container bằng "
        "`docker compose -f docker-compose.local.yml pause products-api`) rồi gọi /bff/products — nếu "
        "treo quá 5 giây thì case đỏ. Nhớ `unpause` sau khi thử.",
    (_F, "ADownstreamFailure_ReturnsProblemDetailsCarryingTheCorrelationId"):
        "Tắt products-api rồi đọc nội dung lỗi từ /bff/products — thiếu correlationId, hoặc "
        "correlationId khác với header X-Correlation-Id của cùng phản hồi, là case đỏ. Khi đó mã tra "
        "cứu đưa cho khách vô dụng.",
    (_F, "ADownstreamFailure_NamesTheLogicalServiceOnly_NeverItsAddress"):
        "Tắt products-api rồi đọc kỹ nội dung lỗi — nếu thấy tên host, chuỗi http://, tên thư viện "
        "Polly, hay stack trace, thì case đỏ. Phải nêu được tên logic 'ProductsApi' nhưng không lộ "
        "địa chỉ.",
    (_F, "EveryRoute_FailsAsAProblemDetails_WhenItsDownstreamIsUnreachable"):
        "Lần lượt tắt baskets-api, orders-api, parties-api rồi gọi route tương ứng qua BFF — route nào "
        "trả 500 trần hoặc nội dung không phải dạng problem+json thì case đỏ. Case này bắt tình huống "
        "một route bị quên nối vào cơ chế xử lý lỗi chung.",
    (_F, "ADownstreamFailure_IsBoundedAndStructured_AgainstARealUnreachableHost"):
        "Trỏ Services__ProductsApi__BaseUrl của bff-api sang một tên miền không tồn tại rồi gọi "
        "/bff/products — case đỏ nếu treo quá 5 giây, hoặc nội dung không phải problem+json, hoặc không "
        "nêu tên ProductsApi. Mã 502 hay 504 đều chấp nhận được.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/GeneratedContractTests.cs"
BREAK_HINTS.update({
    (_F, "TheDocument_DescribesEveryClientFacingRoute"):
        "Mở http://localhost:5301/openapi/v1.json và đối chiếu với 7 đường dẫn /bff/* mà BFF thực sự "
        "phục vụ — thiếu đường dẫn nào là case đỏ, vì mã gọi API của giao diện được sinh ra từ chính "
        "tài liệu này nên đường dẫn thiếu sẽ không có hàm để gọi.",
    (_F, "EveryRoute_DeclaresItsDownstreamFailureResponses"):
        "Trong tài liệu OpenAPI, xem phần responses của từng route GET — thiếu khai báo 502 hoặc 504 là "
        "case đỏ. Hậu quả: giao diện được sinh ra không có kiểu dữ liệu cho tình huống downstream hỏng, "
        "dù thực tế BFF vẫn trả về những mã đó.",
    (_F, "TheByIdRoutes_DeclareNotFound"):
        "Xem responses của ba route theo id (baskets, orders, parties) — thiếu khai báo 404 là case đỏ, "
        "vì giao diện sẽ coi 'không tìm thấy' như một sự cố thay vì một kết quả bình thường.",
    (_F, "TheProductListingSchema_MatchesTheHandAuthoredContract"):
        "Trong tài liệu OpenAPI, kiểm tra schema ProductListResponse có thuộc tính items và "
        "ProductSummary có đủ id, name, price — thiếu bất kỳ trường nào là case đỏ.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/OrdersRouteTests.cs"
BREAK_HINTS.update({
    (_F, "GetOrder_ReturnsShapedOrderFromTheOrdersService"):
        "Đặt một đơn rồi so http://localhost:5041/orders/{id} với http://localhost:5301/bff/orders/{id} "
        "— sai id, sai thời điểm đặt, hoặc sai tổng tiền là case đỏ.",
    (_F, "GetOrder_ReturnsNotFound_WhenTheOrdersServiceHasNoSuchOrder"):
        "Gọi http://localhost:5301/bff/orders/{GUID không tồn tại} — trả 502 hay 200 rỗng thay vì 404 là "
        "case đỏ.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/PartiesRouteTests.cs"
BREAK_HINTS.update({
    (_F, "GetParty_ReturnsShapedPartyFromThePartiesService"):
        "Chèn tay một party vào database parties (localhost,14330) rồi so kết quả gọi thẳng "
        "http://localhost:5204/parties/{id} với qua http://localhost:5301/bff/parties/{id} — sai id "
        "hoặc mất DisplayName là case đỏ.",
    (_F, "GetParty_ReturnsNotFound_WhenThePartiesServiceHasNoSuchParty"):
        "Gọi http://localhost:5301/bff/parties/{GUID không tồn tại} — trả 502 hay 200 rỗng thay vì 404 "
        "là case đỏ.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/ProductsRouteTests.cs"
BREAK_HINTS.update({
    (_F, "GetProducts_ReturnsShapedListingFromTheProductsService"):
        "So từng sản phẩm giữa http://localhost:5088/products và http://localhost:5301/bff/products — "
        "thiếu sản phẩm, sai tên, hoặc sai giá là case đỏ.",
    (_F, "GetProducts_ReturnsEmptyItemsEnvelope_WhenTheCatalogIsEmpty"):
        "Xoá hết sản phẩm trong database rồi gọi http://localhost:5301/bff/products — nếu trả về mảng "
        "trần [] hoặc null thay vì đối tượng {\"items\": []} thì case đỏ, vì giao diện luôn đọc trường "
        "items mà không kiểm tra hình dạng phản hồi.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/SubjectPropagationTests.cs"
BREAK_HINTS.update({
    (_F, "TheBffsOutboundCall_CarriesTheSubjectTheBffReceived"):
        "Gọi http://localhost:5301/bff/basket kèm đủ header tenant và subject — nếu trả về lỗi 500 hoặc "
        "502 thì case đỏ, dấu hiệu BFF nhận được subject nhưng không chuyển tiếp xuống baskets service. "
        "Tái hiện bằng cách gỡ bước truyền header ở BFF dưới local.",
    (_F, "TheBffsOutboundCall_CarriesNoSubject_WhenTheBffItselfHasNone"):
        "Gọi http://localhost:5301/bff/basket chỉ có tenant, KHÔNG có subject — nếu trả về một giỏ nào "
        "đó thay vì lỗi thì case đỏ, vì BFF đã tự bịa ra người mua mặc định.",
})

_F = "services/bff/tests/Bff.Api.IntegrationTests/TenantPropagationTests.cs"
BREAK_HINTS.update({
    (_F, "TheBffsOutboundCall_CarriesTheTenantTheBffReceived"):
        "Gọi http://localhost:5301/bff/products kèm X-Tenant-Id — nếu trả 502 thì case đỏ, dấu hiệu BFF "
        "nhận được tenant nhưng không chuyển tiếp xuống products service.",
    (_F, "TheBffsOutboundCall_CarriesNoTenant_WhenTheBffItselfHasNone"):
        "Gọi http://localhost:5301/bff/products KHÔNG kèm header nào — nếu trả 200 kèm danh sách sản "
        "phẩm thì case đỏ, vì BFF đã tự điền một tenant mặc định. Kết quả đúng là thất bại rõ ràng "
        "(502 hoặc 504).",
})

# ---------------------------------------------------------------- COMMON: Tenancy

_F = "shared/Tenancy.UnitTests/CallerContextTests.cs"
BREAK_HINTS.update({
    (_F, "RequireSubjectId_ReturnsTheResolvedSubject_WhenOneHasBeenSet"):
        "Gọi http://localhost:5188/baskets/current kèm 'X-Subject-Id: alice' — nếu giỏ trả về "
        "customerRef khác alice thì case đỏ, dấu hiệu subject phân giải ra không đúng giá trị nhận được.",
    (_F, "CallerContext_IsUnresolved_BeforeAnythingSetsIt"):
        "Sửa Tenancy dưới local để CallerContext có sẵn một subject mặc định lúc khởi tạo — case đỏ. "
        "Hậu quả thật: mọi request không có header vẫn được coi là của một người mua nào đó.",
    (_F, "RequireSubjectId_Throws_WhenNoSubjectHasBeenResolved"):
        "Gọi http://localhost:5188/baskets/current chỉ có tenant, không có subject — nếu trả về một giỏ "
        "thay vì lỗi 500 thì case đỏ.",
    (_F, "RequireSubjectId_Throws_WhenTheResolvedSubjectIsBlank"):
        "Gọi http://localhost:5188/baskets/current kèm header 'X-Subject-Id:' để trống — nếu vẫn trả về "
        "một giỏ thì case đỏ. Đây là tình huống nguy hiểm nhất: mọi request có subject rỗng sẽ dùng "
        "chung đúng một giỏ hàng.",
})

_F = "shared/Tenancy.UnitTests/CallerContextMiddlewareTests.cs"
BREAK_HINTS.update({
    (_F, "InvokeAsync_ResolvesTheCallerContext_FromTheInboundHeader"):
        "Gọi http://localhost:5188/baskets/current kèm 'X-Subject-Id: alice' rồi lại 'X-Subject-Id: bob' "
        "— nếu hai lần ra cùng một giỏ thì case đỏ, dấu hiệu header không được đọc vào ngữ cảnh request.",
    (_F, "InvokeAsync_LeavesTheCallerContextUnresolved_WhenTheHeaderIsAbsentOrEmpty"):
        "Gọi lần lượt: không có header subject, có header rỗng, có header chỉ khoảng trắng — cả ba đều "
        "phải lỗi 500. Trường hợp nào trả về giỏ là case đỏ.",
    (_F, "InvokeAsync_PushesTheResolvedSubjectIntoTheLoggingScope"):
        "Gọi một request có subject rồi xem log của service (`docker compose -f docker-compose.local.yml "
        "logs baskets-api`) — nếu các dòng log của request đó không kèm SubjectId thì case đỏ, và việc "
        "điều tra khiếu nại của một khách cụ thể trở nên bất khả thi.",
    (_F, "InvokeAsync_PushesNoSubjectScope_WhenTheRequestIsUnresolved"):
        "Gọi một request không có subject rồi xem log — nếu thấy SubjectId rỗng được ghi ra như thể có "
        "người gọi thì case đỏ. Sự vắng mặt mới là tín hiệu đúng.",
    (_F, "InvokeAsync_AlwaysCallsTheRestOfThePipeline"):
        "Sửa middleware dưới local để chặn request ngay khi thiếu subject, rồi gọi "
        "http://localhost:5188/health/live — nếu health probe cũng bị chặn thì case đỏ. Hậu quả thật: "
        "mọi container bị coi là chết và khởi động lại liên tục.",
})

_F = "shared/Tenancy.UnitTests/TenantContextTests.cs"
BREAK_HINTS.update({
    (_F, "RequireTenantId_ReturnsTheResolvedTenant_WhenOneHasBeenSet"):
        "Đặt một đơn rồi đọc lại tại http://localhost:5041/orders/{id} — nếu trường tenantId khác giá "
        "trị gửi trong header X-Tenant-Id thì case đỏ.",
    (_F, "TenantContext_IsUnresolved_BeforeAnythingSetsIt"):
        "Sửa Tenancy dưới local để TenantContext có sẵn tenant mặc định lúc khởi tạo — case đỏ. Hậu quả "
        "thật: request không đi qua gateway vẫn đọc được dữ liệu của tenant mặc định đó.",
    (_F, "RequireTenantId_Throws_WhenNoTenantHasBeenResolved"):
        "Gọi http://localhost:5088/products không kèm X-Tenant-Id — nếu trả 200 kèm catalog thay vì lỗi "
        "500 thì case đỏ.",
    (_F, "RequireTenantId_Throws_WhenTheResolvedTenantIsBlank"):
        "Gọi http://localhost:5088/products kèm 'X-Tenant-Id:' để trống — nếu vẫn trả về dữ liệu thì "
        "case đỏ. Một tenant rỗng phải bị coi là chưa phân giải, không phải một tenant tên rỗng.",
})

_F = "shared/Tenancy.UnitTests/TenantContextMiddlewareTests.cs"
BREAK_HINTS.update({
    (_F, "InvokeAsync_ResolvesTheTenantContext_FromTheInboundHeader"):
        "Gọi http://localhost:5088/products kèm X-Tenant-Id hợp lệ — nếu vẫn lỗi 500 thì case đỏ, dấu "
        "hiệu header không được đọc vào ngữ cảnh request.",
    (_F, "InvokeAsync_LeavesTheTenantContextUnresolved_WhenTheHeaderIsAbsentOrEmpty"):
        "Gọi lần lượt: không có header tenant, header rỗng, header chỉ khoảng trắng — cả ba đều phải "
        "lỗi 500. Trường hợp nào trả về dữ liệu là case đỏ.",
    (_F, "InvokeAsync_PushesTheResolvedTenantIntoTheLoggingScope"):
        "Gọi một request có tenant rồi xem log của service — nếu các dòng log không kèm TenantId thì "
        "case đỏ.",
    (_F, "InvokeAsync_PushesNoTenantScope_WhenTheRequestIsUnresolved"):
        "Gọi một request không có tenant rồi xem log — nếu thấy TenantId rỗng được ghi ra như thể request "
        "thuộc về 'tenant rỗng' thì case đỏ.",
    (_F, "InvokeAsync_AlwaysCallsTheRestOfThePipeline"):
        "Sửa middleware dưới local để trả lỗi ngay khi thiếu tenant, rồi gọi "
        "http://localhost:5088/health/live — health probe bị chặn theo là case đỏ.",
})

# ---------------------------------------------------------------- COMMON: quy ước kiến trúc

_F = "tests/ContainerConventionTests/DockerfileSharedProjectTests.cs"
BREAK_HINTS.update({
    (_F, "EveryServiceImage_ReceivesEverySharedProject_ItCompilesAgainst"):
        "Xoá một dòng COPY thư viện dùng chung (ví dụ shared/Tenancy) khỏi Dockerfile của một service, "
        "rồi chạy `./scripts/local-up.ps1` — case đỏ, và image đó cũng không build được. Đây đúng là "
        "lỗi từng làm 5/6 image hỏng mà không bài test nào phát hiện.",
    (_F, "TheScan_Examined_EveryService"):
        "Thêm một thư mục service thứ 7 dưới services/, hoặc xoá Dockerfile của một service hiện có — "
        "case đỏ. Case này canh cho việc bộ quét không âm thầm bỏ sót service nào.",
    (_F, "TheScan_Observed_SharedProjectReferences"):
        "Sửa bộ quét dưới local để nó không nhận diện được dòng tham chiếu thư viện nào nữa — case đỏ. "
        "Nếu không có case này, một bộ quét bị mù sẽ báo 'không có vi phạm' và trông hệt như repo sạch.",
    (_F, "ServicesThatUseTheTenancyLibrary_CopyIt"):
        "Xoá dòng COPY shared/Tenancy khỏi Dockerfile của baskets, bff, orders, parties hoặc products — "
        "case đỏ và nêu đích danh service vừa bị sửa.",
    (_F, "TheGateway_DoesNotReferenceTheTenancyLibrary"):
        "Thêm tham chiếu shared/Tenancy vào Gateway.Api.csproj — case đỏ. Gateway là bên sinh ra header "
        "tenant/subject chứ không đọc chúng, nên nó không được dùng thư viện đọc.",
})

_F = "tests/CrossServiceIsolation.Tests/ConnectionStringIsolationTests.cs"
BREAK_HINTS.update({
    (_F, "NoServiceConfiguration_NamesAnotherServicesDatabase"):
        "Sửa appsettings của một service để chuỗi kết nối trỏ sang database của service khác (ví dụ cho "
        "Baskets trỏ vào database orders) — case đỏ. Đây là ranh giới cấm tuyệt đối: mỗi service chỉ "
        "được chạm database của chính nó.",
    (_F, "Scan_ActuallyExaminesEveryServicesConfiguration"):
        "Thêm một thư mục service thứ 7, hoặc đổi tên thư mục src/ của một service khiến bộ quét không "
        "tìm thấy file cấu hình — case đỏ.",
    (_F, "NoStatelessService_DeclaresAConnectionString"):
        "Thêm một mục ConnectionStrings vào appsettings của bff hoặc gateway — case đỏ. Hai service này "
        "không sở hữu dữ liệu nên không được cấp database, kể cả database của chính chúng.",
    (_F, "Scan_FlagsAConfigurationThatReachesAnotherServicesDatabase"):
        "Sửa bộ quét dưới local để nó luôn trả về 'không có vi phạm' — case đỏ. Case này dựng sẵn một "
        "cây thư mục giả có lỗi cố ý, nên nó bắt được một bộ quét đã mất khả năng phát hiện.",
    (_F, "Scan_AllowsAServiceToNameItsOwnDatabase"):
        "Sửa bộ quét để nó báo lỗi cả khi service trỏ đúng database của mình — case đỏ (báo động giả). "
        "Cặp với case trên để giữ độ nhạy của bộ quét ở đúng mức.",
})

_F = "tests/CrossServiceIsolation.Tests/TenantGatedConnectionTests.cs"
BREAK_HINTS.update({
    (_F, "EveryDatabaseOwningService_HasExactlyOneDbContextRegistration"):
        "Thêm một điểm đăng ký DbContext thứ hai vào Program.cs của một service sở hữu database — case "
        "đỏ. Nhiều cửa vào database nghĩa là nhiều chỗ phải canh, và sớm muộn có chỗ bị quên.",
    (_F, "EveryDbContextRegistration_IsGatedOnAResolvedTenant"):
        "Xoá lời gọi RequireTenantId() ở điểm đăng ký DbContext của một service — case đỏ. Sau đó gọi "
        "API của service đó không kèm header tenant sẽ trả về dữ liệu thay vì lỗi, đúng lỗ hổng mà case "
        "này chặn.",
    (_F, "NoStatelessService_RegistersADbContext"):
        "Thêm một đăng ký DbContext vào bff hoặc gateway — case đỏ.",
    (_F, "Scan_ActuallyExaminesEveryServicesRegistration"):
        "Thêm thư mục service thứ 7, hoặc đổi cấu trúc thư mục khiến bộ quét không tìm thấy Program.cs "
        "— case đỏ.",
    (_F, "Scan_FlagsAnUngatedRegistration"):
        "Sửa bộ quét dưới local để nó coi mọi đăng ký DbContext là đã được canh — case đỏ, vì case này "
        "dựng sẵn một Program.cs giả cố tình không canh.",
    (_F, "Scan_AcceptsAGatedRegistration"):
        "Sửa bộ quét để nó không nhận ra lời gọi RequireTenantId() hợp lệ — case đỏ (báo động giả trên "
        "mã đúng).",
    (_F, "Scan_DoesNotAcceptAGuardThatOnlyAppearsInAComment"):
        "Sửa bộ quét để nó đếm cả những lần RequireTenantId xuất hiện trong dòng chú thích — case đỏ. "
        "Nếu không, xoá cổng chặn thật mà để lại comment mô tả nó vẫn qua được kiểm tra.",
})

_F = "tests/StructureConventionTests/VerticalSliceStructureTests.cs"
BREAK_HINTS.update({
    (_F, "NoService_HasATopLevelTechnicalLayerFolder"):
        "Tạo một thư mục rỗng tên Controllers, Services hoặc Repositories ngay dưới một dự án *.Api "
        "(ví dụ `mkdir services/products/src/Products.Api/Controllers`) — case đỏ. Chỉ cần thư mục "
        "rỗng là đủ, và vì git không theo dõi thư mục rỗng nên loại lỗi này không thấy được qua diff.",
    (_F, "Scan_ActuallyExaminesEveryServicesApiProject"):
        "Tạo thêm một thư mục bất kỳ dưới services/ (ví dụ `mkdir services/zzz-temp`) — case đỏ vì bộ "
        "quét đếm được 7 service thay vì 6. Thêm service thật cũng làm đỏ, và đó là cố ý: người thêm "
        "buộc phải vào cập nhật danh sách.",
    (_F, "EveryService_OrganisesAtLeastOneCapabilityUnderFeatures"):
        "Đổi tên thư mục Features của một service (ví dụ Gateway.Api/Features thành Features_x) — case "
        "đỏ. Case này bắt tình huống ngược với case trên: service không có gì tổ chức theo nghiệp vụ.",
    (_F, "Scan_FlagsATopLevelTechnicalLayerFolder"):
        "Sửa bộ quét dưới local để so tên thư mục có phân biệt hoa thường — case đỏ ở đúng bộ dữ liệu "
        "'repositories' viết thường (3 bộ kia vẫn xanh). Trên Windows tên thư mục không phân biệt hoa "
        "thường nên bỏ sót nửa số trường hợp thật.",
    (_F, "Scan_AllowsCapabilityFoldersAndNonLayerFolders"):
        "Thêm một tên hợp lệ như Data vào danh sách thư mục bị cấm trong bộ quét — case đỏ (báo động "
        "giả trên cấu trúc đúng).",
    (_F, "Scan_AllowsATechnicalNameNestedInsideACapability"):
        "Sửa bộ quét để soi đệ quy toàn bộ cây thư mục thay vì chỉ cấp 1 — case đỏ, vì "
        "Features/HealthCheck/Services bị bắt oan. Tên tầng kỹ thuật nằm bên trong một nghiệp vụ là "
        "hợp lệ.",
})

# ---------------------------------------------------------------- COMMON: frontend dùng chung

_F = "frontend/apps/web/tests/shared/money.test.ts"
BREAK_HINTS.update({
    (_F, "formats %d as %s"):
        "Xem giá và tổng tiền hiển thị trên giao diện — nếu thấy '12.5' thay vì '$12.50', '48' thay vì "
        "'$48.00', hay '0' thay vì '$0.00', thì case đỏ. Tái hiện bằng cách bỏ bước định dạng tiền.",
    (_F, "groups thousands so a large total stays readable"):
        "Sửa giá một sản phẩm trong database lên mức hàng nghìn rồi xem giỏ — nếu tổng hiện "
        "'$1234.50' thay vì '$1,234.50' thì case đỏ.",
    (_F, "rounds to two decimal places"):
        "Đặt giá sản phẩm thành số có nhiều hơn hai chữ số thập phân (12.499) rồi xem giao diện — nếu "
        "hiện '$12.499' thì case đỏ, vì giá hiển thị sẽ không khớp với tổng tính từ nó.",
    (_F, "keeps the sign on a negative amount"):
        "Cho hệ thống hiển thị một số tiền âm — nếu dấu trừ bị mất và hiện '$5.00' thay vì '-$5.00' thì "
        "case đỏ.",
    (_F, "accepts the string form the contract also permits"):
        "Sửa BFF dưới local để trả giá dưới dạng chuỗi ('12.50') thay vì số — nếu giao diện hiện 'NaN' "
        "hoặc trang vỡ thì case đỏ. Hợp đồng API cho phép cả hai dạng nên giao diện phải chịu được cả hai.",
    (_F, "refuses a value that is not an amount at all"):
        "Sửa BFF để trả về một giá trị không phải số ở trường giá — nếu giao diện hiển thị nguyên chuỗi "
        "rác đó như một mức giá thì case đỏ; nó phải báo lỗi rõ ràng.",
})

_F = "frontend/apps/web/tests/accessibility.test.tsx"
BREAK_HINTS.update({
    (_F, "offers a skip link as the first focusable element"):
        "Mở http://localhost:4173 và nhấn Tab lần đầu tiên — nếu phần tử nhận focus không phải liên kết "
        "'Skip to content' thì case đỏ. Người dùng bàn phím sẽ phải Tab qua toàn bộ thanh điều hướng ở "
        "mỗi trang.",
    (_F, "gives each screen its own document title"):
        "Chuyển giữa trang Products và Basket, nhìn tiêu đề tab trình duyệt — nếu không đổi theo màn "
        "hình thì case đỏ, vì mọi màn hình sẽ tự xưng cùng một tên với trình đọc màn hình.",
    (_F, "moves focus to the new screen after navigating"):
        "Dùng bàn phím bấm vào liên kết Basket rồi nhấn Tab tiếp — nếu focus vẫn kẹt ở thanh điều hướng "
        "trong khi nội dung bên dưới đã đổi thì case đỏ.",
    (_F, "does not steal focus on first render"):
        "Mở trang lần đầu và chưa bấm gì — nếu vùng nội dung chính đã tự chiếm focus thì case đỏ, vì nó "
        "cướp mất điểm bắt đầu của người dùng bàn phím.",
})

_F = "frontend/apps/web/tests/app.test.tsx"
BREAK_HINTS.update({
    (_F, "renders the landing route inside the shell"):
        "Mở http://localhost:4173 — nếu trang trắng hoặc không có tiêu đề Products thì case đỏ. Đây là "
        "case cơ bản nhất: vỏ ứng dụng dựng được và route mặc định hiển thị được.",
    (_F, "exposes navigation to the shopper"):
        "Mở trang chủ và tìm thanh điều hướng — nếu thiếu vùng điều hướng chính hoặc thiếu liên kết "
        "Basket thì case đỏ, khách không còn đường sang trang giỏ hàng.",
})

# ==========================================================================================
# DỰNG FILE EXCEL
# ==========================================================================================

FEATURES = ["Basket", "Gateway", "Order", "Party", "Product", "Common"]

ID_PREFIX = {
    "Basket": "TC-BSK",
    "Gateway": "TC-GTW",
    "Order": "TC-ORD",
    "Party": "TC-PTY",
    "Product": "TC-PRD",
    "Common": "TC-CMN",
}

HEADERS = [
    "Mã TC (ID)",
    "Tính năng chính (Master Feature)",
    "Nhóm chức năng (Sub-module)",
    "Tầng (Layer)",
    "Loại test (Test Type)",
    "Tên test gốc (Test Name)",
    "File nguồn (Source File)",
    "Trạng thái (Status)",
    "Điều kiện (Given)",
    "Hành động (When)",
    "Kết quả mong đợi (Then)",
    "Cách làm cho test fail (QA)",
]

COLUMN_WIDTHS = [14, 22, 34, 12, 14, 56, 62, 14, 50, 50, 50, 72]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical="top", wrap_text=True)
TOP_PLAIN = Alignment(vertical="top")


def commit_hash() -> str:
    """Hash của HEAD, để báo cáo tự nói rõ nó chụp lại phiên bản nào."""
    try:
        return subprocess.run(
            ["git", "rev-parse", "HEAD"],
            capture_output=True,
            text=True,
            check=True,
        ).stdout.strip()
    except (subprocess.CalledProcessError, FileNotFoundError):
        return "unknown"


def assign_ids(rows: list[dict[str, str]]) -> None:
    counters: Counter[str] = Counter()
    for row in rows:
        feature = row["feature"]
        counters[feature] += 1
        row["id"] = f"{ID_PREFIX[feature]}-{counters[feature]:03d}"


def write_feature_sheet(workbook: Workbook, feature: str, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(feature)
    sheet.append(HEADERS)

    for row in rows:
        sheet.append([
            row["id"],
            row["feature"],
            row["sub"],
            row["layer"],
            row["type"],
            row["name"],
            row["file"],
            "Passed",
            row["given"],
            row["when"],
            row["then"],
            BREAK_HINTS[(row["file"], row["name"])],
        ])

    style_sheet(sheet, len(rows))


def style_sheet(sheet, data_row_count: int) -> None:
    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER

    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{data_row_count + 1}"

    # Sub-module, tên test, đường dẫn file, Given/When/Then và cột cách làm fail đều xuống dòng
    # trong ô; các cột còn lại chỉ căn lề trên.
    wrapped = {3, 6, 7, 9, 10, 11, 12}
    for row in sheet.iter_rows(min_row=2, max_row=data_row_count + 1):
        for cell in row:
            cell.alignment = TOP_WRAP if cell.column in wrapped else TOP_PLAIN
            cell.border = BORDER


def write_summary_sheet(workbook: Workbook, rows: list[dict[str, str]], generated_on: date) -> None:
    sheet = workbook.create_sheet("Summary", 0)

    sheet["A1"] = "Tổng hợp test case đã pass theo master feature"
    sheet["A1"].font = TITLE_FONT
    sheet["A3"] = "Ngày sinh báo cáo:"
    sheet["B3"] = generated_on.isoformat()
    sheet["A4"] = "Commit (HEAD):"
    sheet["B4"] = commit_hash()
    sheet["A5"] = "Nhánh:"
    sheet["B5"] = "master"
    sheet["A6"] = "Ghi chú:"
    sheet["B6"] = (
        "Tổng hợp từ source code test tại HEAD; không chạy lại test. "
        "Một [Theory]/it.each nhiều bộ dữ liệu được tính là một test case. "
        "Cột cuối mô tả cách làm cho case đó fail, viết theo góc nhìn QA: thao tác trên giao diện, "
        "gọi API, bật/tắt container, hoặc sửa mã nguồn dưới máy local để tái hiện."
    )

    for row in range(3, 7):
        sheet.cell(row=row, column=1).font = Font(bold=True)

    types = ["Unit", "Integration", "E2E"]
    layers = ["Backend", "Frontend"]

    header_row = 8
    headers = ["Tính năng chính (Master Feature)", "Tầng (Layer)", *types, "Tổng cộng"]
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER

    counts = Counter((row["feature"], row["layer"], row["type"]) for row in rows)

    current = header_row + 1
    for feature in FEATURES:
        for layer in layers:
            per_type = [counts[(feature, layer, test_type)] for test_type in types]
            if sum(per_type) == 0:
                continue

            values = [feature, layer, *per_type, sum(per_type)]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=current, column=column, value=value)
                cell.border = BORDER
                cell.alignment = TOP_PLAIN
            current += 1

    totals = [sum(1 for row in rows if row["type"] == test_type) for test_type in types]
    for column, value in enumerate(["TỔNG CỘNG", "", *totals, len(rows)], start=1):
        cell = sheet.cell(row=current, column=column, value=value)
        cell.font = Font(bold=True)
        cell.border = BORDER

    for column, width in enumerate([34, 14, 14, 16, 10, 14], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    sheet.column_dimensions["B"].width = 20
    sheet.freeze_panes = f"A{header_row + 1}"


def main() -> None:
    # Console Windows mac dinh la cp1252 va khong in duoc tieng Viet.
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    repository_root = Path(__file__).resolve().parent.parent
    generated_on = date.today()
    output_path = repository_root / "docs" / f"test-cases-{generated_on.isoformat()}.xlsx"

    unknown = {row["feature"] for row in ROWS} - set(FEATURES)
    if unknown:
        raise SystemExit(f"Master feature không hợp lệ: {sorted(unknown)}")

    # Thiếu một mô tả nào cũng phải dừng ngay, thay vì lặng lẽ ghi ra một ô trống.
    missing = [(row["file"], row["name"]) for row in ROWS
               if (row["file"], row["name"]) not in BREAK_HINTS]
    if missing:
        raise SystemExit(f"Thiếu mô tả cách làm fail cho {len(missing)} case: {missing[:3]}")

    assign_ids(ROWS)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for feature in FEATURES:
        write_feature_sheet(workbook, feature, [row for row in ROWS if row["feature"] == feature])

    write_summary_sheet(workbook, ROWS, generated_on)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"Đã ghi: {output_path}")
    print(f"Tổng số test case: {len(ROWS)}")
    for feature in FEATURES:
        feature_rows = [row for row in ROWS if row["feature"] == feature]
        backend = sum(1 for row in feature_rows if row["layer"] == "Backend")
        frontend = len(feature_rows) - backend
        print(f"  {feature:<8} {len(feature_rows):>4}  (Backend {backend}, Frontend {frontend})")


if __name__ == "__main__":
    main()
