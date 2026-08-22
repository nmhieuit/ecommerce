#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Tong hop toan bo specs/*/spec.md thanh MOT file Excel tieng Viet cho software manager.

Truc cha la tung muc Functional Requirement; cac dong con chi gom Acceptance Scenario /
Measurable Outcome co THAM CHIEU TUONG MINH (nhac dich danh ma FR-xxx / SC-xxx).

Noi dung tieng Viet lay tu file dich docs/spec-summary-vi.json (khoa theo slug feature + ma muc).
Script khong tu dich va khong tu suy luan lien ket.

Cach dung:
    python scripts/export-spec-summary.py                # sinh docs/tong-hop-yeu-cau-<YYYY-MM-DD>.xlsx
    python scripts/export-spec-summary.py --dump a.json  # xuat du lieu trich xuat (ban goc) ra JSON
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
SPECS_DIR = REPO_ROOT / "specs"
DOCS_DIR = REPO_ROOT / "docs"
TRANSLATION_DIR = DOCS_DIR / "spec-summary-vi"

KHONG_CO = "Không có trong spec"
CHUA_LIEN_KET = "Chưa có liên kết tường minh trong spec"
CHUA_DICH = "[CHƯA DỊCH] "

RE_TITLE = re.compile(r"^#\s+Feature Specification:\s*(.+?)\s*$")
RE_USER_STORY = re.compile(r"^###\s+User Story\s+(\d+)\s*[-–—]\s*(.+?)\s*\(Priority:\s*(P\d)\)\s*$")
RE_HEADING = re.compile(r"^(#{2,4})\s+(.*?)\s*$")
RE_BULLET = re.compile(r"^-\s+\*\*(FR-\d+[a-z]?|SC-\d+[a-z]?)\*\*:\s*(.*)$")
RE_NUMBERED = re.compile(r"^(\d+)\.\s+(.*)$")
RE_FR_REF = re.compile(r"\bFR-\d{3}[a-z]?\b")
RE_SC_REF = re.compile(r"\bSC-\d{3}[a-z]?\b")


def clean(text: str) -> str:
    """Bo cu phap markdown gay nhieu, giu nguyen chu."""
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"~~(.+?)~~", r"\1", text)
    text = re.sub(r"`([^`]+)`", r"\1", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


@dataclass
class Item:
    code: str
    text: str
    line: int
    kind: str  # "FR" | "SC" | "AS"
    story: int = 0
    story_title: str = ""
    priority: str = ""
    refs_fr: list = field(default_factory=list)
    refs_sc: list = field(default_factory=list)


@dataclass
class Feature:
    slug: str
    sheet: str
    path: str
    title: str
    stories: dict
    frs: list
    scs: list
    scenarios: list
    co_refs: list = field(default_factory=list)   # (ma FR, ma SC, so dong) neu cung mot cum ngoac


def _flush(buf, out, kind, story=0, story_title="", priority=""):
    if not buf:
        return
    code, parts, line = buf
    text = clean(" ".join(parts))
    item = Item(code=code, text=text, line=line, kind=kind, story=story,
                story_title=story_title, priority=priority)
    item.refs_fr = sorted(set(RE_FR_REF.findall(text)) - {code})
    item.refs_sc = sorted(set(RE_SC_REF.findall(text)) - {code})
    out.append(item)


def parse_spec(path: Path) -> Feature:
    lines = path.read_text(encoding="utf-8").splitlines()
    slug = path.parent.name
    title = ""
    frs: list = []
    scs: list = []
    scenarios: list = []
    stories: dict = {}

    section = ""          # "FR" | "SC" | ""
    story_no = 0
    story_title = ""
    priority = ""
    in_scenarios = False
    buf = None
    rel = path.relative_to(REPO_ROOT).as_posix()

    def target_list():
        return frs if section == "FR" else scs if section == "SC" else scenarios

    for idx, raw in enumerate(lines, start=1):
        line = raw.rstrip()

        m = RE_TITLE.match(line)
        if m and not title:
            title = clean(m.group(1))
            continue

        m = RE_USER_STORY.match(line)
        if m:
            _flush(buf, scenarios, "AS", story_no, story_title, priority)
            buf = None
            story_no = int(m.group(1))
            story_title = clean(m.group(2))
            priority = m.group(3)
            stories[story_no] = {"title": story_title, "priority": priority, "line": idx}
            section = ""
            in_scenarios = False
            continue

        m = RE_HEADING.match(line)
        if m:
            _flush(buf, target_list(), section or "AS", story_no, story_title, priority)
            buf = None
            heading = m.group(2).lower()
            in_scenarios = False
            if heading.startswith("functional requirements"):
                section = "FR"
            elif heading.startswith("measurable outcomes"):
                section = "SC"
            else:
                section = ""
            continue

        if line.strip().startswith("**Acceptance Scenarios**"):
            _flush(buf, scenarios, "AS", story_no, story_title, priority)
            buf = None
            in_scenarios = True
            continue

        if line.strip() == "---":
            _flush(buf, scenarios, "AS", story_no, story_title, priority)
            buf = None
            in_scenarios = False
            continue

        if section in ("FR", "SC"):
            m = RE_BULLET.match(line)
            if m:
                _flush(buf, target_list(), section, story_no, story_title, priority)
                buf = (m.group(1), [m.group(2)], idx)
                continue
            if buf and line.startswith((" ", "\t")) and line.strip():
                buf[1].append(line.strip())
                continue
            if not line.strip():
                _flush(buf, target_list(), section, story_no, story_title, priority)
                buf = None
            continue

        if in_scenarios:
            m = RE_NUMBERED.match(line)
            if m:
                _flush(buf, scenarios, "AS", story_no, story_title, priority)
                buf = ("US%d-KB%s" % (story_no, m.group(1)), [m.group(2)], idx)
                continue
            if buf and line.startswith((" ", "\t")) and line.strip():
                buf[1].append(line.strip())
                continue
            if not line.strip():
                _flush(buf, scenarios, "AS", story_no, story_title, priority)
                buf = None
            continue

    _flush(buf, target_list(), section or "AS", story_no, story_title, priority)

    # Tham chieu tuong minh dang "(FR-012, SC-006)" / "(... FR-025 / SC-011)":
    # chi nhan khi ma FR va ma SC nam trong CUNG MOT cum ngoac don, khong doc ra ngoai cum.
    co_refs = []
    for idx, raw in enumerate(lines, start=1):
        for group in re.findall(r"\(([^()]*)\)", raw):
            fr_hits = RE_FR_REF.findall(group)
            sc_hits = RE_SC_REF.findall(group)
            for fr_code in sorted(set(fr_hits)):
                for sc_code in sorted(set(sc_hits)):
                    co_refs.append((fr_code, sc_code, idx))

    return Feature(slug=slug, sheet=slug.split("-")[0], path=rel, title=title,
                   stories=stories, frs=frs, scs=scs, scenarios=scenarios, co_refs=co_refs)


def link(feature: Feature):
    """Chi noi khi co tham chieu tuong minh.

    Tra ve (map ma FR -> [(Item con, nguon dan chung)], AS mo coi, SC mo coi).
    """
    fr_codes = set(fr.code for fr in feature.frs)
    sc_by_code = dict((s.code, s) for s in feature.scs)
    children = dict((fr.code, []) for fr in feature.frs)
    seen = set()
    linked_as = set()
    linked_sc = set()

    def add(fr_code, item, source):
        key = (fr_code, item.code)
        if fr_code not in children or key in seen:
            return
        seen.add(key)
        children[fr_code].append((item, source))
        (linked_as if item.kind == "AS" else linked_sc).add(item.code)

    def where(item):
        return "%s:%d" % (feature.path, item.line)

    for scenario in feature.scenarios:                # Acceptance Scenario nhac dich danh FR-xxx
        for ref in scenario.refs_fr:
            add(ref, scenario, where(scenario))

    for outcome in feature.scs:                       # Measurable Outcome nhac dich danh FR-xxx
        for ref in outcome.refs_fr:
            add(ref, outcome, where(outcome))

    for fr in feature.frs:                            # FR nhac dich danh SC-xxx
        for ref in fr.refs_sc:
            target = sc_by_code.get(ref)
            if target is not None:
                add(fr.code, target, where(target))

    for fr_code, sc_code, line_no in feature.co_refs:  # cap ma nam cung mot cum ngoac
        target = sc_by_code.get(sc_code)
        if fr_code in fr_codes and target is not None:
            add(fr_code, target,
                "%s (liên kết nêu tại %s:%d)" % (where(target), feature.path, line_no))

    orphan_as = [s for s in feature.scenarios if s.code not in linked_as]
    orphan_sc = [s for s in feature.scs if s.code not in linked_sc]
    return children, orphan_as, orphan_sc


def load_translations() -> dict:
    """Gom cac file dich docs/spec-summary-vi/<slug>.json thanh mot dict."""
    out = {}
    if TRANSLATION_DIR.exists():
        for path in sorted(TRANSLATION_DIR.glob("*.json")):
            out[path.stem] = json.loads(path.read_text(encoding="utf-8"))
    return out


def build_workbook(features, translations, out_path: Path) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEAD_FILL = PatternFill("solid", fgColor="FF1F4E79")
    HEAD_FONT = Font(bold=True, color="FFFFFFFF", size=11)
    FR_FILL = PatternFill("solid", fgColor="FFDCE6F1")
    TITLE_FONT = Font(bold=True, size=13)
    BLOCK_FONT = Font(bold=True, size=11, color="FF1F4E79")
    WRAP = Alignment(vertical="top", wrap_text=True)
    WRAP_IND = Alignment(vertical="top", wrap_text=True, indent=2)
    CENTER = Alignment(vertical="top", horizontal="center")
    HEAD_ALIGN = Alignment(vertical="center", wrap_text=True, horizontal="center")

    missing = []

    def vi(feature, code, fallback):
        entry = translations.get(feature.slug, {}).get("items", {}).get(code)
        if entry:
            return entry
        missing.append("%s/%s" % (feature.slug, code))
        return CHUA_DICH + fallback

    def vi_title(feature):
        entry = translations.get(feature.slug, {}).get("title")
        if entry:
            return entry
        missing.append("%s/title" % feature.slug)
        return CHUA_DICH + feature.title

    wb = Workbook()
    overview = wb.active
    overview.title = "Tổng quan"
    ov_headers = ["Mã feature", "Tên feature", "Số yêu cầu chức năng", "Số kịch bản chấp nhận",
                  "Số kết quả đo lường được", "Số FR chưa có liên kết tường minh",
                  "Tên sheet chi tiết"]
    overview.append(ov_headers)
    for cell in overview[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = HEAD_ALIGN

    stats = []
    for feature in features:
        children, orphan_as, orphan_sc = link(feature)
        no_link = sum(1 for fr in feature.frs if not children[fr.code])
        stats.append((feature, children, orphan_as, orphan_sc, no_link))
        overview.append([feature.slug, vi_title(feature), len(feature.frs), len(feature.scenarios),
                         len(feature.scs), no_link, feature.sheet])

    for row in overview.iter_rows(min_row=2, max_row=overview.max_row, max_col=len(ov_headers)):
        for cell in row:
            cell.alignment = CENTER if cell.column > 2 else WRAP

    note_row = overview.max_row + 2
    overview.cell(row=note_row, column=1, value="Ghi chú phương pháp tổng hợp").font = BLOCK_FONT
    notes = [
        "Toàn bộ nội dung được trích từ specs/*/spec.md rồi dịch sang tiếng Việt; không suy luận, không bổ sung.",
        "Một FR chỉ được nối với Kịch bản chấp nhận / Kết quả đo lường được khi trong spec có tham chiếu tường minh (nhắc đích danh mã FR-xxx hoặc SC-xxx).",
        "FR không có tham chiếu tường minh nào được đánh dấu \"" + CHUA_LIEN_KET + "\"; các mục còn lại nằm ở hai khối cuối mỗi sheet.",
    ]
    for offset, text in enumerate(notes, start=1):
        overview.cell(row=note_row + offset, column=1, value=text).alignment = WRAP

    for i, w in enumerate([30, 62, 14, 14, 16, 16, 14], start=1):
        overview.column_dimensions[get_column_letter(i)].width = w
    overview.freeze_panes = "A2"
    overview.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(ov_headers)), 1 + len(features))

    headers = ["Cấp", "Mã", "Loại", "Nội dung", "Nguồn"]
    for feature, children, orphan_as, orphan_sc, _ in stats:
        ws = wb.create_sheet(feature.sheet)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        title_cell = ws.cell(row=1, column=1, value="%s — %s" % (feature.slug, vi_title(feature)))
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.append(headers)
        for cell in ws[2]:
            cell.fill = HEAD_FILL
            cell.font = HEAD_FONT
            cell.alignment = HEAD_ALIGN

        for fr in feature.frs:
            ws.append([1, fr.code, "Yêu cầu chức năng", vi(feature, fr.code, fr.text),
                       "%s:%d" % (feature.path, fr.line)])
            for cell in ws[ws.max_row]:
                cell.fill = FR_FILL
                cell.font = Font(bold=True)
                cell.alignment = WRAP
            ws.cell(row=ws.max_row, column=1).alignment = CENTER

            kids = children[fr.code]
            if not kids:
                ws.append([2, KHONG_CO, KHONG_CO, CHUA_LIEN_KET, KHONG_CO])
                for cell in ws[ws.max_row]:
                    cell.alignment = WRAP
                    cell.font = Font(italic=True, color="FF808080")
                ws.cell(row=ws.max_row, column=4).alignment = WRAP_IND
                ws.cell(row=ws.max_row, column=1).alignment = CENTER
                continue

            for kid, source in kids:
                loai = "Kịch bản chấp nhận" if kid.kind == "AS" else "Kết quả đo lường được"
                ws.append([2, kid.code, loai, vi(feature, kid.code, kid.text), source])
                for cell in ws[ws.max_row]:
                    cell.alignment = WRAP
                ws.cell(row=ws.max_row, column=4).alignment = WRAP_IND
                ws.cell(row=ws.max_row, column=1).alignment = CENTER

        last_hierarchy_row = ws.max_row
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = "A2:E%d" % last_hierarchy_row

        def block(title, items, loai):
            row = ws.max_row + 2
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=1, value=title).font = BLOCK_FONT
            if not items:
                ws.append(["", KHONG_CO, loai,
                           "Không có mục nào — mọi mục đều đã gắn với ít nhất một FR.", KHONG_CO])
                for cell in ws[ws.max_row]:
                    cell.alignment = WRAP
                return
            for item in items:
                ws.append(["", item.code, loai, vi(feature, item.code, item.text),
                           "%s:%d" % (feature.path, item.line)])
                for cell in ws[ws.max_row]:
                    cell.alignment = WRAP

        block("Kịch bản chấp nhận chưa gắn với FR nào", orphan_as, "Kịch bản chấp nhận")
        block("Kết quả đo lường được chưa gắn với FR nào", orphan_sc, "Kết quả đo lường được")

        for i, w in enumerate([6, 14, 24, 110, 44], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 24

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"missing": missing, "stats": stats}


def main() -> int:
    for stream in (sys.stdout, sys.stderr):   # console Windows mac dinh cp1252 -> ep UTF-8
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--dump", help="Xuat du lieu trich xuat (ban goc) ra file JSON roi dung")
    ap.add_argument("--out", help="Duong dan file Excel dau ra")
    args = ap.parse_args()

    spec_files = sorted(SPECS_DIR.glob("*/spec.md"))
    if not spec_files:
        print("Không tìm thấy file spec.md nào trong specs/", file=sys.stderr)
        return 1
    features = [parse_spec(p) for p in spec_files]

    if args.dump:
        payload = {}
        for f in features:
            children, orphan_as, orphan_sc = link(f)
            payload[f.slug] = {
                "title": f.title,
                "stories": f.stories,
                "items": dict((i.code, i.text) for i in f.frs + f.scenarios + f.scs),
                "lines": dict((i.code, i.line) for i in f.frs + f.scenarios + f.scs),
                "links": dict((code, [k.code for k, _ in kids]) for code, kids in children.items()),
                "orphan_as": [i.code for i in orphan_as],
                "orphan_sc": [i.code for i in orphan_sc],
            }
        Path(args.dump).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print("Đã xuất dữ liệu trích xuất ra %s" % args.dump)
        return 0

    out_path = Path(args.out) if args.out else DOCS_DIR / ("tong-hop-yeu-cau-%s.xlsx" % date.today().isoformat())
    result = build_workbook(features, load_translations(), out_path)

    total_fr = sum(len(f.frs) for f in features)
    print("Đã tạo: %s" % out_path.relative_to(REPO_ROOT).as_posix())
    print("Số feature đã xử lý: %d" % len(features))
    print("Tổng số Functional Requirement: %d" % total_fr)
    print("Danh sách FR không có liên kết tường minh:")
    for feature, children, _, _, _ in result["stats"]:
        codes = [fr.code for fr in feature.frs if not children[fr.code]]
        print("  - %s (%d/%d): %s" % (feature.slug, len(codes), len(feature.frs),
                                      ", ".join(codes) if codes else "không có"))
    if result["missing"]:
        print("CẢNH BÁO: %d mục chưa có bản dịch tiếng Việt:" % len(result["missing"]), file=sys.stderr)
        for key in result["missing"]:
            print("  - %s" % key, file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
